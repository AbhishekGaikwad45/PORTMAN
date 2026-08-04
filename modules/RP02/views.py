import re
from functools import wraps
from flask import render_template, request, jsonify, session, redirect, url_for, Response

from database import get_user_permissions
from . import bp
from . import model


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def _can_upload():
    if session.get('is_admin'):
        return True
    perms = get_user_permissions(session['user_id'], 'RP02')
    return bool(perms['can_add'] or perms['can_edit'])


def upload_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if not _can_upload():
            return jsonify({'error': 'No permission to modify bill master data'}), 403
        return f(*args, **kwargs)
    return decorated


@bp.route('/module/RP02/')
@login_required
def index():
    return render_template('rp02.html', username=session.get('username'),
                           can_upload=_can_upload())


@bp.route('/module/RP02/bill-master/')
@login_required
def bill_master_index():
    return render_template('bill_master.html', username=session.get('username'),
                           status=model.get_status(), can_upload=_can_upload())


@bp.route('/module/RP02/bill-master-report/')
@login_required
def bill_master_report_index():
    return render_template('bill_master_report.html', username=session.get('username'))


@bp.route('/api/module/RP02/bill-master-report/data')
@login_required
def bill_master_report_data():
    return jsonify({'data': model.get_bill_master_report()})


@bp.route('/api/module/RP02/bill-master/template')
@login_required
def bill_master_template():
    return Response(
        model.build_template_csv(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename="RP02_bill_master_template.csv"'},
    )


@bp.route('/api/module/RP02/bill-master/preview', methods=['POST'])
@upload_required
def bill_master_preview():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file provided'}), 400
    rows, errors = model.parse_upload(f)
    customers = model.get_customer_master()
    recon = model.reconcile(rows, customers)
    years = sorted({r['financial_year'] for r in rows if r.get('financial_year')})
    opts = {col: customers for col, info in recon.items() if info['unknown']}
    return jsonify({'total_rows': len(rows), 'format_errors': errors,
                    'years': years,
                    'reconciliation': recon,
                    'master_options': opts,
                    'addable_columns': ['customer_name']})


@bp.route('/api/module/RP02/bill-master/apply', methods=['POST'])
@upload_required
def bill_master_apply():
    import json as _json
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file provided'}), 400
    try:
        resolutions = _json.loads(request.form.get('resolutions') or '{}')
    except (ValueError, TypeError):
        resolutions = {}

    rows, errors = model.parse_upload(f)
    if errors:
        return jsonify({'error': 'Fix format errors before applying',
                        'format_errors': errors}), 400

    # 1) Add-to-master actions (customer master only).
    added, add_errors = [], []
    for value, res in (resolutions.get('customer_name') or {}).items():
        if isinstance(res, dict) and res.get('action') == 'add':
            try:
                if model.add_customer(value):
                    added.append({'column': 'customer_name', 'value': value})
            except Exception as e:  # noqa: BLE001 — surface, don't abort
                add_errors.append({'column': 'customer_name', 'value': value, 'error': str(e)})

    # 2) Replace actions rewrite the parsed rows, then per-FY replace insert.
    rows = model.apply_resolutions(rows, resolutions)
    inserted, years = model.replace_years(rows, session.get('user_id'))
    return jsonify({'inserted': inserted, 'years': years,
                    'added_to_master': added, 'add_errors': add_errors})


@bp.route('/api/module/RP02/bill-master/rows')
@login_required
def bill_master_rows():
    import json as _json
    try:
        page = int(request.args.get('page', 1))
        size = int(request.args.get('size', 50))
    except (TypeError, ValueError):
        page, size = 1, 50
    try:
        filters = _json.loads(request.args.get('colfilters') or '[]')
        if not isinstance(filters, list):
            filters = []
    except (ValueError, TypeError):
        filters = []
    rows, total = model.get_rows(page, size, filters)
    return jsonify({'data': rows, 'last_page': max(1, (total + size - 1) // size), 'total': total})


@bp.route('/api/module/RP02/bill-master/row/update', methods=['POST'])
@upload_required
def bill_master_row_update():
    data = request.json or {}
    if not data.get('id'):
        return jsonify({'error': 'Missing id'}), 400
    res = model.update_row(data['id'], data)
    if res.get('error'):
        return jsonify(res), 400
    return jsonify(res)


@bp.route('/api/module/RP02/bill-master/row/delete', methods=['POST'])
@upload_required
def bill_master_row_delete():
    data = request.json or {}
    if not data.get('id'):
        return jsonify({'error': 'Missing id'}), 400
    model.delete_row(data['id'])
    return jsonify({'success': True})


# ── Billing Pipeline dashboard ───────────────────────────────────────────────

@bp.route('/module/RP02/billing-dashboard/')
@login_required
def billing_dashboard_index():
    return render_template('billing_dashboard.html', username=session.get('username'))


@bp.route('/api/module/RP02/billing-dashboard/data')
@login_required
def billing_dashboard_data():
    return jsonify(model.get_billing_dashboard())


# Columns for the Excel dump — quantity and dates only, no rates or amounts.
_EXPORT_COLUMNS = [
    ('Status', 'status'),
    ('Vessel / MBC', 'vessel_name'),
    ('Material PO', 'material_po'),
    ('Customer', 'customer_name'),
    ('Type of Cargo', 'cargo_type'),
    ('Cargo', 'cargo_name'),
    ('Pending Qty (MT)', 'bl_qty'),
    ('Load Port', 'load_port'),
    ('MV/MBC', 'mv_mbc'),
    ('Discharge Commence', 'discharge_commence'),
    ('Discharge Completed', 'discharge_completed'),
    ('Age (days)', 'age_days'),
    ('Age Bucket', 'age_bucket'),
    ('Doc Status', 'doc_status'),
]


@bp.route('/api/module/RP02/billing-dashboard/export')
@login_required
def billing_dashboard_export():
    """Excel dump of the pending lines, plus a summary sheet of the same numbers
    the dashboard shows. Deliberately carries no rates or amounts.

    ?dim=&val=&scope= exports one breakdown bar's drilldown instead of the whole
    pipeline, filtered by the same model helper the modal uses."""
    import io
    from datetime import datetime as _dt
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    dim = request.args.get('dim')
    val = request.args.get('val')
    scope = request.args.get('scope', 'all')

    data = model.get_billing_dashboard()
    if dim or scope != 'all':
        rows, drill_desc = model.filter_pending(dim, val, scope, rows=data['rows'])
        data = {**data, 'rows': rows}
    else:
        drill_desc = None

    thin = Side(style='thin', color='000000')
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    cell_font = Font(name='Calibri', size=10)
    blocked_font = Font(name='Calibri', size=10, color='8A5A00')
    lft = Alignment(horizontal='left', vertical='center')
    rgt = Alignment(horizontal='right', vertical='center')
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb = openpyxl.Workbook()

    # Sheet 1: the lines
    ws = wb.active
    ws.title = 'Pending Lines'
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 22
    for i, (label, _k) in enumerate(_EXPORT_COLUMNS, 1):
        c = ws.cell(1, i, label)
        c.font = hdr_font; c.fill = hdr_fill; c.border = bdr; c.alignment = ctr
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(label) + 3)

    numeric = {'bl_qty', 'age_days'}
    for row_i, r in enumerate(data['rows'], 2):
        blocked = r.get('status') != 'Ready to bill'
        # not `val` — that name holds the drilldown query param used below
        for col_i, (_label, key) in enumerate(_EXPORT_COLUMNS, 1):
            cell_val = r.get(key)
            c = ws.cell(row_i, col_i, '' if cell_val is None else cell_val)
            c.border = bdr
            c.font = blocked_font if blocked else cell_font
            c.alignment = rgt if key in numeric else lft
            if key == 'bl_qty' and cell_val is not None:
                c.number_format = '#,##0.00'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(_EXPORT_COLUMNS))}{max(len(data["rows"]) + 1, 1)}'

    # Sheet 2: the dashboard's own numbers, so the file stands alone
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 34
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 14
    k = data['kpi']
    blocks = [
        ('Billing Pipeline — generated ' + data['generated_at'], None, None),
        ('Quantities are cargo tonnage (MT) and line counts only. No rates or amounts.', None, None),
    ]
    if drill_desc:
        # a drilldown export says what it is a slice of, so the file is not
        # mistaken for the whole pipeline once it leaves this screen
        blocks += [
            (None, None, None),
            ('Drilldown: ' + drill_desc, None, None),
            ('Lines in this export', len(data['rows']),
             round(sum(r['bl_qty'] or 0 for r in data['rows']), 2)),
        ]
    blocks += [
        (None, None, None),
        ('Whole pipeline', 'Lines', 'Qty (MT)'),
        ('Ready to bill', k['ready_lines'], k['ready_qty']),
        ('Blocked upstream', k['blocked_lines'], k['blocked_qty']),
        ('Oldest waiting (days)', k['oldest_days'], None),
        ('Customers waiting', k['customers'], None),
        ('Billed to date (bill master lines)', k['billed_lines'], None),
        (f"Billed in {k['month_label']}", k['billed_month'], None),
    ]
    for title, group in (('Ready to bill by age', data['ageing']),
                         ('Ready to bill by customer', data['by_customer']),
                         ('Ready to bill by cargo type', data['by_cargo']),
                         ('Blocked by document status', data['blocked_by_status'])):
        blocks += [(None, None, None), (title, 'Lines', 'Qty (MT)')]
        blocks += [(g['label'], g['lines'], g['qty']) for g in group]

    for row_i, (a, b, c_) in enumerate(blocks, 1):
        is_head = b in ('Lines',) or (row_i == 1)
        ca = ws2.cell(row_i, 1, a if a is not None else '')
        ca.font = Font(name='Calibri', size=10, bold=bool(is_head))
        ca.alignment = lft
        # not `val` — that name holds the drilldown query param used below
        for col_i, cell_val in ((2, b), (3, c_)):
            cc = ws2.cell(row_i, col_i, '' if cell_val is None else cell_val)
            cc.font = Font(name='Calibri', size=10, bold=bool(is_head))
            cc.alignment = rgt
            if col_i == 3 and isinstance(cell_val, (int, float)):
                cc.number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug = ''
    if drill_desc:
        safe = re.sub(r'[^A-Za-z0-9]+', '_', (val or scope)).strip('_')[:40]
        slug = f'_{safe}' if safe else '_drilldown'
    fname = f"RP02_billing_pipeline{slug}_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )
