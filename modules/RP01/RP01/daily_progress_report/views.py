from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from flask import request, jsonify
from openpyxl.styles import PatternFill

from flask import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime, timedelta
from flask import (
    render_template,
    request,
    session,
    redirect,
    url_for,
    jsonify
)

from functools import wraps

from .. import bp
from database import get_db, get_cursor

_thin = Side(style='thin', color='C7CDD4')

_bdr = Border(
    left=_thin,
    right=_thin,
    top=_thin,
    bottom=_thin
)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

_ctr = Alignment(
    horizontal='center',
    vertical='center',
    wrap_text=True
)

_left = Alignment(
    horizontal='left',
    vertical='center',
    wrap_text=True
)

_TITLE_FILL = 'F6F8FB'
_HEADER_FILL = 'EEF4FB'
_BODY_FILL = 'FFFFFF'
_TOTAL_FILL = 'E9EEF5'
_TEXT = '2C3E50'


def _fill(hex_color):

    return PatternFill(
        'solid',
        fgColor=hex_color
    )


def _parse_flexible_dt(raw, out_fmt):
    """Parse a stored timestamp string that may or may not include
    seconds/microseconds, and format it for display.

    Several tables store timestamps as free text (not real timestamp
    columns), and different rows have ended up with different levels
    of precision ("...T07:35" vs "...T07:35:00" vs "...T07:35:00.123").
    A fixed-format strptime blows up on whichever variant it wasn't
    written for, which throws inside the endpoint, makes the whole
    API call return success:false, and the frontend then silently
    skips rendering that entire report section. Always use this
    helper instead of a bare strptime() on these text columns.
    """
    if not raw:
        return ''
    s = str(raw).replace('T', ' ').strip()
    for f in ('%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M'):
        try:
            return datetime.strptime(s, f).strftime(out_fmt)
        except ValueError:
            continue
    try:
        return datetime.strptime(s[:16], '%Y-%m-%d %H:%M').strftime(out_fmt)
    except ValueError:
        return s


def _font(
    bold=False,
    size=11,
    color='000000'
):

    return Font(
        name='Calibri',
        bold=bold,
        size=size,
        color=color
    )
def _cell(
    ws,
    row,
    col,
    value='',
    bold=False,
    fill_color='FFFFFF',
    align=_ctr,
    font_color='000000'
):

    cell = ws.cell(row, col, value)

    cell.font = _font(
        bold=bold,
        color=font_color
    )

    cell.fill = _fill(fill_color)

    cell.alignment = align

    cell.border = _bdr

    return cell
yellow_fill = PatternFill(
    start_color='FFFF00',
    end_color='FFFF00',
    fill_type='solid'
)

heading_font = Font(
    bold=True,
    size=14
)

sub_heading_font = Font(
    bold=True,
    size=11
)

# =========================================================
# LOGIN REQUIRED
# =========================================================

def login_required(f):

    @wraps(f)
    def decorated(*args, **kwargs):

        if 'user_id' not in session:
            return redirect(url_for('login'))

        return f(*args, **kwargs)

    return decorated


# =========================================================
# INDEX PAGE
# =========================================================

@bp.route(
    '/module/RP01/daily_progress_report/',
    endpoint='daily_progress_report_index'
)
@login_required
def daily_progress_report_index():

    return render_template(
        'daily_progress_report/daily_progress_report.html',
        username=session.get('username')
    )


# =========================================================
# DAILY PROGRESS REPORT API
# =========================================================

@bp.route(
    '/api/module/RP01/daily_progress_report',
    methods=['GET']
)

# Assumed already available in your project:
# from your_auth_module import login_required
# from your_db_module import get_db, get_cursor


@login_required
def daily_progress_report_data():

    report_date = request.args.get('report_date')

    print("\n========== DPR API START ==========")
    print("REPORT DATE:", report_date)

    if not report_date:
        return jsonify({
            "success": False,
            "message": "Report date required"
        })

    conn = get_db()
    cur = get_cursor(conn)

    try:
        # ─────────────────────────────────────────────────────────────
        # REPORT WINDOW:
        #   A "report date" covers a 24-hour operational window ending
        #   at 08:00 on the selected date (i.e. 08:00 previous day ->
        #   08:00 selected day). window_end / window_start define that
        #   boundary and are used (instead of the raw date string) for
        #   all timestamp comparisons below.
        #
        # COLUMN MAPPING:
        #   ldud_anchorage.anchored            : anchorage arrival time
        #   ldud_anchorage.discharge_started   : discharge start time
        #   ldud_anchorage.discharge_commenced : misnamed column — it
        #                                        actually stores the
        #                                        DISCHARGE COMPLETED
        #                                        datetime.
        #
        # FILTER LOGIC:
        #   - Vessel's discharge must have STARTED before window_end.
        #   - Vessel must either still be discharging (no completion
        #     time yet) OR have COMPLETED at/after window_start (i.e.
        #     completed inside this reporting window), OR have an open
        #     barge line (no completed_discharge_berth / cast_off_berth)
        #     that keeps it "active" for this report regardless of date.
        #
        # SEQUENCE:
        #   Vessels are ordered purely chronologically by discharge
        #   start time — NOT by completed/in-progress status. A vessel
        #   that started discharging earlier always ranks higher,
        #   whether or not it has since completed.
        # ─────────────────────────────────────────────────────────────

        report_dt = datetime.strptime(report_date, "%Y-%m-%d")

        window_end = datetime(
            report_dt.year,
            report_dt.month,
            report_dt.day,
            8, 0, 0
        )

        window_start = window_end - timedelta(hours=24)

        query = """

        SELECT
            lh.id,
            lh.vessel_name,
            vcd.cargo_name,
            vcd.bl_quantity,
            first_anchor.arrived_mfl,
            first_anchor.arrived_mbpt,
            first_anchor.discharge_started,
            last_anchor.discharge_completed AS discharge_commenced

        FROM ldud_header lh

        LEFT JOIN LATERAL (

            SELECT
                STRING_AGG(
                    cargo_name,
                    ' + '
                    ORDER BY cargo_name
                ) AS cargo_name,

                SUM(bl_qty) AS bl_quantity

            FROM (

                SELECT
                    TRIM(cargo_name) AS cargo_name,
                    SUM(COALESCE(bl_quantity, 0)) AS bl_qty
                FROM vcn_cargo_declaration
                WHERE vcn_id = lh.vcn_id
                GROUP BY TRIM(cargo_name)

            ) cargo

        ) vcd ON TRUE

        LEFT JOIN LATERAL (

            SELECT
                MIN(
                    CASE
                        WHEN anchorage_name ILIKE '%%PLA%%'
                        THEN anchored
                    END
                ) AS arrived_mfl,

                MIN(
                    CASE
                        WHEN anchorage_name NOT ILIKE '%%PLA%%'
                        THEN anchored
                    END
                ) AS arrived_mbpt,

                MIN(discharge_started) AS discharge_started

            FROM ldud_anchorage
            WHERE ldud_id = lh.id

        ) first_anchor ON TRUE

        LEFT JOIN LATERAL (

            SELECT
                CASE
                    WHEN EXISTS (
                        SELECT 1
                        FROM ldud_anchorage x
                        WHERE x.ldud_id = lh.id
                        AND x.discharge_started IS NOT NULL
                        AND x.discharge_commenced IS NULL
                    )
                    THEN NULL
                    ELSE MAX(discharge_commenced)
                END AS discharge_completed

            FROM ldud_anchorage
            WHERE ldud_id = lh.id

        ) last_anchor ON TRUE

        WHERE
            first_anchor.discharge_started IS NOT NULL
            AND first_anchor.discharge_started < %s

            AND (
                last_anchor.discharge_completed IS NULL
                OR last_anchor.discharge_completed >= %s
                OR EXISTS (
                    SELECT 1
                    FROM ldud_barge_lines b
                    WHERE b.ldud_id = lh.id
                    AND (
                        b.completed_discharge_berth IS NULL
                        OR b.cast_off_berth IS NULL
                    )
                )
            )

        ORDER BY
            UPPER(TRIM(lh.vessel_name)),
            lh.id

        """

        print("\nEXECUTING QUERY:")
        print(query)

        print("\nQUERY PARAMS:")
        print(window_end, window_start)

        cur.execute(
            query,
            (
                window_end,
                window_start
            )
        )

        rows = cur.fetchall()

        seq_map = {}

        for index, row in enumerate(rows, start=1):
            seq_map[row["id"]] = index

        print("SEQ MAP:", seq_map)

        print("\nTOTAL ROWS FETCHED:")
        print(len(rows))

        data = []

        for row in rows:

            print("\nROW:")
            print(row)

            vessel_name = row['vessel_name'] or ""
            cargo_name = row['cargo_name'] or ""
            bl_quantity = row['bl_quantity'] or ""

            arrived_mfl_raw = row['arrived_mfl']
            arrived_mbpt_raw = row['arrived_mbpt']

            # Actual discharge start
            discharge_start_raw = row['discharge_started']

            # Actual discharge completed (aliased from the misnamed column)
            discharge_completed_raw = row['discharge_commenced']

            arrived_mfl = (
                arrived_mfl_raw.strftime('%d-%m-%Y %H:%M')
                if arrived_mfl_raw else ""
            )

            arrived_mbpt = (
                arrived_mbpt_raw.strftime('%d-%m-%Y %H:%M')
                if arrived_mbpt_raw else ""
            )

            discharge_commenced = (
                discharge_start_raw.strftime('%d-%m-%Y %H:%M')
                if discharge_start_raw else ""
            )

            # Pre-Berthing Delay — fall back to the MbPT arrival time if
            # the vessel never anchored at MFL/PLA at all
            arrival_for_delay = arrived_mfl_raw or arrived_mbpt_raw

            pre_berthing_delay = ""
            if discharge_start_raw and arrival_for_delay:
                delay = discharge_start_raw - arrival_for_delay
                total_hours = round(delay.total_seconds() / 3600, 2)
                pre_berthing_delay = f"{total_hours} Hrs"

            # Discharge Completed — only show the timestamp if completion
            # actually falls inside this report's 08:00-to-08:00 window
            discharge_completed = ""
            if discharge_completed_raw and window_start <= discharge_completed_raw < window_end:
                discharge_completed = discharge_completed_raw.strftime('%d-%m-%Y %H:%M')

            print("VESSEL:", vessel_name)
            print("CARGO:", cargo_name)
            print("BL:", bl_quantity)
            print("ARRIVED MFL:", arrived_mfl)
            print("ARRIVED MBPT:", arrived_mbpt)
            print("DISCHARGE COMMENCED:", discharge_commenced)
            print("PRE BERTHING DELAY:", pre_berthing_delay)
            print("DISCHARGE COMPLETED:", discharge_completed)

            data.append({
                "vessel_seq": seq_map[row["id"]],
                "vessel_name": vessel_name,
                "cargo_name": cargo_name,
                "bl_quantity": bl_quantity,
                "arrived_mfl": arrived_mfl,
                "arrived_mbpt": arrived_mbpt,
                "discharge_commenced": discharge_commenced,
                "pre_berthing_delay": pre_berthing_delay,
                "discharge_completed": discharge_completed
            })

        print("\nFINAL JSON DATA:")
        print(data)

        print("\n========== DPR API END ==========\n")

        return jsonify({
            "success": True,
            "data": data
        })

    except Exception as e:
        print("\n========== DPR ERROR ==========")
        print(str(e))
        print("========== DPR ERROR END ==========\n")

        return jsonify({
            "success": False,
            "message": str(e)
        })

    finally:
        cur.close()
        conn.close()


@bp.route(
    '/api/module/RP01/monthly_cargo_report',
    methods=['GET']
)
@login_required
def monthly_cargo_report():

    report_date = request.args.get('report_date')

    print("\n========== MONTHLY REPORT START ==========")
    print("REPORT DATE:", report_date)

    if not report_date:
        return jsonify({"success": False, "message": "Report date required"})

    conn = get_db()
    cur = get_cursor(conn)

    try:
        report_dt = datetime.strptime(report_date, "%Y-%m-%d")

        window_end = datetime(
            report_dt.year, report_dt.month, report_dt.day, 8, 0, 0
        )
        window_start = window_end - timedelta(hours=24)
        cutoff_date = (report_dt - timedelta(days=1)).strftime('%Y-%m-%d')

        def format_hrs_to_hms(hours):
            """Convert decimal hours (e.g. 13.75) to 'H:MM:SS' string."""
            if hours is None:
                return "-"
            total_seconds = int(round(hours * 3600))
            h = total_seconds // 3600
            m = (total_seconds % 3600) // 60
            s = total_seconds % 60
            return f"{h}:{m:02d}:{s:02d}"

        query = """
        WITH vessel_order AS (
            SELECT
                lh.id,
                lh.vcn_id,
                lh.vessel_name,
                first_anchor.discharge_started,
                last_anchor.discharge_completed,

                CASE
                    WHEN last_anchor.discharge_completed IS NOT NULL
                    THEN 0
                    ELSE 1
                END AS sort_order,

                ROW_NUMBER() OVER (
                    ORDER BY
                        first_anchor.discharge_started,
                        lh.id
                ) AS vessel_seq

            FROM ldud_header lh

            LEFT JOIN LATERAL (
                SELECT MIN(discharge_started) AS discharge_started
                FROM ldud_anchorage
                WHERE ldud_id = lh.id
            ) first_anchor ON TRUE

            LEFT JOIN LATERAL (
                SELECT MAX(discharge_commenced) AS discharge_completed
                FROM ldud_anchorage
                WHERE ldud_id = lh.id
                AND discharge_commenced IS NOT NULL
            ) last_anchor ON TRUE

            WHERE
                first_anchor.discharge_started IS NOT NULL
                AND first_anchor.discharge_started < %s
                AND (
                    last_anchor.discharge_completed IS NULL
                    OR last_anchor.discharge_completed >= %s
                    OR EXISTS (
                        SELECT 1
                        FROM ldud_barge_lines b
                        WHERE b.ldud_id = lh.id
                        AND (
                            b.completed_discharge_berth IS NULL
                            OR b.cast_off_berth IS NULL
                        )
                    )
                )
        ),

        vessel_list AS (
            SELECT DISTINCT
                vo.id,
                vo.vcn_id,
                vo.vessel_name,
                vo.vessel_seq,
                vo.sort_order,
                vo.discharge_started,
                vo.discharge_completed,
                TRIM(vcd.cargo_name) AS cargo_name
            FROM vessel_order vo
            LEFT JOIN vcn_cargo_declaration vcd
                ON vcd.vcn_id = vo.vcn_id
        )

        SELECT
            vl.id,
            vl.vcn_id,
            vl.vessel_name,
            vl.cargo_name,
            vl.vessel_seq,
            vl.sort_order,
            vl.discharge_started,
            vl.discharge_completed,

            DATE(lco.start_time) AS cargo_date,
            TO_CHAR(DATE(lco.start_time), 'DD-MM-YYYY') AS day_label,

            COALESCE(SUM(lco.quantity),0) AS total_qty

        FROM vessel_list vl

        LEFT JOIN ldud_vessel_operations lco
            ON lco.ldud_id = vl.id
        AND LOWER(TRIM(lco.cargo_name)) = LOWER(TRIM(vl.cargo_name))
        AND DATE(lco.start_time)
            BETWEEN DATE_TRUNC('month', %s::date)::date
                AND %s::date
        AND (
                vl.discharge_completed IS NULL
                OR DATE(lco.start_time) <= DATE(vl.discharge_completed)
        )

        GROUP BY
            vl.id,
            vl.vcn_id,
            vl.vessel_name,
            vl.cargo_name,
            vl.vessel_seq,
            vl.sort_order,
            vl.discharge_started,
            vl.discharge_completed,
            DATE(lco.start_time)

        ORDER BY
            vl.vessel_seq,
            vl.cargo_name,
            DATE(lco.start_time)
        """

        cur.execute(
            query,
            (
                window_end,
                window_start,
                report_date,
                cutoff_date
            )
        )
        rows = cur.fetchall()

        # --- Combined BL + Discharged + Balance, computed directly in SQL per vcn_id ---
        vcn_ids = list({r["vcn_id"] for r in rows if r["vcn_id"]})

        bl_map = {}

        if vcn_ids:
            cur.execute("""
                SELECT
                    vcn_id,
                    COALESCE(SUM(bl_quantity), 0) AS total_bl
                FROM vcn_cargo_declaration
                WHERE vcn_id = ANY(%s)
                GROUP BY vcn_id
            """, (vcn_ids,))

            for r in cur.fetchall():
                bl_map[r["vcn_id"]] = float(r["total_bl"])

        # --- Fetch delay windows (Mother Vessel Agent / Force Majeure / MBP) ---
        # NOTE: ldud_delays has no FK to vessel_delay_types, so we join on name (TRIMmed).
        # type mapping:
        #   MOTHER VESSEL ACCOUNT -> "Mother Vessel Agent"
        #   FORCE MAJEURE         -> "Force Majeure"
        #   MbPT                  -> "MBP"
        ldud_ids = list({row["id"] for row in rows})

        delay_map = {}  # { ldud_id: [ (start_dt, end_dt), ... ] }

        if ldud_ids:
            cur.execute("""
                SELECT
                    d.ldud_id,
                    d.start_datetime,
                    d.end_datetime,
                    vdt.type AS delay_type
                FROM ldud_delays d
                JOIN vessel_delay_types vdt
                    ON TRIM(LOWER(vdt.name)) = TRIM(LOWER(d.delay_name))
                WHERE d.ldud_id = ANY(%s)
                AND vdt.type IN ('MOTHER VESSEL ACCOUNT', 'FORCE MAJEURE', 'MbPT')
                AND d.start_datetime IS NOT NULL
                AND d.end_datetime IS NOT NULL
            """, (ldud_ids,))

            for r in cur.fetchall():
                try:
                    d_start = datetime.fromisoformat(r["start_datetime"])
                    d_end = datetime.fromisoformat(r["end_datetime"])
                except (ValueError, TypeError):
                    continue
                delay_map.setdefault(r["ldud_id"], []).append((d_start, d_end))

        print("\nTOTAL ROWS:", len(rows))
        print("VCN IDS:", vcn_ids)
        print("BL MAP:", bl_map)
        print("DELAY MAP:", delay_map)

        report_data = {}

        for row in rows:
            vessel_name          = row['vessel_name']       or '-'
            cargo_name           = row['cargo_name']        or '-'
            vessel_seq           = row['vessel_seq']
            sort_order           = row['sort_order']
            discharge_started    = row['discharge_started']
            discharge_completed  = row['discharge_completed']

            key = f"{row['id']}"

            if key not in report_data:
                bl_qty = bl_map.get(row["vcn_id"], 0)

                report_data[key] = {
                    "ldud_id": row["id"],
                    "vessel_name": vessel_name,
                    "cargo_names": [],
                    "vessel_seq": vessel_seq,
                    "sort_order": sort_order,
                    "discharge_started": discharge_started,
                    "discharge_completed_dt": discharge_completed,
                    "discharge_completed": str(discharge_completed) if discharge_completed else "",

                    "bl_quantity": bl_qty,
                    "discharged_quantity": 0,
                    "balance_on_board": 0,

                    "daily_data": {}
                }
            if cargo_name != "-":
                if cargo_name not in report_data[key]["cargo_names"]:
                    report_data[key]["cargo_names"].append(cargo_name)

            if row["day_label"]:
                day = row["day_label"]

                if day not in report_data[key]["daily_data"]:
                    report_data[key]["daily_data"][day] = {
                        "date_day": day,
                        "cargo_name": "",
                        "total_qty": 0,
                        "ww_hrs": "-"
                    }

                report_data[key]["daily_data"][day]["total_qty"] += float(row["total_qty"] or 0)

        for vessel in report_data.values():

            vessel["cargo_name"] = " + ".join(sorted(vessel["cargo_names"]))

            vessel["daily_data"] = list(vessel["daily_data"].values())

            total_discharged = sum(
                float(d["total_qty"] or 0)
                for d in vessel["daily_data"]
            )

            vessel["discharged_quantity"] = total_discharged

            vessel["balance_on_board"] = max(
                vessel["bl_quantity"] - total_discharged,
                0
            )

            # ---------- W/W HRS CALCULATION ----------
            # Window for a given day_label D is 8:00 AM (D) -> 8:00 AM (D+1)
            d_start = vessel["discharge_started"]
            d_end   = vessel["discharge_completed_dt"]  # may be None if still discharging
            vessel_id = vessel["ldud_id"]
            vessel_delays = delay_map.get(vessel_id, [])

            total_ww_hours = 0.0  # running total in decimal hours, for the TOTAL row

            for day in vessel["daily_data"]:
                day_win_start = datetime.strptime(day["date_day"], "%d-%m-%Y").replace(
                    hour=8, minute=0, second=0, microsecond=0
                )
                day_win_end = day_win_start + timedelta(hours=24)

                if d_start is None:
                    day["ww_hrs"] = "-"
                    continue

                gross_start = max(d_start, day_win_start)
                gross_end = min(d_end, day_win_end) if d_end else day_win_end

                gross_seconds = (gross_end - gross_start).total_seconds()
                gross_hours = max(gross_seconds / 3600, 0)
                gross_hours = min(gross_hours, 24)

                # --- Subtract delay hours (Mother Vessel Agent / Force Majeure / MBP) ---
                deduction_hours = 0
                for delay_start, delay_end in vessel_delays:
                    clip_start = max(delay_start, day_win_start)
                    clip_end = min(delay_end, day_win_end)
                    overlap = (clip_end - clip_start).total_seconds() / 3600
                    if overlap > 0:
                        deduction_hours += overlap

                ww_hours = max(gross_hours - deduction_hours, 0)

                total_ww_hours += ww_hours
                day["ww_hrs"] = format_hrs_to_hms(ww_hours)

            # TOTAL row value for this vessel, same HH:MM:SS format
            vessel["ww_hrs_total"] = format_hrs_to_hms(total_ww_hours) if total_ww_hours > 0 else "-"

            # ---------- AVG DISCHARGE RATE PWWD ----------
            # = total discharged quantity / total W/W hours
            if total_ww_hours > 0:
                vessel["avg_discharge_rate"] = round(total_discharged / total_ww_hours, 2)
            else:
                vessel["avg_discharge_rate"] = 0

            del vessel["cargo_names"]
            del vessel["discharge_started"]
            del vessel["discharge_completed_dt"]
            del vessel["ldud_id"]

        final_data = sorted(
            report_data.values(),
            key=lambda x: x["vessel_seq"]
        )

        for item in final_data:
            item["daily_data"].sort(
                key=lambda d: (
                    datetime.strptime(d["date_day"], "%d-%m-%Y"),
                    d["cargo_name"]
                )
            )

        print("\nFINAL DATA COUNT:", len(final_data))
        for item in final_data:
            print(
                f"[seq={item['vessel_seq']}]"
                f"[sort_order={item['sort_order']}]"
                f"[discharge_completed={item['discharge_completed']}]",
                item['vessel_name'],
                item['cargo_name'],
                "BL:", item['bl_quantity'],
                "Discharged:", item['discharged_quantity'],
                "Balance:", item['balance_on_board'],
                "WW Total:", item['ww_hrs_total'],
                "Avg Rate:", item['avg_discharge_rate']
            )
            for d in item['daily_data']:
                print("   ", d['date_day'], "qty:", d['total_qty'], "ww_hrs:", d['ww_hrs'])

        print("\n========== MONTHLY REPORT END ==========\n")

        return jsonify({"success": True, "data": final_data})

    except Exception as e:
        import traceback
        traceback.print_exc()
        print("\n========== MONTHLY REPORT ERROR ==========")
        print(str(e))
        print("========== ERROR END ==========\n")
        return jsonify({"success": False, "message": str(e)})

    finally:
        cur.close()
        conn.close()
@bp.route(
    '/api/module/RP01/vessel_delay_report',
    methods=['GET']
)
@login_required
def vessel_delay_report():

    report_date = request.args.get('report_date')

    print("\n========== VESSEL DELAY REPORT START ==========")
    print("REPORT DATE:", report_date)

    if not report_date:

        return jsonify({
            "success": False,
            "message": "Report date required"
        })

    conn = get_db()
    cur = get_cursor(conn)

    try:
        report_dt = datetime.strptime(report_date, "%Y-%m-%d")

        window_end = datetime(
            report_dt.year,
            report_dt.month,
            report_dt.day,
            8, 0, 0
        )

        window_start = window_end - timedelta(hours=24)

        query = """

        WITH vessel_order AS (

            SELECT

                lh.id,
                lh.vessel_name,
                first_anchor.discharge_started,
                last_anchor.discharge_completed,

                ROW_NUMBER() OVER (
                    ORDER BY
                        first_anchor.discharge_started,
                        lh.id
                ) AS vessel_seq

            FROM ldud_header lh

            LEFT JOIN LATERAL (
                SELECT
                    MIN(discharge_started) AS discharge_started
                FROM ldud_anchorage
                WHERE ldud_id = lh.id
            ) first_anchor ON TRUE

            LEFT JOIN LATERAL (
                SELECT
                    CASE
                        WHEN EXISTS (
                            SELECT 1
                            FROM ldud_anchorage x
                            WHERE x.ldud_id = lh.id
                            AND x.discharge_started IS NOT NULL
                            AND x.discharge_commenced IS NULL
                        )
                        THEN NULL
                        ELSE MAX(discharge_commenced)
                    END AS discharge_completed
                FROM ldud_anchorage
                WHERE ldud_id = lh.id
            ) last_anchor ON TRUE

            WHERE
                first_anchor.discharge_started IS NOT NULL
                AND first_anchor.discharge_started < %s

                AND (
                    last_anchor.discharge_completed IS NULL
                    OR last_anchor.discharge_completed >= %s
                    OR EXISTS (
                        SELECT 1
                        FROM ldud_barge_lines b
                        WHERE b.ldud_id = lh.id
                        AND (
                                b.completed_discharge_berth IS NULL
                            OR b.cast_off_berth IS NULL
                        )
                    )
                )
        ),

        delay_summary AS (

            SELECT

                vo.id,
                vo.vessel_seq,
                vo.vessel_name,

                COALESCE(ld.delay_name,'') AS delay_name,

                COALESCE(ld.crane_number,'All') AS crane_number,

                SUM(COALESCE(ld.total_time_mins,0)) AS total_mins

            FROM vessel_order vo

            LEFT JOIN ldud_delays ld
                ON ld.ldud_id = vo.id

            WHERE

                TO_TIMESTAMP(
                    ld.start_datetime,
                    'YYYY-MM-DD"T"HH24:MI'
                ) < %s

                AND (

                    ld.end_datetime IS NULL

                    OR

                    TO_TIMESTAMP(
                        ld.end_datetime,
                        'YYYY-MM-DD"T"HH24:MI'
                    ) >= %s

                )

            GROUP BY

                vo.id,
                vo.vessel_seq,
                vo.vessel_name,
                ld.delay_name,
                ld.crane_number
        )

        SELECT

            vessel_seq,
            vessel_name,
            delay_name,

            STRING_AGG(

                CONCAT(

                    'Crane ',
                    crane_number,
                    ' : ',
                    FLOOR(total_mins / 60),
                    ' Hrs ',
                    MOD(total_mins::int,60),
                    ' Mins'

                ),

                E'\n'

                ORDER BY crane_number

            ) AS crane_details

        FROM delay_summary

        GROUP BY

            vessel_seq,
            vessel_name,
            delay_name

        ORDER BY

            vessel_seq,
            delay_name;

        """

        cur.execute(
            query,
            (
                window_end,
                window_start,
                window_end,
                window_start
            )
        )
        rows = cur.fetchall()

        print("\nTOTAL ROWS:")
        print(len(rows))

        grouped_data = {}

        for row in rows:

            vessel_name = (
                row['vessel_name']
                or '-'
            )

            if vessel_name not in grouped_data:
                grouped_data[vessel_name] = {
                    "vessel_seq": row["vessel_seq"],
                    "delays": []
                }

            if row['delay_name']:

                grouped_data[vessel_name]["delays"].append({

                    "delay_name":
                        row['delay_name'],

                    "crane_details":
                        row['crane_details']
                        or ''
                })

        data = []

        for vessel_name, vessel in sorted(
            grouped_data.items(),
            key=lambda x: x[1]["vessel_seq"]
        ):

            data.append({
                "vessel_seq": vessel["vessel_seq"],
                "vessel_name": vessel_name,
                "delays": vessel["delays"]
            })

        print("\nFINAL DATA:")
        print(data)

        print(
            "\n========== VESSEL DELAY REPORT END ==========\n"
        )

        return jsonify({

            "success": True,

            "data": data
        })

    except Exception as e:

        print(
            "\n========== VESSEL DELAY REPORT ERROR =========="
        )

        import traceback

        traceback.print_exc()

        print(str(e))

        print(
            "========== ERROR END ==========\n"
        )

        return jsonify({

            "success": False,

            "message": str(e)
        })

    finally:

        cur.close()
        conn.close()




@bp.route('/api/module/RP01/barge_status_report', methods=['GET'])
@login_required
def barge_status_report():

    print("\n================ BARGE STATUS REPORT START ================")

    report_date = request.args.get('report_date')
    print("REPORT DATE:", report_date)

    if not report_date:
        return jsonify({"success": False, "message": "Report date required"})

    conn = get_db()
    cur = get_cursor(conn)

    try:
        report_dt = datetime.strptime(report_date, "%Y-%m-%d")

        # Window for Step 1 (vessel selection) - unchanged, 8 AM to 8 AM.
        window_end = datetime(report_dt.year, report_dt.month, report_dt.day, 8, 0, 0)
        window_start = window_end - timedelta(hours=24)
        ws_str = window_start.strftime("%Y-%m-%dT%H:%M")


        # Barge window: Previous day 06:00 AM to selected day 06:00 AM
        barge_window_end = datetime(
            report_dt.year,
            report_dt.month,
            report_dt.day,
            6, 0, 0
        )

        barge_window_start = barge_window_end - timedelta(hours=24)

        print("WINDOW START:", window_start)
        print("WINDOW END:", window_end)
        print("BARGE WINDOW START:", barge_window_start)
        print("BARGE WINDOW END:", barge_window_end)

        # =========================================================
        # STEP 1 : FETCH HEADER ROWS
        # =========================================================

        print("\nSTEP 1 : FETCH HEADER ROWS")

        cur.execute("""
            SELECT lh.id, lh.vcn_id, lh.vessel_name
            FROM ldud_header lh
            LEFT JOIN LATERAL (
                SELECT MIN(discharge_started) AS discharge_started
                FROM ldud_anchorage
                WHERE ldud_id = lh.id
            ) first_anchor ON TRUE
            LEFT JOIN LATERAL (
                SELECT MAX(discharge_commenced) AS discharge_completed
                FROM ldud_anchorage
                WHERE ldud_id = lh.id
                AND discharge_commenced IS NOT NULL
            ) last_anchor ON TRUE
            WHERE
                first_anchor.discharge_started IS NOT NULL
                AND first_anchor.discharge_started < %s
                AND (
                    last_anchor.discharge_completed IS NULL
                    OR last_anchor.discharge_completed >= %s
                    OR EXISTS (
                        SELECT 1
                        FROM ldud_barge_lines b
                        WHERE b.ldud_id = lh.id
                        AND (
                            b.completed_discharge_berth IS NULL
                            OR b.cast_off_berth IS NULL
                        )
                    )
                )
        """, (window_end, window_start))

        header_rows = cur.fetchall()

        print("HEADER ROW COUNT:", len(header_rows))
        for row in header_rows:
            print(row)

        ldud_ids = [r["id"] for r in header_rows]

        print("LDUD IDS:", ldud_ids)

        ldud_to_vcn = {
            r["id"]: r["vcn_id"]
            for r in header_rows
        }

        print("LDUD TO VCN:", ldud_to_vcn)

        ldud_to_vessel_name = {
            r["id"]: (r["vessel_name"] or f"Vessel {r['id']}")
            for r in header_rows
        }

        print("LDUD TO VESSEL NAME:", ldud_to_vessel_name)

        vcn_ids = list({
            r["vcn_id"]
            for r in header_rows
            if r["vcn_id"]
        })

        print("VCN IDS:", vcn_ids)

        # =========================================================
        # STEP 2 : FETCH LUEU ACTUAL
        # =========================================================

        print("\nSTEP 2 : FETCH LUEU ACTUAL")

        barge_actual = {}

        if vcn_ids:

            cur.execute("""
                SELECT
                    source_id,
                    UPPER(TRIM(SPLIT_PART(barge_name,'/',1))) AS base_barge,
                    TRIM(SPLIT_PART(barge_name,'/',2)) AS trip_no,

                    MAX(route_name) AS route_name,
                    MAX(equipment_name) AS crane,

                    SUM(COALESCE(quantity,0)) AS actual_qty

                FROM lueu_lines

                WHERE source_type = 'VCN'
                AND is_deleted = false
                AND source_id = ANY(%s)
                AND barge_name IS NOT NULL
                AND quantity IS NOT NULL
                AND TO_DATE(entry_date,'YYYY-MM-DD') <= %s

                GROUP BY
                    source_id,
                    UPPER(TRIM(SPLIT_PART(barge_name,'/',1))),
                    TRIM(SPLIT_PART(barge_name,'/',2))
            """, (
                vcn_ids,
                # FIX: was `report_date - timedelta(...)` - report_date is
                # the raw request string, not a date. Subtracting a
                # timedelta from a string raises TypeError. Use the
                # already-parsed report_dt.date() instead.
                report_dt.date() - timedelta(days=1)
            ))

            actual_rows = cur.fetchall()

            print("LUEU ROW COUNT:", len(actual_rows))

            for r in actual_rows:

                print(r)

                barge_actual[
                    (
                        r["source_id"],
                        r["base_barge"],
                        r["trip_no"]
                    )
                ] = {
                    "actual_qty": float(r["actual_qty"]),
                    "route_name": r["route_name"] or "",
                    "crane": r["crane"] or ""
                }

        print("BARGE ACTUAL DICT SIZE:", len(barge_actual))

        # =========================================================
        # STEP 3 : BUILD BARGE STATUS BUCKETS
        # =========================================================

        print("\nSTEP 3 : BUILD BARGE STATS")

        barge_stats = {
            lid: {
                "all": set(),

                # Logic implemented
                "at_jetty": [],
                "waiting_discharge": [],
                "waiting_empty_jetty": [],
                "at_gull_loaded": [],
                "under_loading": [],
                "waiting_loading": [],
                "remarks": [],
               

                # Logic not implemented yet
                "r19_waiting_loaded": "NA",
                # "transit_mv_to_jetty_loaded": "NA",
                "empty_at_gull_r19": "NA",
                "in_transit_jetty_to_mv": [],
                "breakdown": [],
            }
            for lid in ldud_ids
        }

        if ldud_ids:

            cur.execute("""
            SELECT
                h.vcn_id,

                b.ldud_id,
                b.barge_name,
                b.discharge_quantity,
                b.port_crane,

                b.along_side_vessel,
                b.commenced_loading,
                b.completed_loading,
                b.cast_off_mv,

                b.anchored_gull_island,
                b.aweigh_gull_island,
                b.amf_at_port,

                b.along_side_berth,
                b.commence_discharge_berth,
                b.completed_discharge_berth,
                b.cast_off_berth,
                b.cast_off_port,

                ROW_NUMBER() OVER (
                    PARTITION BY
                        h.vcn_id,
                        UPPER(TRIM(b.barge_name))
                    ORDER BY
                        b.commence_discharge_berth
                ) AS trip_no

            FROM ldud_barge_lines b

            JOIN ldud_header h
                ON h.id = b.ldud_id

            WHERE b.ldud_id = ANY(%s)
            AND (
                b.cast_off_port IS NULL
                OR b.cast_off_port > %s
            )
            ORDER BY
                h.vcn_id,
                b.barge_name,
                trip_no

            """, (
                ldud_ids,
                ws_str
            ))
            barge_line_rows = cur.fetchall()

            print("\n========== ALL BARGES ==========")
            print("COUNT:", len(barge_line_rows))

            for r in barge_line_rows:
                print(
                    r["ldud_id"],
                    r["barge_name"],
                    r["trip_no"],
                    r["along_side_vessel"],
                    r["commence_discharge_berth"],
                    r["completed_discharge_berth"],
                    r["cast_off_berth"],
                    r["cast_off_port"]
                )

            print("BARGE LINE ROW COUNT:", len(barge_line_rows))

            for r in barge_line_rows:

                print(r)

                lid = r["ldud_id"]

                bn = (r["barge_name"] or "").strip()

                if not bn:
                    continue

                barge_stats[lid]["all"].add(bn)

                # --------------------------
                # Existing status logic
                # --------------------------

                # Determine barge status

                if r['cast_off_port']:
                    status = 'None'
                elif r['completed_discharge_berth'] and not r['cast_off_berth']:
                    status = 'waiting_empty_jetty'
                elif r['commence_discharge_berth'] and not r['cast_off_berth']:
                    status = 'at_jetty'
                elif r['along_side_berth'] and not r['commence_discharge_berth']:
                    status = 'waiting_discharge'
                elif r['cast_off_mv'] and not r['along_side_berth']:
                    status = 'at_gull_loaded'
                elif r['commenced_loading'] and not r['completed_loading']:
                    status = 'under_loading'
                elif r['along_side_vessel'] and not r['commenced_loading']:
                    status = 'waiting_loading'
                else:
                    status = None

                # Add barge to the appropriate status list
                if status and status in barge_stats[lid]:
                    barge_stats[lid][status].append(bn)

        json_safe = {}

        for lid, stats in barge_stats.items():

            vessel = {
                "vessel_name": ldud_to_vessel_name.get(lid, f"Vessel {lid}")
            }

            for key, value in stats.items():

                if key == "all":
                    vessel[key] = sorted(value)

                elif isinstance(value, list):
                    # Keep implemented statuses as lists
                    vessel[key] = value

                else:
                    # Only unimplemented statuses remain "NA"
                    vessel[key] = value

            json_safe[lid] = vessel

        print("JSON SAFE:")
        print(json_safe)

        return jsonify({
            "success": True,
            "data": json_safe
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": str(e)
        })

    finally:
        cur.close()
        conn.close()

@bp.route(
    '/api/module/RP01/barge_discharge_report',
    methods=['GET']
)
@login_required
def barge_discharge_report():

    report_date = request.args.get('report_date')

    if not report_date:
        return jsonify({"success": False, "message": "Report date required"})

    try:
        report_date_obj = datetime.strptime(report_date, "%Y-%m-%d")
    except ValueError:
        return jsonify({"success": False, "message": "Invalid report_date format, expected YYYY-MM-DD"})

    # Selecting "13" means: show data for 12th (report_date - 1 day)
    data_date_obj = report_date_obj - timedelta(days=1)
    data_date = data_date_obj.strftime("%Y-%m-%d")

    # Month-to-date: 1st of that month -> data_date
    month_start = data_date_obj.replace(day=1).strftime("%Y-%m-%d")

    # Fiscal year start (assumption: Apr 1)
    fy_start_year = data_date_obj.year if data_date_obj.month >= 4 else data_date_obj.year - 1
    fy_start = f"{fy_start_year}-04-01"

    # Historic table (rp01_historical_lueu) covers April only.
    # Live table (lueu_lines) covers May 1 onward for this FY.
    historic_cutoff = f"{fy_start_year}-04-30"
    live_start = f"{fy_start_year}-05-01"

    EQUIP_EXPR = """
        CASE TRIM(equipment_name)
            WHEN 'Barge Unloader 1' THEN 'BUL-01'
            WHEN 'Barge Unloader 2' THEN 'BUL-02'
            ELSE COALESCE(TRIM(equipment_name), 'Others')
        END
    """

    def cat_key(ctype, ccat):
        return f"{ctype}||{ccat}"

    conn = get_db()
    cur = get_cursor(conn)

    try:
        # 0) FULL master list of cargo_type -> cargo_category
        cur.execute("""
            SELECT
                COALESCE(vc.cargo_type, 'Others') AS cargo_type,
                COALESCE(vc.cargo_category, 'Others') AS cargo_category
            FROM vessel_cargo vc
            GROUP BY vc.cargo_type, vc.cargo_category
            ORDER BY vc.cargo_type, vc.cargo_category
        """)
        master_cargo_rows = cur.fetchall()

        # 0b) FULL master list of equipment (live table only)
        cur.execute(f"""
            SELECT DISTINCT {EQUIP_EXPR} AS equipment
            FROM lueu_lines
            WHERE is_deleted IS NOT TRUE
            ORDER BY 1
        """)
        master_equipment_rows = cur.fetchall()

        # 1) Day-level matrix: equipment x (cargo_type, cargo_category) for entry_date = data_date ONLY (live table only)
        cur.execute(f"""
            SELECT
                COALESCE(vc.cargo_type, 'Others') AS cargo_type,
                COALESCE(vc.cargo_category, 'Others') AS cargo_category,
                {EQUIP_EXPR.replace('equipment_name', 'lbl.equipment_name')} AS equipment,
                SUM(lbl.quantity) AS day_qty
            FROM lueu_lines lbl
            LEFT JOIN vessel_cargo vc
                ON LOWER(TRIM(vc.cargo_name)) = LOWER(TRIM(lbl.cargo_name))
            WHERE lbl.is_deleted IS NOT TRUE
              AND lbl.quantity IS NOT NULL
              AND lbl.entry_date::date = %(data_date)s::date
            GROUP BY vc.cargo_type, vc.cargo_category, {EQUIP_EXPR.replace('equipment_name', 'lbl.equipment_name')}
        """, {"data_date": data_date})
        day_rows = cur.fetchall()

        # 2a) Historic total (rp01_historical_lueu): fy_start -> historic_cutoff (April 1 - April 30)
        cur.execute("""
            SELECT
                COALESCE(vc.cargo_type, 'Others') AS cargo_type,
                COALESCE(vc.cargo_category, 'Others') AS cargo_category,
                SUM(COALESCE(h.quantity, 0)) AS historic_total
            FROM rp01_historical_lueu h
            LEFT JOIN vessel_cargo vc
                ON LOWER(TRIM(vc.cargo_name)) = LOWER(TRIM(h.cargo_name))
            WHERE h.quantity IS NOT NULL
              AND h.entry_date BETWEEN %(fy_start)s::date AND %(historic_cutoff)s::date
            GROUP BY vc.cargo_type, vc.cargo_category
        """, {"fy_start": fy_start, "historic_cutoff": historic_cutoff})
        historic_rows = cur.fetchall()

        # 2b) Live total (lueu_lines): month_total (calendar month -> data_date),
        #     live_fy_total (live_start -> data_date), live_cumulative_total (live_start -> data_date)
        cur.execute("""
            SELECT
                COALESCE(vc.cargo_type, 'Others') AS cargo_type,
                COALESCE(vc.cargo_category, 'Others') AS cargo_category,
                SUM(COALESCE(lbl.quantity, 0)) AS live_fy_total,
                SUM(CASE WHEN lbl.entry_date::date BETWEEN %(month_start)s::date AND %(data_date)s::date
                    THEN COALESCE(lbl.quantity, 0) ELSE 0 END) AS month_total,
                SUM(CASE WHEN lbl.entry_date::date <= %(data_date)s::date
                    THEN COALESCE(lbl.quantity, 0) ELSE 0 END) AS live_cumulative_total
            FROM lueu_lines lbl
            LEFT JOIN vessel_cargo vc
                ON LOWER(TRIM(vc.cargo_name)) = LOWER(TRIM(lbl.cargo_name))
            WHERE lbl.is_deleted IS NOT TRUE
              AND lbl.quantity IS NOT NULL
              AND lbl.entry_date::date BETWEEN %(live_start)s::date AND %(data_date)s::date
            GROUP BY vc.cargo_type, vc.cargo_category
        """, {"live_start": live_start, "data_date": data_date, "month_start": month_start})
        live_rows = cur.fetchall()

        # 3) Per-equipment totals: day total (live only) and month total (live only)
        cur.execute(f"""
            SELECT
                {EQUIP_EXPR} AS equipment,
                SUM(CASE WHEN entry_date::date = %(data_date)s::date
                    THEN quantity ELSE 0 END) AS day_total,
                SUM(CASE WHEN entry_date::date BETWEEN %(month_start)s::date AND %(data_date)s::date
                    THEN quantity ELSE 0 END) AS month_total
            FROM lueu_lines
            WHERE is_deleted IS NOT TRUE AND quantity IS NOT NULL
            GROUP BY {EQUIP_EXPR}
        """, {"data_date": data_date, "month_start": month_start})
        equipment_totals_rows = cur.fetchall()

        # ---- Build FULL cargo hierarchy: cargo_type -> [cargo_categories] ----
        cargo_hierarchy = {}
        for row in master_cargo_rows:
            ctype = row['cargo_type']
            ccat = row['cargo_category']
            cargo_hierarchy.setdefault(ctype, [])
            if ccat not in cargo_hierarchy[ctype]:
                cargo_hierarchy[ctype].append(ccat)

        all_keys = [cat_key(row['cargo_type'], row['cargo_category']) for row in master_cargo_rows]

        # ---- Seed equipment_rows with the FULL equipment list (zeros by default) ----
        equipment_rows = {}
        for row in master_equipment_rows:
            equipment_rows[row['equipment']] = {k: 0 for k in all_keys}

        for row in day_rows:
            equip = row['equipment']
            key = cat_key(row['cargo_type'], row['cargo_category'])
            equipment_rows.setdefault(equip, {k: 0 for k in all_keys})
            equipment_rows[equip][key] = int(row['day_qty'] or 0)

        # ---- Historic totals (April) per category ----
        historic_totals = {}
        for row in historic_rows:
            key = cat_key(row['cargo_type'], row['cargo_category'])
            historic_totals[key] = int(row['historic_total'] or 0)

        # ---- Live totals (May onward) merged with historic to form FY / cumulative ----
        month_totals = {}
        fy_totals = {}
        cumulative_totals = {}
        for row in live_rows:
            key = cat_key(row['cargo_type'], row['cargo_category'])
            month_totals[key] = int(row['month_total'] or 0)
            live_fy = int(row['live_fy_total'] or 0)
            live_cum = int(row['live_cumulative_total'] or 0)
            hist = historic_totals.get(key, 0)
            fy_totals[key] = hist + live_fy
            cumulative_totals[key] = hist + live_cum

        # Categories that ONLY had historic (April) activity, with nothing live yet
        for key, hist_val in historic_totals.items():
            if key not in fy_totals:
                fy_totals[key] = hist_val
                cumulative_totals[key] = hist_val
                month_totals.setdefault(key, 0)

        # ---- Seed equipment_totals with full equipment list too ----
        equipment_totals = {row['equipment']: {"total_day": 0, "total_month": 0} for row in master_equipment_rows}
        for row in equipment_totals_rows:
            equipment_totals[row['equipment']] = {
                "total_day": int(row['day_total'] or 0),
                "total_month": int(row['month_total'] or 0),
            }

        final_data = {
            "cargo_hierarchy": cargo_hierarchy,
            "equipment_rows": equipment_rows,
            "equipment_totals": equipment_totals,
            "month_totals": month_totals,
            "fy_totals": fy_totals,
            "cumulative_totals": cumulative_totals,
        }

        return jsonify({"success": True, "data": final_data})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)})

    finally:
        cur.close()
        conn.close()

# from flask import request, jsonify
# from datetime import datetime, timedelta
# # from your_app import bp, login_required, get_db, get_cursor  # keep your existing imports


        
@bp.route(
    '/api/module/RP01/mbc_report',
    methods=['GET']
)
@login_required
def mbc_report():

    report_date = request.args.get('report_date')

    print("\n========== MBC REPORT START ==========")
    print("REPORT DATE:", report_date)

    if not report_date:

        return jsonify({

            "success": False,

            "message": "Report date required"
        })

    conn = get_db()

    cur = get_cursor(conn)

    try:
        report_dt = datetime.strptime(report_date, "%Y-%m-%d")

        window_end = datetime(
            report_dt.year,
            report_dt.month,
            report_dt.day,
            8, 0, 0
        )

        window_start = window_end - timedelta(hours=24)

        print("WINDOW START:", window_start)
        print("WINDOW END:", window_end)

        query = """

        SELECT

            mh.id,
            mh.doc_num,
            mh.mbc_name,
            mh.cargo_name,
            mh.bl_quantity,
            mh.load_port,

            dpl.vessel_arrival_port,
            dpl.unloading_commenced,
            dpl.cleaning_commenced,
            dpl.unloading_completed,
            dpl.sailed_out_load_port

        FROM mbc_header mh

        LEFT JOIN mbc_discharge_port_lines dpl
            ON dpl.mbc_id = mh.id

        WHERE
            NULLIF(TRIM(dpl.unloading_completed), '') IS NOT NULL
            AND NULLIF(TRIM(dpl.unloading_completed), '')::timestamp >= %s
            AND NULLIF(TRIM(dpl.unloading_completed), '')::timestamp < %s

    
        ORDER BY
            NULLIF(TRIM(dpl.unloading_completed), '')::timestamp

        """
        print("\nEXECUTING QUERY:")
        print(query)

        print("\nQUERY PARAM:")
        print(report_date)

        cur.execute(
        query,
        (
            window_start,
            window_end
        )
    )

        print("\nQUERY EXECUTED SUCCESSFULLY")

        rows = cur.fetchall()

        print("\nTOTAL ROWS:")
        print(len(rows))

        data = []

        sr_no = 1

        for row in rows:

            print("\nCURRENT ROW:")
            print(row)

            data.append({

            "mbc_no": sr_no,

            "mbc_name": row["mbc_name"] or "",

            "cargo": row["cargo_name"] or "",

            "source": row["load_port"] or "",

            "qty": int(row["bl_quantity"] or 0),

            "arrived_at_jetty":
                _parse_flexible_dt(row["vessel_arrival_port"], "%d-%m-%Y : %H:%M"),

            "commence_loading":
                _parse_flexible_dt(row["unloading_commenced"], "%d-%m-%Y : %H:%M"),

            "cleaning_start":
                _parse_flexible_dt(row["cleaning_commenced"], "%d-%m-%Y : %H:%M"),

            "completed_loading":
                _parse_flexible_dt(row["unloading_completed"], "%d-%m-%Y : %H:%M"),

            "cast_off_jetty":
                row["sailed_out_load_port"].strftime("%d-%m-%Y : %H:%M")
                if row["sailed_out_load_port"] else ""

        })

            sr_no += 1

        print("\nFINAL DATA:")
        print(data)

        print("\n========== MBC REPORT END ==========\n")

        return jsonify({

            "success": True,

            "data": data
        })

    except Exception as e:

        print("\n========== MBC REPORT ERROR ==========")

        import traceback

        traceback.print_exc()

        print(str(e))

        print("========== ERROR END ==========\n")

        return jsonify({

            "success": False,

            "message": str(e)
        })

    finally:

        print("\nCLOSING DB CONNECTION")

        cur.close()

        conn.close()

@bp.route(
    '/api/module/RP01/arrived_vessel_report',
    methods=['GET']
)
@login_required
def arrived_vessel_report():

    report_date = request.args.get("report_date")

    print("\n========== ARRIVED VESSEL REPORT START ==========")
    print("REPORT DATE:", report_date)

    if not report_date:

        return jsonify({

            "success": False,

            "message": "Report date required"

        })

    conn = get_db()
    cur = get_cursor(conn)

    try:

        query = """

        SELECT

            lh.id,

            lh.vessel_name,

            STRING_AGG(
                DISTINCT TRIM(vcd.cargo_name),
                ', '
                ORDER BY TRIM(vcd.cargo_name)
            ) AS cargo,

            SUM(
                COALESCE(vcd.bl_quantity,0)
            ) AS bl_quantity,

            vh.load_port,

            lh.nor_accepted

        FROM ldud_header lh

        LEFT JOIN vcn_header vh
            ON vh.id = lh.vcn_id

        LEFT JOIN vcn_cargo_declaration vcd
            ON vcd.vcn_id = lh.vcn_id

        WHERE

            NULLIF(TRIM(lh.nor_accepted), '') IS NOT NULL

            AND DATE(
                NULLIF(TRIM(lh.nor_accepted), '')::timestamp
            )
            BETWEEN
                (%s::date - INTERVAL '1 day')
            AND
                %s::date

        GROUP BY

            lh.id,
            lh.vessel_name,
            vh.load_port,
            lh.nor_accepted

        ORDER BY

            NULLIF(TRIM(lh.nor_accepted), '')::timestamp

        """

        print("\nEXECUTING QUERY:")
        print(query)

        print("\nQUERY PARAM:")
        print(report_date)

        cur.execute(

            query,

            (
                report_date,
                report_date
            )

        )

        rows = cur.fetchall()

        print("\nTOTAL ROWS:")
        print(len(rows))

        data = []

        sr_no = 1

        for row in rows:

            print("\nCURRENT ROW:")
            print(row)

            arrived_mumbai = _parse_flexible_dt(row["nor_accepted"], "%d-%m-%Y %H:%M")

            data.append({

                "vsl_no": sr_no,

                "vessel_name":
                    row["vessel_name"] or "",

                "cargo":
                    row["cargo"] or "",

                "bl_qty":
                    int(row["bl_quantity"] or 0),

                "load_port":
                    row["load_port"] or "",

                "arrived_mumbai":
                    arrived_mumbai

            })

            sr_no += 1

        print("\nFINAL DATA:")
        print(data)

        print("\n========== ARRIVED VESSEL REPORT END ==========\n")

        return jsonify({

            "success": True,

            "data": data

        })

    except Exception as e:

        import traceback

        traceback.print_exc()

        return jsonify({

            "success": False,

            "message": str(e)

        })

    finally:

        cur.close()
        conn.close()

@bp.route(
    '/api/module/RP01/mbc_arrived_report',
    methods=['GET']
)
@login_required
def mbc_arrived_report():

    report_date = request.args.get("report_date")

    print("\n========== MBC ARRIVED REPORT START ==========")
    print("REPORT DATE:", report_date)

    if not report_date:
        return jsonify({
            "success": False,
            "message": "Report date required"
        })

    conn = get_db()
    cur = get_cursor(conn)

    try:

        query = """

        SELECT

            mh.id,

            mh.mbc_name,

            mh.cargo_name,

            mh.bl_quantity,

            mh.load_port,

            dpl.vessel_arrival_port

        FROM mbc_header mh

        LEFT JOIN mbc_discharge_port_lines dpl
            ON dpl.mbc_id = mh.id

        WHERE

            NULLIF(TRIM(dpl.vessel_arrival_port), '') IS NOT NULL

            AND DATE(
                NULLIF(TRIM(dpl.vessel_arrival_port), '')::timestamp
            )
            BETWEEN
                (%s::date - INTERVAL '1 day')
            AND
                %s::date

        ORDER BY

            NULLIF(TRIM(dpl.vessel_arrival_port), '')::timestamp

        """

        cur.execute(
            query,
            (
                report_date,
                report_date
            )
        )

        rows = cur.fetchall()

        data = []

        sr_no = 1

        for row in rows:

            arrived = _parse_flexible_dt(row["vessel_arrival_port"], "%d-%m-%Y %H:%M")

            data.append({

                "mbc_no": sr_no,

                "mbc_name": row["mbc_name"] or "",

                "cargo": row["cargo_name"] or "",

                "bl_qty": int(row["bl_quantity"] or 0),

                "load_port": row["load_port"] or "",

                "arrived_dharamtar": arrived

            })

            sr_no += 1

        return jsonify({

            "success": True,

            "data": data

        })

    except Exception as e:

        import traceback
        traceback.print_exc()

        return jsonify({

            "success": False,

            "message": str(e)

        })

    finally:

        cur.close()
        conn.close()

@bp.route(
    '/api/module/RP01/upcoming_vessel_report',
    methods=['GET']
)
@login_required
def upcoming_vessel_report():

    report_date = request.args.get('report_date')

    print("\n========== UPCOMING VESSEL REPORT START ==========")
    print("REPORT DATE:", report_date)

    if not report_date:

        return jsonify({

            "success": False,

            "message": "Report date required"
        })

    conn = get_db()

    cur = get_cursor(conn)

    try:

        query = """

        SELECT

            vn.id,

            vh.vessel_name,

            STRING_AGG(DISTINCT TRIM(vcd.cargo_name), ', ' ORDER BY TRIM(vcd.cargo_name)) AS cargo,

            SUM(COALESCE(vcd.bl_quantity,0)) AS bl_quantity,

            vh.load_port,

            vn.eta

        FROM vcn_nominations vn

        LEFT JOIN vcn_header vh
            ON vh.id = vn.vcn_id

        LEFT JOIN vcn_cargo_declaration vcd
            ON vcd.vcn_id = vh.id

        WHERE DATE(vn.eta) >= %s

        GROUP BY
            vn.id,
            vh.vessel_name,
            vh.load_port,
            vn.eta

        ORDER BY
            vn.eta

        """

        print("\nEXECUTING QUERY:")
        print(query)

        print("\nQUERY PARAM:")
        print(report_date)

        cur.execute(

            query,

            (
                report_date,
            )
        )

        print("\nQUERY EXECUTED SUCCESSFULLY")

        rows = cur.fetchall()

        print("\nTOTAL ROWS:")
        print(len(rows))

        data = []

        sr_no = 1

        for row in rows:

            print("\nCURRENT ROW:")
            print(row)

            eta_mumbai = _parse_flexible_dt(row["eta"], "%d-%m-%Y %H:%M")

            current_row = {

                "vsl_no": sr_no,

                "vessel_name": row["vessel_name"] or "",

                "cargo": row["cargo"] or "",

                "bl_qty": int(row["bl_quantity"] or 0),

                "load_port": row["load_port"] or "",

                "eta_mumbai": eta_mumbai

            }

            print("\nFORMATTED ROW:")
            print(current_row)

            data.append(current_row)

            sr_no += 1

        print("\nFINAL DATA:")
        print(data)

        print("\n========== UPCOMING VESSEL REPORT END ==========\n")

        return jsonify({

            "success": True,

            "data": data

        })

    except Exception as e:

        print("\n========== UPCOMING VESSEL REPORT ERROR ==========")

        import traceback

        traceback.print_exc()

        print(str(e))

        print("========== ERROR END ==========\n")

        return jsonify({

            "success": False,

            "message": str(e)

        })

    finally:

        print("\nCLOSING DB CONNECTION")

        cur.close()

        conn.close()

@bp.route(
    '/api/module/RP01/mbc_expected_report',
    methods=['GET']
)
@login_required
def mbc_expected_report():

    report_date = request.args.get("report_date")

    print("\n========== MBC EXPECTED REPORT START ==========")
    print("REPORT DATE:", report_date)

    if not report_date:

        return jsonify({
            "success": False,
            "message": "Report date required"
        })

    conn = get_db()
    cur = get_cursor(conn)

    try:

        query = """

        SELECT

            mh.id,

            mh.mbc_name,

            mh.cargo_name,

            mh.bl_quantity,

            mh.load_port,

            dpl.arrival_gull_island

        FROM mbc_header mh

        LEFT JOIN mbc_discharge_port_lines dpl
            ON dpl.mbc_id = mh.id

        WHERE

            NULLIF(TRIM(dpl.arrival_gull_island), '') IS NOT NULL

            AND DATE(
                NULLIF(TRIM(dpl.arrival_gull_island), '')::timestamp
            )
            BETWEEN
                (%s::date - INTERVAL '1 day')
            AND
                %s::date

        ORDER BY

            NULLIF(TRIM(dpl.arrival_gull_island), '')::timestamp

        """

        print(query)
        print(report_date)

        cur.execute(
            query,
            (
                report_date,
                report_date
            )
        )

        rows = cur.fetchall()

        print("TOTAL ROWS:", len(rows))

        data = []

        sr_no = 1

        for row in rows:

            eta_mumbai = _parse_flexible_dt(row["arrival_gull_island"], "%d-%m-%Y %H:%M")

            data.append({

                "mbc_no": sr_no,

                "mbc_name": row["mbc_name"] or "",

                "cargo": row["cargo_name"] or "",

                "bl_qty": int(row["bl_quantity"] or 0),

                "load_port": row["load_port"] or "",

                "eta_mumbai": eta_mumbai

            })

            sr_no += 1

        return jsonify({

            "success": True,

            "data": data

        })

    except Exception as e:

        import traceback
        traceback.print_exc()

        return jsonify({

            "success": False,

            "message": str(e)

        })

    finally:

        cur.close()
        conn.close()

@bp.route(
    '/api/module/RP01/tide_report',
    methods=['GET']
)
@login_required
def tide_report():

    report_date = request.args.get("report_date")

    print("\n========== TIDE REPORT START ==========")
    print("REPORT DATE:", report_date)

    if not report_date:

        return jsonify({
            "success": False,
            "message": "Report date required"
        })

    conn = get_db()
    cur = get_cursor(conn)

    try:

        cur.execute("""

            SELECT

                tide_datetime,

                tide_meters

            FROM tide_master

            WHERE

                NULLIF(TRIM(tide_datetime), '') IS NOT NULL

                AND DATE(
                    NULLIF(TRIM(tide_datetime), '')::timestamp
                ) = %s

            ORDER BY

                NULLIF(TRIM(tide_datetime), '')::timestamp

        """, (report_date,))

        rows = cur.fetchall()

        print("ROWS:", rows)

        if not rows:

            return jsonify({
                "success": True,
                "data": []
            })

        heights = [float(r["tide_meters"] or 0) for r in rows]

        if len(heights) >= 2:
            current = "HW" if heights[0] > heights[1] else "LW"
        else:
            current = "HW"

        tide_data = []

        for row in rows:

            time_disp = _parse_flexible_dt(row["tide_datetime"], "%H:%M")

            tide_data.append({

                "type": current,

                "time": time_disp,

                "mtrs": float(row["tide_meters"] or 0)

            })

            current = "LW" if current == "HW" else "HW"

        print(tide_data)

        return jsonify({

            "success": True,

            "data": tide_data

        })

    except Exception as e:

        import traceback
        traceback.print_exc()

        return jsonify({

            "success": False,

            "message": str(e)

        })

    finally:

        cur.close()
        conn.close()


@bp.route(
    '/api/module/RP01/vessel_discharge_summary',
    methods=['GET']
)
@login_required
def vessel_discharge_summary():

    report_date = request.args.get('report_date')

    print("\n========== VESSEL DISCHARGE SUMMARY START ==========")
    print("REPORT DATE:", report_date)

    if not report_date:

        print("ERROR: REPORT DATE MISSING")

        return jsonify({

            "success": False,

            "message": "Report date required"
        })

    conn = get_db()

    cur = get_cursor(conn)

    try:
        report_dt = datetime.strptime(report_date, "%Y-%m-%d")

        # ----------------------------------------------------------------
        # WINDOW LOGIC
        #
        # window_end   = selected report_date at 08:00 AM
        # window_start = 1st of the "active" month at 08:00 AM
        #
        # If report_date is the 1st of a month, there is no "month to
        # date" data for the current month yet, so we roll back and show
        # the ENTIRE previous month instead.
        #
        # Example: report_date = 2026-07-01
        #   -> window_start = 2026-06-01 08:00:00
        #   -> window_end   = 2026-07-01 08:00:00   (covers all of June)
        #
        # Example: report_date = 2026-07-15
        #   -> window_start = 2026-07-01 08:00:00
        #   -> window_end   = 2026-07-15 08:00:00   (month-to-date)
        # ----------------------------------------------------------------

        if report_dt.day == 1:
            prev_month_last_day = report_dt.replace(day=1) - timedelta(days=1)
            month_start_date = prev_month_last_day.replace(day=1)
        else:
            month_start_date = report_dt.replace(day=1)

        window_start = month_start_date.replace(
            hour=8, minute=0, second=0, microsecond=0
        )

        window_end = report_dt.replace(
            hour=8, minute=0, second=0, microsecond=0
        )

        print("WINDOW START:", window_start)
        print("WINDOW END:", window_end)

        query = """

        SELECT

            lh.id,

            lh.vessel_name,

            STRING_AGG(
                    DISTINCT TRIM(vcd.cargo_name),
                    ', '
                    ORDER BY TRIM(vcd.cargo_name)
                ) AS cargo_name,

                SUM(
                    COALESCE(vcd.bl_quantity, 0)
                ) AS bl_quantity,

            ----------------------------------------------------------------
            -- VCN DOC NUMBER
            ----------------------------------------------------------------
            vh.vcn_doc_num,

            ----------------------------------------------------------------
            -- LOAD PORT
            ----------------------------------------------------------------
            vh.load_port,

            ----------------------------------------------------------------
            -- DISCHARGE COMMENCED
            ----------------------------------------------------------------
            MIN(
                la.discharge_started
            ) AS discharge_commenced,

            ----------------------------------------------------------------
            -- DISCHARGE COMPLETED
            ----------------------------------------------------------------
            MAX(
                la.discharge_commenced
            ) AS discharge_completed,

            ----------------------------------------------------------------
            -- TIME TAKEN IN HOURS
            ----------------------------------------------------------------
            ROUND(

                EXTRACT(
                    EPOCH FROM
                    (
                        MAX(la.discharge_commenced)
                        -
                        MIN(la.discharge_started)
                    )
                ) / 3600,

                2

            ) AS time_taken_hrs

        FROM ldud_header lh

        LEFT JOIN vcn_cargo_declaration vcd
            ON lh.vcn_id = vcd.vcn_id

        LEFT JOIN vcn_header vh
            ON vh.id = lh.vcn_id

        LEFT JOIN ldud_anchorage la
            ON la.ldud_id = lh.id

        WHERE

        la.discharge_commenced IS NOT NULL

        AND la.discharge_commenced >= %s
        AND la.discharge_commenced <  %s

        GROUP BY

            lh.id,
            lh.vessel_name,
            vh.load_port,
            vh.vcn_doc_num

        ORDER BY

            MAX(la.discharge_commenced)

        """

        print("\nEXECUTING QUERY:")
        print(query)

        print("\nQUERY PARAMS:")
        print(window_start, window_end)

        cur.execute(

            query,

            (
                window_start,
                window_end
            )

        )

        print("\nQUERY EXECUTED SUCCESSFULLY")

        rows = cur.fetchall()

        print("\nRAW FETCHED ROWS:")
        print(rows)

        print("\nTOTAL ROWS:")
        print(len(rows))

        data = []

        sr_no = 1

        for row in rows:

            print("\n========== CURRENT ROW ==========")
            print(row)

            discharge_commenced = ""

            if row['discharge_commenced']:

                discharge_commenced = (
                    row['discharge_commenced']
                    .strftime('%d-%m-%Y %H:%M')
                )

            discharge_completed = ""

            if row['discharge_completed']:

                discharge_completed = (
                    row['discharge_completed']
                    .strftime('%d-%m-%Y %H:%M')
                )

            current_row = {

                "vsl_no":
                    sr_no,

                "vessel_name":
                    row['vessel_name']
                    or "",

                "cargo":
                    row['cargo_name']
                    or "",

                "bl_qty":
                    int(
                        row['bl_quantity']
                        or 0
                    ),

                "load_port":
                    row['load_port']
                    or "",

                "vcn_call_no":
                    row['vcn_doc_num']
                    or "",

                "discharge_commenced":
                    discharge_commenced,

                "discharge_completed":
                    discharge_completed,

                "time_taken_hrs":
                    float(
                        row['time_taken_hrs']
                        or 0
                    )

            }

            print("\nFORMATTED ROW:")
            print(current_row)

            data.append(current_row)

            sr_no += 1

        print("\n========== FINAL RESPONSE ==========")

        print("\nTYPE OF DATA:")
        print(type(data))

        print("\nFINAL DATA:")
        print(data)

        print("\nTOTAL RECORDS:")
        print(len(data))

        print("\n========== VESSEL DISCHARGE SUMMARY END ==========\n")

        return jsonify({

            "success": True,

            "data": data

        })

    except Exception as e:

        print("\n========== VESSEL DISCHARGE SUMMARY ERROR ==========")

        import traceback

        traceback.print_exc()

        print(str(e))

        print("========== ERROR END ==========\n")

        return jsonify({

            "success": False,

            "message": str(e)

        })

    finally:

        print("\nCLOSING DB CONNECTION")

        cur.close()

        conn.close()


# =========================================================
# FULL DAILY MIS REPORT EXCEL
#
# Rebuilt to mirror the layout of the reference MIS workbook
# (single continuous sheet, no artificial "section title"
# banners like "MONTHLY CARGO REPORT" / "MBC REPORT" — only
# the row captions that genuinely appear in the reference,
# e.g. "MBC'S DISCHARGE COMPLETED").
#
# Anything the current schema has no data source for is left
# blank / "NA", same convention already used in
# barge_status_report() above, with a comment explaining why.
# =========================================================

@bp.route(
    '/api/module/RP01/daily_progress_report_excel',
    methods=['GET']
)
@login_required
def daily_progress_report_excel():

    report_date = request.args.get('report_date')

    if not report_date:
        return jsonify({
            "success": False,
            "message": "Report date required"
        })

    conn = get_db()
    cur = get_cursor(conn)

    try:
        report_dt = datetime.strptime(report_date, "%Y-%m-%d")

        window_end = datetime(
            report_dt.year, report_dt.month, report_dt.day, 8, 0, 0
        )
        window_start = window_end - timedelta(hours=24)
        month_start = report_dt.replace(day=1).date()

        # =====================================================
        # STYLES — matched cell-by-cell against the reference
        # workbook (see header comment above for the mapping).
        # =====================================================

        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        yellow_fill = PatternFill('solid', fgColor='FFFF00')
        white_fill = PatternFill('solid', fgColor='FFFFFF')

        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Main title bar: "DAILY REPORT OF JSW DHARAMTAR PORT OPERATIONS"
        title_font = Font(name='Calibri', bold=True, size=16)

        # Plain row captions ("Vessel Name", dates, barge-status
        # labels) — NOT bold, NOT red, white fill.
        caption_font = Font(name='Calibri', bold=False, size=16)

        # Bold black values inside the vessel grid (vessel name,
        # cargo, quantities, timestamps).
        value_font = Font(name='Calibri', bold=False, size=12)

        # Section headers that open a new block: "Date / Day",
        # "Mother Vessel Name", "Remarks:", "MBC'S DISCHARGE
        # COMPLETED", "Vessel Completed for The Month …" — bold
        # RED text on YELLOW fill.
        section_font = Font(name='Calibri', bold=True, size=16, color='FF0000')

        # Delay-row text ("Bad Weather : ...", "Want of Barge : ...")
        # sits on the same yellow fill as a section header, but the
        # text itself should be BLACK, not red.
        delay_font = Font(name='Calibri', bold=True, size=16, color='000000')

        # Row height (in points) used for the delay row and for the
        # whole "Mother Vessel Name" -> barge-status -> "Remarks:"
        # block below it, so multi-line delay text and long barge
        # name lists aren't clipped.
        TALL_ROW_HEIGHT = 60

        # Column headers inside data tables (MBC Name, Cargo,
        # Source, Qty …) — bold black, no fill.
        header_font = Font(name='Calibri', bold=True, size=13)

        # Ordinary table data rows.
        data_font = Font(name='Calibri', bold=False, size=11)

        def caption(row, col, text, span=1, font=caption_font, fill=white_fill, align=left):
            """Plain (non-highlighted) row/column caption."""
            if span > 1:
                ws.merge_cells(start_row=row, start_column=col,
                                end_row=row, end_column=col + span - 1)
            c = ws.cell(row, col, text)
            c.font = font
            c.alignment = align
            c.fill = fill
            c.border = border
            if span > 1:
                for extra in range(col + 1, col + span):
                    ec = ws.cell(row, extra)
                    ec.fill = fill
                    ec.border = border
            return c

        def section(row, col, text, span=1, align=left):
            """Bold red-on-yellow section header."""
            return caption(row, col, text, span=span, font=section_font,
                            fill=yellow_fill, align=align)

        def value(row, col, val, span=1, font=value_font, align=center):
            return caption(row, col, val, span=span, font=font,
                            fill=white_fill, align=align)

        def header(row, col, text, span=1, align=center):
            return caption(row, col, text, span=span, font=header_font,
                            fill=yellow_fill, align=align)

        def data(row, col, val, align=center, span=1):
            """Data cell. When span>1 the caller is expected to have
            already merged the range (ws.merge_cells) — this applies
            the border/fill to EVERY cell in that merged range, not
            just the top-left one. Excel only draws the border you
            explicitly set on each cell of a merge; leaving the extra
            cells unset is what caused the missing borders under the
            vessel columns."""
            c = ws.cell(row, col, val if val not in (None, '') else '')
            c.font = data_font
            c.alignment = align
            c.fill = white_fill
            c.border = border
            if span > 1:
                for extra in range(col + 1, col + span):
                    ec = ws.cell(row, extra)
                    ec.fill = white_fill
                    ec.border = border
            return c

        def fmt_dt(dt_value):
            if not dt_value:
                return ''
            if hasattr(dt_value, 'strftime'):
                return dt_value.strftime('%d-%m-%Y %H:%M')
            return _parse_flexible(str(dt_value), '%d-%m-%Y %H:%M')

        def _parse_flexible(raw, out_fmt):
            """Parse a stored timestamp string that may or may not include
            seconds/microseconds, and format it for display. Falls back to
            the raw string instead of raising if nothing matches, so a
            single odd value never breaks the whole report."""
            if not raw:
                return ''
            s = str(raw).replace('T', ' ').strip()
            for f in ('%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M'):
                try:
                    return datetime.strptime(s, f).strftime(out_fmt)
                except ValueError:
                    continue
            try:
                return datetime.strptime(s[:16], '%Y-%m-%d %H:%M').strftime(out_fmt)
            except ValueError:
                return s

        wb = Workbook()
        ws = wb.active
        ws.title = "MIS"

        # =====================================================
        # TITLE BAR (matches reference row 1: yellow fill, bold,
        # 16pt, plus "Document No.OPE/0100/F/10" at the far right)
        # =====================================================

        ws.merge_cells('A1:W1')
        t = ws.cell(1, 1, 'DAILY REPORT OF JSW DHARAMTAR PORT OPERATIONS')
        t.font = title_font
        t.alignment = center
        t.fill = yellow_fill
        t.border = border
        for c in range(2, 24):
            cc = ws.cell(1, c)
            cc.fill = yellow_fill
            cc.border = border
        caption(1, 24, 'Document No.OPE/0100/F/10', span=6, font=title_font, align=center)

        # =====================================================
        # REPORT DATE + DOCUMENT CONTROL ROW (reference row 2)
        # =====================================================

        row_no = 2
        caption(row_no, 1, 'Report Date :', font=title_font)
        value(row_no, 16,
              report_dt.strftime('%d/%m/%Y') + ' @ 06:00 Hrs',
              span=6, align=center)
        caption(row_no, 22, 'Rev. No: 00', span=2, font=title_font, align=center)
        caption(row_no, 24, 'Issue No: 01', span=4, font=title_font, align=center)
        caption(row_no, 28, 'Issue Date:10.12.2022', span=6, font=title_font, align=left)

        row_no = 3

        # =====================================================
        # STEP 1: VESSEL LIST FOR THIS REPORT WINDOW
        # (same selection rule used by daily_progress_report_data)
        # =====================================================

        vessel_query = """
        SELECT
            lh.id,
            lh.vessel_name,
            vcd.cargo_name,
            vcd.bl_quantity,
            first_anchor.arrived_mfl,
            first_anchor.arrived_mbpt,
            first_anchor.discharge_started,
            last_anchor.discharge_completed AS discharge_commenced
        FROM ldud_header lh
        LEFT JOIN LATERAL (
            SELECT
                STRING_AGG(cargo_name, ' + ' ORDER BY cargo_name) AS cargo_name,
                SUM(bl_qty) AS bl_quantity
            FROM (
                SELECT TRIM(cargo_name) AS cargo_name,
                       SUM(COALESCE(bl_quantity, 0)) AS bl_qty
                FROM vcn_cargo_declaration
                WHERE vcn_id = lh.vcn_id
                GROUP BY TRIM(cargo_name)
            ) cargo
        ) vcd ON TRUE
        LEFT JOIN LATERAL (
            SELECT
                MIN(CASE WHEN anchorage_name ILIKE '%%PLA%%' THEN anchored END) AS arrived_mfl,
                MIN(CASE WHEN anchorage_name NOT ILIKE '%%PLA%%' THEN anchored END) AS arrived_mbpt,
                MIN(discharge_started) AS discharge_started
            FROM ldud_anchorage
            WHERE ldud_id = lh.id
        ) first_anchor ON TRUE
        LEFT JOIN LATERAL (
            SELECT
                CASE
                    WHEN EXISTS (
                        SELECT 1 FROM ldud_anchorage x
                        WHERE x.ldud_id = lh.id
                        AND x.discharge_started IS NOT NULL
                        AND x.discharge_commenced IS NULL
                    ) THEN NULL
                    ELSE MAX(discharge_commenced)
                END AS discharge_completed
            FROM ldud_anchorage
            WHERE ldud_id = lh.id
        ) last_anchor ON TRUE
        WHERE
            first_anchor.discharge_started IS NOT NULL
            AND first_anchor.discharge_started < %s
            AND (
                last_anchor.discharge_completed IS NULL
                OR last_anchor.discharge_completed >= %s
                OR EXISTS (
                    SELECT 1 FROM ldud_barge_lines b
                    WHERE b.ldud_id = lh.id
                    AND (b.completed_discharge_berth IS NULL OR b.cast_off_berth IS NULL)
                )
            )
        ORDER BY first_anchor.discharge_started, lh.id
        """

        cur.execute(vessel_query, (window_end, window_start))
        vessels = cur.fetchall()
        vessel_ids = [v['id'] for v in vessels]

        # =====================================================
        # VESSEL DAILY GRID
        # Row captions match the reference EXACTLY, including
        # trailing spaces ("B/L QTY.(MT) ", "Discharge Commenced ",
        # "Discharge Completed ").
        # =====================================================

        row_labels = [
            'Vessel Name', 'Cargo / Source', 'B/L QTY.(MT) ',
            'Arrived at MFL', 'Arrived at MbPT', 'Discharge Commenced ',
            'Pre-Berthing Delay', 'Discharge Completed '
        ]

        VESSEL_BLOCK_WIDTH = 3

        grid_start = row_no
        for i, label in enumerate(row_labels):
            caption(grid_start + i, 1, label, span=2)

        col = 3
        vessel_col_map = {}
        for v in vessels:
            vessel_col_map[v['id']] = col

            arrival_for_delay = v['arrived_mfl'] or v['arrived_mbpt']
            pre_delay = ''
            if v['discharge_started'] and arrival_for_delay:
                hrs = round((v['discharge_started'] - arrival_for_delay).total_seconds() / 3600, 2)
                pre_delay = f"{hrs} Hrs"

            discharge_completed_disp = ''
            if v['discharge_commenced'] and window_start <= v['discharge_commenced'] < window_end:
                discharge_completed_disp = fmt_dt(v['discharge_commenced'])

            values = [
                v['vessel_name'] or '',
                v['cargo_name'] or '',
                v['bl_quantity'] or '',
                fmt_dt(v['arrived_mfl']),
                fmt_dt(v['arrived_mbpt']),
                fmt_dt(v['discharge_started']),
                pre_delay,
                discharge_completed_disp,
            ]

            # Vessel name (row 0 of the block) — bold black, white fill,
            # same as every other row (the reference does NOT highlight
            # the vessel-name cell itself, only the true section headers).
            for r_off, v2 in enumerate(values):
                value(grid_start + r_off, col, v2, span=VESSEL_BLOCK_WIDTH)

            col += VESSEL_BLOCK_WIDTH

        row_no = grid_start + len(row_labels) + 1

        # =====================================================
        # STEP 2: DAILY CARGO QUANTITY PER VESSEL
        # =====================================================

        daily_query = """
        SELECT
            lh.id AS ldud_id,
            TRIM(vcd.cargo_name) AS cargo_name,
            TO_CHAR(DATE(lco.start_time), 'DD-MM-YYYY') AS day_label,
            DATE(lco.start_time) AS cargo_date,
            COALESCE(SUM(lco.quantity), 0) AS total_qty
        FROM ldud_header lh
        JOIN vcn_cargo_declaration vcd ON vcd.vcn_id = lh.vcn_id
        LEFT JOIN ldud_vessel_operations lco
            ON lco.ldud_id = lh.id
            AND LOWER(TRIM(lco.cargo_name)) = LOWER(TRIM(vcd.cargo_name))
            AND DATE(lco.start_time) BETWEEN DATE_TRUNC('month', %s::date)::date AND %s::date
        WHERE lh.id = ANY(%s)
        GROUP BY lh.id, TRIM(vcd.cargo_name), DATE(lco.start_time)
        ORDER BY DATE(lco.start_time)
        """

        daily_by_vessel = {v['id']: {} for v in vessels}

        if vessel_ids:
            cur.execute(daily_query, (report_date, report_date, vessel_ids))
            for r in cur.fetchall():
                if not r['day_label']:
                    continue
                day = r['day_label']
                qty = float(r['total_qty'] or 0)
                if not qty:
                    continue
                bucket = daily_by_vessel.setdefault(r['ldud_id'], {})
                entry = bucket.setdefault(day, {'qty': 0.0, 'cargoes': []})
                entry['qty'] += qty
                cname = r['cargo_name'] or ''
                if cname and cname not in entry['cargoes']:
                    entry['cargoes'].append(cname)

        all_days = sorted(
            {d for bucket in daily_by_vessel.values() for d in bucket},
            key=lambda s: datetime.strptime(s, '%d-%m-%Y')
        )

        # =====================================================
        # DATE / DAY MATRIX
        # "Date / Day" is a true section header in the reference
        # (bold red on yellow) — the per-vessel "Cargo / Source /
        # Qty in MT / W/W Hrs." sub-headers are plain bold black.
        # =====================================================

        matrix_header_row = row_no
        section(matrix_header_row, 1, 'Date / Day')
        caption(matrix_header_row, 2, 'Total MV Disch', font=header_font)

        for v in vessels:
            c = vessel_col_map[v['id']]
            header(matrix_header_row, c, 'Cargo / Source')
            header(matrix_header_row, c + 1, 'Qty in MT')
            header(matrix_header_row, c + 2, 'W/W Hrs.')

        matrix_row = matrix_header_row + 1
        vessel_totals = {v['id']: 0.0 for v in vessels}

        for i, day in enumerate(all_days, start=1):
            day_total = 0.0
            caption(matrix_row, 1, f"{day}/{i:02d}", align=left)

            for v in vessels:
                c = vessel_col_map[v['id']]
                entry = daily_by_vessel.get(v['id'], {}).get(day)
                qty = entry['qty'] if entry else 0
                cargo_text = ' + '.join(entry['cargoes']) if entry else ''
                if qty:
                    data(matrix_row, c, cargo_text, align=left)
                    data(matrix_row, c + 1, qty)
                    day_total += qty
                    vessel_totals[v['id']] += qty
                else:
                    data(matrix_row, c, '')
                    data(matrix_row, c + 1, '')
                # W/W Hrs — not tracked in the schema, left blank
                data(matrix_row, c + 2, '')

            data(matrix_row, 2, day_total)
            matrix_row += 1

        # TOTAL row
        caption(matrix_row, 2, 'TOTAL', font=header_font)
        for v in vessels:
            c = vessel_col_map[v['id']]
            header(matrix_row, c, 'TOTAL')
            data(matrix_row, c + 1, vessel_totals[v['id']])
            data(matrix_row, c + 2, '')
        matrix_row += 1

        # BALANCE ON BOARD row
        caption(matrix_row, 2, 'Balance on Board', font=header_font)
        for v in vessels:
            c = vessel_col_map[v['id']]
            bl_qty = float(v['bl_quantity'] or 0)
            balance = max(bl_qty - vessel_totals[v['id']], 0)
            header(matrix_row, c, 'Balance on Board')
            data(matrix_row, c + 1, balance)
            data(matrix_row, c + 2, '')
        matrix_row += 1

        # AVG DISCHARGE RATE PWWD row — no working-hours data source
        # in the schema, so this cannot be computed; left as NA.
        caption(matrix_row, 2, 'Avg Discharge Rate PWWD', font=header_font)
        for v in vessels:
            c = vessel_col_map[v['id']]
            header(matrix_row, c, 'Avg Discharge Rate PWWD')
            data(matrix_row, c + 1, 'NA')
            data(matrix_row, c + 2, '')
        matrix_row += 1

        # HOOKS AVAILABLE row — no data source in schema; left as NA.
        caption(matrix_row, 2, 'Hooks Available', font=header_font)
        for v in vessels:
            c = vessel_col_map[v['id']]
            header(matrix_row, c, 'Hooks Available')
            data(matrix_row, c + 1, 'NA')
            data(matrix_row, c + 2, '')
        matrix_row += 1

        # =====================================================
        # DELAYS (per vessel) — this whole row is a section
        # header row in the reference: bold red on yellow.
        # =====================================================

        delay_query = """
        SELECT
            ld.ldud_id,
            COALESCE(ld.delay_name, '') AS delay_name,
            COALESCE(ld.crane_number, 'All') AS crane_number,
            SUM(COALESCE(ld.total_time_mins, 0)) AS total_mins
        FROM ldud_delays ld
        WHERE ld.ldud_id = ANY(%s)
        AND TO_TIMESTAMP(ld.start_datetime, 'YYYY-MM-DD"T"HH24:MI') < %s
        AND (
            ld.end_datetime IS NULL
            OR TO_TIMESTAMP(ld.end_datetime, 'YYYY-MM-DD"T"HH24:MI') >= %s
        )
        GROUP BY ld.ldud_id, ld.delay_name, ld.crane_number
        ORDER BY ld.ldud_id, ld.delay_name, ld.crane_number
        """

        delay_text_by_vessel = {v['id']: [] for v in vessels}

        if vessel_ids:
            cur.execute(delay_query, (vessel_ids, window_end, window_start))
            grouped = {}
            for r in cur.fetchall():
                key = (r['ldud_id'], r['delay_name'])
                grouped.setdefault(key, []).append(
                    f"Crane {r['crane_number']} : "
                    f"{int(r['total_mins']) // 60} Hrs {int(r['total_mins']) % 60} Mins"
                )
            for (ldud_id, delay_name), parts in grouped.items():
                if delay_name:
                    delay_text_by_vessel[ldud_id].append(
                        f"{delay_name} :  " + "  ".join(parts) + "."
                    )

        for v in vessels:
            c = vessel_col_map[v['id']]
            ws.merge_cells(start_row=matrix_row, start_column=c,
                            end_row=matrix_row, end_column=c + VESSEL_BLOCK_WIDTH - 1)
            text = "\n".join(delay_text_by_vessel[v['id']]) or 'NA'
            cell = ws.cell(matrix_row, c, text)
            cell.font = delay_font
            cell.alignment = left
            cell.fill = yellow_fill
            cell.border = border
            for extra in range(c + 1, c + VESSEL_BLOCK_WIDTH):
                ec = ws.cell(matrix_row, extra)
                ec.fill = yellow_fill
                ec.border = border

        # Taller row so multi-line delay text isn't clipped.
        ws.row_dimensions[matrix_row].height = TALL_ROW_HEIGHT

        matrix_row += 1

        row_no = matrix_row

        # =====================================================
        # MOTHER VESSEL NAME ROW — section header (red/yellow)
        # =====================================================

        # Start of the "tall block" (Mother Vessel Name -> last barge
        # status row) whose row height gets bumped to match the delay
        # row above, once we know where it ends.
        tall_block_start = row_no

        section(row_no, 1, 'Mother Vessel Name', span=2)
        for i, v in enumerate(vessels, start=1):
            c = vessel_col_map[v['id']]
            data(row_no, c, f'Vessel # {i} :', align=left)
            ws.merge_cells(start_row=row_no, start_column=c + 1,
                            end_row=row_no, end_column=c + VESSEL_BLOCK_WIDTH - 1)
            # span=VESSEL_BLOCK_WIDTH-1 so the border/fill is applied to
            # EVERY cell of the merged range, not just its first cell —
            # this is what was leaving the cells under the vessel name
            # borderless.
            value(row_no, c + 1, v['vessel_name'] or '', span=VESSEL_BLOCK_WIDTH - 1, align=left)
        row_no += 1

        # =====================================================
        # BARGE STATUS — captions copied verbatim from the
        # reference (note exact spacing/hyphenation quirks below
        # match the source file character-for-character). These
        # are plain captions, NOT section headers, in the
        # reference — white fill, not bold, not red.
        # =====================================================

        barge_status_rows = [
            ('At Jetty -Under Discharge / Loading', 'at_jetty'),
            ('At Jetty - Waiting for Discharge', 'waiting_discharge'),
            ('At R-19 - Waiting (Loaded)', None),
            ('In transit - From MV/Gull to Jetty (Loaded )', None),
            ('At Gull - Waiting (Loaded)', 'at_gull_loaded'),
            ('Under Loading at MV', 'under_loading'),
            ('Waiting for loading', 'waiting_loading'),
            ('Waiting at Jetty- Empty', 'waiting_empty_jetty'),
            ('Empty at Gull / R-19', None),
            ('In transit - from Jetty to  MV', 'in_transit_jetty_to_mv'),
            ('Breakdown / Off Hired/ Coastal', 'breakdown'),
        ]

        barge_stats = {
            lid: {
                'at_jetty': [], 'waiting_discharge': [], 'waiting_empty_jetty': [],
                'at_gull_loaded': [], 'under_loading': [], 'waiting_loading': [],
                'in_transit_jetty_to_mv': [], 'breakdown': [],
            }
            for lid in vessel_ids
        }

        if vessel_ids:
            ws_str = window_start.strftime('%Y-%m-%dT%H:%M')
            cur.execute("""
                SELECT
                    b.ldud_id, b.barge_name,
                    b.along_side_vessel, b.commenced_loading, b.completed_loading,
                    b.cast_off_mv, b.along_side_berth, b.commence_discharge_berth,
                    b.completed_discharge_berth, b.cast_off_berth, b.cast_off_port
                FROM ldud_barge_lines b
                WHERE b.ldud_id = ANY(%s)
                AND (b.cast_off_port IS NULL OR b.cast_off_port > %s)
            """, (vessel_ids, ws_str))

            for r in cur.fetchall():
                bn = (r['barge_name'] or '').strip()
                if not bn:
                    continue
                if r['cast_off_port']:
                    status = None
                elif r['completed_discharge_berth'] and not r['cast_off_berth']:
                    status = 'waiting_empty_jetty'
                elif r['commence_discharge_berth'] and not r['cast_off_berth']:
                    status = 'at_jetty'
                elif r['along_side_berth'] and not r['commence_discharge_berth']:
                    status = 'waiting_discharge'
                elif r['cast_off_mv'] and not r['along_side_berth']:
                    status = 'at_gull_loaded'
                elif r['commenced_loading'] and not r['completed_loading']:
                    status = 'under_loading'
                elif r['along_side_vessel'] and not r['commenced_loading']:
                    status = 'waiting_loading'
                else:
                    status = None
                if status:
                    barge_stats[r['ldud_id']][status].append(bn)

        for label, key in barge_status_rows:
            caption(row_no, 1, label, span=2)
            for v in vessels:
                c = vessel_col_map[v['id']]
                ws.merge_cells(start_row=row_no, start_column=c,
                                end_row=row_no, end_column=c + VESSEL_BLOCK_WIDTH - 1)
                barges = None if key is None else ' + '.join(barge_stats.get(v['id'], {}).get(key, []))
                # Never leave the cell truly blank — show "NA" when
                # there's nothing to report for that status.
                text = barges if barges else 'NA'
                # span=VESSEL_BLOCK_WIDTH fixes the same missing-border
                # issue as above for every barge-status row.
                data(row_no, c, text, align=left, span=VESSEL_BLOCK_WIDTH)
            row_no += 1

        # End of the tall block: bump every row from "Mother Vessel
        # Name" through the last barge-status row to the same height
        # as the delay row above, so long barge lists aren't clipped.
        tall_block_end = row_no - 1
        for r in range(tall_block_start, tall_block_end + 1):
            ws.row_dimensions[r].height = TALL_ROW_HEIGHT

        # "Remarks:" — section header (red/yellow), exact trailing
        # space kept to match the reference.
        # Merge the Remarks row across all columns (change 6 to your last column)
        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=2)   # A:B
        ws.merge_cells(start_row=row_no, start_column=3, end_row=row_no, end_column=5)   # C:E
        ws.merge_cells(start_row=row_no, start_column=6, end_row=row_no, end_column=8)   # F:H
        ws.merge_cells(start_row=row_no, start_column=9, end_row=row_no, end_column=11)  # I:K
        ws.merge_cells(start_row=row_no, start_column=12, end_row=row_no, end_column=14) # L:N

        # First merged cell
        cell = ws.cell(row=row_no, column=1)
        cell.value = "Remarks:"
        cell.font = header_font
        cell.fill = yellow_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # Apply border to all cells
        for col in range(1, 15):   # Columns A to N
            ws.cell(row=row_no, column=col).border = thin

        row_no += 2

        # =====================================================
        # CARGO TYPE WISE DISCHARGE
        # Rebuilt on lueu_lines (live) + rp01_historical_lueu
        # (historic, April only) — the old ldud_barge_lines /
        # discharge_quantity / completed_loading columns do not
        # exist on this schema.
        #
        # Columns are now cargo_type + cargo_category (NOT
        # cargo_name) to match the barge_discharge_report API.
        # Equipment names are normalized the same way (legacy
        # "Barge Unloader 1/2" labels merged into "BUL-01/02").
        #
        # Day total   = entry_date = data_date (report_date - 1)
        # Month total = 1st of that month -> data_date  (live only)
        # FY total    = historic (Apr 1 - Apr 30, rp01_historical_lueu)
        #               + live (May 1 -> data_date, lueu_lines)
        # =====================================================

        data_date_obj = report_dt - timedelta(days=1)
        data_date = data_date_obj.strftime("%Y-%m-%d")
        cargo_month_start = data_date_obj.replace(day=1).strftime("%Y-%m-%d")

        fy_start_year = data_date_obj.year if data_date_obj.month >= 4 else data_date_obj.year - 1
        fy_start = f"{fy_start_year}-04-01"
        historic_cutoff = f"{fy_start_year}-04-30"
        live_start = f"{fy_start_year}-05-01"

        EQUIP_EXPR = """
            CASE TRIM(equipment_name)
                WHEN 'Barge Unloader 1' THEN 'BUL-01'
                WHEN 'Barge Unloader 2' THEN 'BUL-02'
                ELSE COALESCE(TRIM(equipment_name), 'Others')
            END
        """

        def cat_key(ctype, ccat):
            return f"{ctype}||{ccat}"

        # 0) FULL master list of cargo_type -> cargo_category
        cur.execute("""
            SELECT
                COALESCE(vc.cargo_type, 'Others') AS cargo_type,
                COALESCE(vc.cargo_category, 'Others') AS cargo_category
            FROM vessel_cargo vc
            GROUP BY vc.cargo_type, vc.cargo_category
            ORDER BY vc.cargo_type, vc.cargo_category
        """)
        master_cargo_rows = cur.fetchall()

        # 0b) FULL master list of equipment (live table only)
        cur.execute(f"""
            SELECT DISTINCT {EQUIP_EXPR} AS equipment
            FROM lueu_lines
            WHERE is_deleted IS NOT TRUE
            ORDER BY 1
        """)
        master_equipment_rows = cur.fetchall()

        # 1) Day-level matrix: equipment x (cargo_type, cargo_category), entry_date = data_date ONLY (live only)
        cur.execute(f"""
            SELECT
                COALESCE(vc.cargo_type, 'Others') AS cargo_type,
                COALESCE(vc.cargo_category, 'Others') AS cargo_category,
                {EQUIP_EXPR.replace('equipment_name', 'lbl.equipment_name')} AS equipment,
                SUM(lbl.quantity) AS day_qty
            FROM lueu_lines lbl
            LEFT JOIN vessel_cargo vc
                ON LOWER(TRIM(vc.cargo_name)) = LOWER(TRIM(lbl.cargo_name))
            WHERE lbl.is_deleted IS NOT TRUE
              AND lbl.quantity IS NOT NULL
              AND lbl.entry_date::date = %(data_date)s::date
            GROUP BY vc.cargo_type, vc.cargo_category, {EQUIP_EXPR.replace('equipment_name', 'lbl.equipment_name')}
        """, {"data_date": data_date})
        cargo_day_rows = cur.fetchall()

        # 2a) Historic total (rp01_historical_lueu): fy_start -> historic_cutoff (April 1 - April 30)
        cur.execute("""
            SELECT
                COALESCE(vc.cargo_type, 'Others') AS cargo_type,
                COALESCE(vc.cargo_category, 'Others') AS cargo_category,
                SUM(COALESCE(h.quantity, 0)) AS historic_total
            FROM rp01_historical_lueu h
            LEFT JOIN vessel_cargo vc
                ON LOWER(TRIM(vc.cargo_name)) = LOWER(TRIM(h.cargo_name))
            WHERE h.quantity IS NOT NULL
              AND h.entry_date BETWEEN %(fy_start)s::date AND %(historic_cutoff)s::date
            GROUP BY vc.cargo_type, vc.cargo_category
        """, {"fy_start": fy_start, "historic_cutoff": historic_cutoff})
        cargo_historic_rows = cur.fetchall()

        # 2b) Live total (lueu_lines): month_total, live_fy_total, live_cumulative_total
        cur.execute("""
            SELECT
                COALESCE(vc.cargo_type, 'Others') AS cargo_type,
                COALESCE(vc.cargo_category, 'Others') AS cargo_category,
                SUM(COALESCE(lbl.quantity, 0)) AS live_fy_total,
                SUM(CASE WHEN lbl.entry_date::date BETWEEN %(month_start)s::date AND %(data_date)s::date
                    THEN COALESCE(lbl.quantity, 0) ELSE 0 END) AS month_total
            FROM lueu_lines lbl
            LEFT JOIN vessel_cargo vc
                ON LOWER(TRIM(vc.cargo_name)) = LOWER(TRIM(lbl.cargo_name))
            WHERE lbl.is_deleted IS NOT TRUE
              AND lbl.quantity IS NOT NULL
              AND lbl.entry_date::date BETWEEN %(live_start)s::date AND %(data_date)s::date
            GROUP BY vc.cargo_type, vc.cargo_category
        """, {"live_start": live_start, "data_date": data_date, "month_start": cargo_month_start})
        cargo_live_rows = cur.fetchall()

        # ---- Build FULL cargo hierarchy: cargo_type -> [cargo_categories] ----
        cargo_hierarchy = {}
        for r in master_cargo_rows:
            ctype = r['cargo_type']
            ccat = r['cargo_category']
            cargo_hierarchy.setdefault(ctype, [])
            if ccat not in cargo_hierarchy[ctype]:
                cargo_hierarchy[ctype].append(ccat)

        all_keys = [cat_key(r['cargo_type'], r['cargo_category']) for r in master_cargo_rows]

        # ---- Seed equipment_rows with the FULL equipment list (zeros by default) ----
        equipment_rows = {}
        for r in master_equipment_rows:
            equipment_rows[r['equipment']] = {k: 0 for k in all_keys}

        for r in cargo_day_rows:
            equip = r['equipment']
            key = cat_key(r['cargo_type'], r['cargo_category'])
            equipment_rows.setdefault(equip, {k: 0 for k in all_keys})
            equipment_rows[equip][key] = int(r['day_qty'] or 0)

        # ---- Historic totals (April) per category ----
        historic_totals = {}
        for r in cargo_historic_rows:
            key = cat_key(r['cargo_type'], r['cargo_category'])
            historic_totals[key] = int(r['historic_total'] or 0)

        # ---- Live totals (May onward) merged with historic to form FY total ----
        month_totals = {}
        year_totals = {}
        for r in cargo_live_rows:
            key = cat_key(r['cargo_type'], r['cargo_category'])
            month_totals[key] = int(r['month_total'] or 0)
            live_fy = int(r['live_fy_total'] or 0)
            hist = historic_totals.get(key, 0)
            year_totals[key] = hist + live_fy

        # Categories that ONLY had historic (April) activity, nothing live yet
        for key, hist_val in historic_totals.items():
            if key not in year_totals:
                year_totals[key] = hist_val
                month_totals.setdefault(key, 0)

        # cargo_columns is now a flat list of (type, category) keys, in
        # cargo_hierarchy order, replacing the old cargo_name list.
        cargo_columns = [
            cat_key(ctype, ccat)
            for ctype, ccats in cargo_hierarchy.items()
            for ccat in ccats
        ]
        cargo_column_labels = {
            cat_key(ctype, ccat): ccat
            for ctype, ccats in cargo_hierarchy.items()
            for ccat in ccats
        }

        if cargo_columns:
            type_row = row_no
            cname_row = row_no + 1
            cur_col = 3
            for ctype, ccats in cargo_hierarchy.items():
                span = len(ccats)
                if span:
                    # Category header ("IBRM", "CBRM", "Fluxes" …) —
                    # bold on yellow, black text (not red — this is a
                    # column-group header, not a section header).
                    caption(type_row, cur_col, ctype, span=span,
                            font=header_font, fill=yellow_fill)
                for ccat in ccats:
                    header(cname_row, cur_col, ccat)
                    cur_col += 1

            row_no = cname_row + 1
            for equipment, qty_map in equipment_rows.items():
                caption(row_no, 1, equipment, span=2)
                cur_col = 3
                for key in cargo_columns:
                    data(row_no, cur_col, qty_map.get(key, 0))
                    cur_col += 1
                row_no += 1

            caption(row_no, 1, 'Total For The Day', span=2, font=header_font)
            cur_col = 3
            for key in cargo_columns:
                header(row_no, cur_col, sum(qm.get(key, 0) for qm in equipment_rows.values()))
                cur_col += 1
            row_no += 1

            caption(row_no, 1, 'Total Receipts for the Month', span=2, font=header_font)
            cur_col = 3
            for key in cargo_columns:
                header(row_no, cur_col, month_totals.get(key, 0))
                cur_col += 1
            row_no += 1

            fy_label = f"FY {str(fy_start_year)[-2:]}-{str(fy_start_year + 1)[-2:]}"
            caption(row_no, 1, f'Total Receipts {fy_label}', span=2, font=header_font)
            cur_col = 3
            for key in cargo_columns:
                header(row_no, cur_col, year_totals.get(key, 0))
                cur_col += 1
            row_no += 2
        else:
            caption(row_no, 1, 'No barge discharge recorded for this window', span=2)
            row_no += 2

        # =====================================================
        # MBC'S DISCHARGE COMPLETED — section header, exact
        # trailing space kept ("...COMPLETED ").
        # =====================================================

        cur.execute("""
            SELECT
                mh.mbc_name, mh.cargo_name, mh.bl_quantity, mh.load_port,
                dpl.vessel_arrival_port, dpl.unloading_commenced,
                dpl.cleaning_commenced, dpl.unloading_completed,
                dpl.sailed_out_load_port
            FROM mbc_header mh
            LEFT JOIN mbc_discharge_port_lines dpl ON dpl.mbc_id = mh.id
            WHERE
                NULLIF(TRIM(dpl.unloading_completed), '') IS NOT NULL
                AND NULLIF(TRIM(dpl.unloading_completed), '')::timestamp >= %s
                AND NULLIF(TRIM(dpl.unloading_completed), '')::timestamp < %s
            ORDER BY NULLIF(TRIM(dpl.unloading_completed), '')::timestamp
        """, (window_start, window_end))

        mbc_completed = cur.fetchall()

        def str_to_disp(v):
            return _parse_flexible(v, '%d-%m-%Y : %H:%M')

        section(row_no, 1, "MBC'S DISCHARGE COMPLETED ", span=8)
        row_no += 1
        headers = ['MBC Name', 'Cargo', 'Source', 'Qty', 'Arrived at Jetty',
                   'Unloading / Loading   Commence', 'Cleaning Start',
                   'Unloading / Loading   Completed', 'C/off from Jetty']
        for i, h in enumerate(headers):
            header(row_no, 1 + i, h)
        row_no += 1
        if not mbc_completed:
            ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=len(headers))
            caption(row_no, 1, 'No records for this window', span=len(headers))
            row_no += 1
        for r in mbc_completed:
            vals_row = [
                r['mbc_name'] or '', r['cargo_name'] or '', r['load_port'] or '',
                int(r['bl_quantity'] or 0), str_to_disp(r['vessel_arrival_port']),
                str_to_disp(r['unloading_commenced']), str_to_disp(r['cleaning_commenced']),
                str_to_disp(r['unloading_completed']),
                r['sailed_out_load_port'].strftime('%d-%m-%Y : %H:%M') if r['sailed_out_load_port'] else '',
            ]
            for i, v2 in enumerate(vals_row):
                data(row_no, 1 + i, v2)
            row_no += 1
        row_no += 1

        # =====================================================
        # VESSELS / MBC — ARRIVED, EXPECTED (row-based tables)
        # Titles are plain captions in the reference for these
        # sub-tables (not the red/yellow section style).
        # =====================================================

        def simple_table(title, rows, headers, keys):
            nonlocal row_no
            caption(row_no, 1, title, span=len(headers), font=header_font)
            row_no += 1
            for i, h in enumerate(headers):
                header(row_no, 1 + i, h)
            row_no += 1
            if not rows:
                ws.merge_cells(start_row=row_no, start_column=1,
                                end_row=row_no, end_column=len(headers))
                caption(row_no, 1, 'No records for this window', span=len(headers))
                row_no += 1
            for idx, r in enumerate(rows, start=1):
                data(row_no, 1, idx)
                for i, k in enumerate(keys):
                    data(row_no, 2 + i, r.get(k, ''), align=left)
                row_no += 1
            row_no += 1

        cur.execute("""
            SELECT
                lh.vessel_name,
                STRING_AGG(DISTINCT TRIM(vcd.cargo_name), ', ' ORDER BY TRIM(vcd.cargo_name)) AS cargo,
                SUM(COALESCE(vcd.bl_quantity, 0)) AS bl_quantity,
                vh.load_port, lh.nor_accepted
            FROM ldud_header lh
            LEFT JOIN vcn_header vh ON vh.id = lh.vcn_id
            LEFT JOIN vcn_cargo_declaration vcd ON vcd.vcn_id = lh.vcn_id
            WHERE
                NULLIF(TRIM(lh.nor_accepted), '') IS NOT NULL
                AND DATE(NULLIF(TRIM(lh.nor_accepted), '')::timestamp)
                    BETWEEN (%s::date - INTERVAL '1 day') AND %s::date
            GROUP BY lh.id, lh.vessel_name, vh.load_port, lh.nor_accepted
            ORDER BY NULLIF(TRIM(lh.nor_accepted), '')::timestamp
        """, (report_date, report_date))

        arrived_rows = [{
            "vessel_name": r["vessel_name"] or "",
            "cargo": r["cargo"] or "",
            "bl_qty": int(r["bl_quantity"] or 0),
            "load_port": r["load_port"] or "",
            "arrived_mumbai": _parse_flexible(r["nor_accepted"], "%d-%m-%Y %H:%M"),
        } for r in cur.fetchall()]

        simple_table(
            "VESSELS ARRIVED AT MUMBAI",
            arrived_rows,
            ["SR.NO.", "M.Vessel Name", "Cargo", "B/L Qty. (MT)", "Load  Port", "Arrived @ Mumbai"],
            ["vessel_name", "cargo", "bl_qty", "load_port", "arrived_mumbai"],
        )

        cur.execute("""
            SELECT mh.mbc_name, mh.cargo_name, mh.bl_quantity, mh.load_port,
                   dpl.vessel_arrival_port
            FROM mbc_header mh
            LEFT JOIN mbc_discharge_port_lines dpl ON dpl.mbc_id = mh.id
            WHERE
                NULLIF(TRIM(dpl.vessel_arrival_port), '') IS NOT NULL
                AND DATE(NULLIF(TRIM(dpl.vessel_arrival_port), '')::timestamp)
                    BETWEEN (%s::date - INTERVAL '1 day') AND %s::date
            ORDER BY NULLIF(TRIM(dpl.vessel_arrival_port), '')::timestamp
        """, (report_date, report_date))

        mbc_arrived_rows = [{
            "mbc_name": r["mbc_name"] or "",
            "cargo": r["cargo_name"] or "",
            "bl_qty": int(r["bl_quantity"] or 0),
            "load_port": r["load_port"] or "",
            "arrived_dharamtar": _parse_flexible(r["vessel_arrival_port"], "%d-%m-%Y %H:%M"),
        } for r in cur.fetchall()]

        simple_table(
            "MBC ARRIVED AT DHARAMTAR",
            mbc_arrived_rows,
            ["SR.NO.", "MBC Name", "Cargo", "B/L Qty. (MT)", "Load  Port", "Arrived  @ Dharamtar"],
            ["mbc_name", "cargo", "bl_qty", "load_port", "arrived_dharamtar"],
        )

        cur.execute("""
            SELECT vn.id, vh.vessel_name,
                   STRING_AGG(DISTINCT TRIM(vcd.cargo_name), ', ' ORDER BY TRIM(vcd.cargo_name)) AS cargo,
                   SUM(COALESCE(vcd.bl_quantity, 0)) AS bl_quantity,
                   vh.load_port, vn.eta
            FROM vcn_nominations vn
            LEFT JOIN vcn_header vh ON vh.id = vn.vcn_id
            LEFT JOIN vcn_cargo_declaration vcd ON vcd.vcn_id = vh.id
            WHERE DATE(vn.eta) >= %s
            GROUP BY vn.id, vh.vessel_name, vh.load_port, vn.eta
            ORDER BY vn.eta
        """, (report_date,))

        upcoming_rows = [{
            "vessel_name": r["vessel_name"] or "",
            "cargo": r["cargo"] or "",
            "bl_qty": int(r["bl_quantity"] or 0),
            "load_port": r["load_port"] or "",
            "eta_mumbai": _parse_flexible(r["eta"], "%d-%m-%Y %H:%M"),
        } for r in cur.fetchall()]

        simple_table(
            "VESSELS EXPECTED AT MUMBAI",
            upcoming_rows,
            ["SR.NO.","M.Vessel Name", "Cargo", "B/L Qty. (MT)", "Load  Port", "ETA @ Mumbai"],
            ["vessel_name", "cargo", "bl_qty", "load_port", "eta_mumbai"],
        )

        cur.execute("""
            SELECT mh.mbc_name, mh.cargo_name, mh.bl_quantity, mh.load_port,
                   dpl.arrival_gull_island
            FROM mbc_header mh
            LEFT JOIN mbc_discharge_port_lines dpl ON dpl.mbc_id = mh.id
            WHERE
                NULLIF(TRIM(dpl.arrival_gull_island), '') IS NOT NULL
                AND DATE(NULLIF(TRIM(dpl.arrival_gull_island), '')::timestamp)
                    BETWEEN (%s::date - INTERVAL '1 day') AND %s::date
            ORDER BY NULLIF(TRIM(dpl.arrival_gull_island), '')::timestamp
        """, (report_date, report_date))

        mbc_expected_rows = [{
            "mbc_name": r["mbc_name"] or "",
            "cargo": r["cargo_name"] or "",
            "bl_qty": int(r["bl_quantity"] or 0),
            "load_port": r["load_port"] or "",
            "eta_mumbai": _parse_flexible(r["arrival_gull_island"], "%d-%m-%Y %H:%M"),
        } for r in cur.fetchall()]

        simple_table(
            "MBC EXPECTED AT MUMBAI",
            mbc_expected_rows,
            ["SR.NO.", "MBC Name", "Cargo", "B/L Qty. (MT)", "Load  Port", "ETA @ Mumbai"],
            ["mbc_name", "cargo", "bl_qty", "load_port", "eta_mumbai"],
        )

        # =====================================================
        # TIDE TABLE
        # =====================================================

        cur.execute("""
            SELECT tide_datetime, tide_meters
            FROM tide_master
            WHERE
                NULLIF(TRIM(tide_datetime), '') IS NOT NULL
                AND DATE(NULLIF(TRIM(tide_datetime), '')::timestamp) = %s
            ORDER BY NULLIF(TRIM(tide_datetime), '')::timestamp
        """, (report_date,))

        tide_rows = cur.fetchall()

        caption(row_no, 1, 'Tide Table', span=3, font=header_font)
        row_no += 1
        header(row_no, 1, 'Type'); header(row_no, 2, 'Time'); header(row_no, 3, 'Mtrs')
        row_no += 1

        if tide_rows:
            heights = [float(r['tide_meters'] or 0) for r in tide_rows]
            current = 'HW' if len(heights) >= 2 and heights[0] > heights[1] else 'HW'
            for r in tide_rows:
                time_disp = _parse_flexible(r['tide_datetime'], '%H:%M')
                data(row_no, 1, current, align=left)
                data(row_no, 2, time_disp)
                data(row_no, 3, float(r['tide_meters'] or 0))
                current = 'LW' if current == 'HW' else 'HW'
                row_no += 1
        else:
            ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=3)
            caption(row_no, 1, 'No tide data for this date', span=3)
            row_no += 1
        row_no += 1

        # =====================================================
        # VESSELS COMPLETED FOR THE MONTH — section header
        # =====================================================

        cur.execute("""
            SELECT
                lh.vessel_name,
                STRING_AGG(DISTINCT TRIM(vcd.cargo_name), ', ' ORDER BY TRIM(vcd.cargo_name)) AS cargo_name,
                SUM(COALESCE(vcd.bl_quantity, 0)) AS bl_quantity,
                vh.load_port, vh.vcn_doc_num,
                MIN(la.discharge_started) AS discharge_commenced,
                MAX(la.discharge_commenced) AS discharge_completed,
                ROUND(EXTRACT(EPOCH FROM (
                    MAX(la.discharge_commenced) - MIN(la.discharge_started)
                )) / 3600, 2) AS time_taken_hrs
            FROM ldud_header lh
            LEFT JOIN vcn_cargo_declaration vcd ON lh.vcn_id = vcd.vcn_id
            LEFT JOIN vcn_header vh ON vh.id = lh.vcn_id
            LEFT JOIN ldud_anchorage la ON la.ldud_id = lh.id
            WHERE
                la.discharge_commenced IS NOT NULL
                AND DATE(la.discharge_commenced) BETWEEN %s AND %s
            GROUP BY lh.id, lh.vessel_name, vh.load_port, vh.vcn_doc_num
            ORDER BY MAX(la.discharge_commenced)
        """, (month_start, report_date))

        completed_rows = [{
            "vessel_name": r["vessel_name"] or "",
            "cargo": r["cargo_name"] or "",
            "bl_qty": int(r["bl_quantity"] or 0),
            "load_port": r["load_port"] or "",
            "vcn_call_no": r["vcn_doc_num"] or "",
            "discharge_commenced": fmt_dt(r["discharge_commenced"]),
            "discharge_completed": fmt_dt(r["discharge_completed"]),
            "time_taken_hrs": float(r["time_taken_hrs"] or 0),
        } for r in cur.fetchall() if r["discharge_completed"]]

        month_label = report_dt.strftime('%B')

        section(row_no, 1, f'Vessel Completed for The Month {month_label}', span=9)
        row_no += 1
        headers = ['SR.NO.', 'M.Vessel Name', 'Cargo ', 'B/L Qty. (MT)', 'Load  Port',
                   'Discharge Commence', 'Discharge Completed', 'Time Taken (Hrs)']
        for i, h in enumerate(headers):
            header(row_no, 1 + i, h)
        row_no += 1
        if not completed_rows:
            ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=len(headers))
            caption(row_no, 1, 'No records for this window', span=len(headers))
            row_no += 1
        for idx, r in enumerate(completed_rows, start=1):
            data(row_no, 1, idx)
            data(row_no, 2, r['vessel_name'], align=left)
            data(row_no, 3, r['cargo'], align=left)
            data(row_no, 4, r['bl_qty'])
            data(row_no, 5, r['load_port'], align=left)
            data(row_no, 6, r['discharge_commenced'])
            data(row_no, 7, r['discharge_completed'])
            data(row_no, 8, r['time_taken_hrs'])
            row_no += 1

        # =====================================================
        # COLUMN WIDTHS — match the reference's actual widths for
        # the fixed left-hand columns; remaining vessel columns
        # get a consistent, readable width.
        # =====================================================

        fixed_widths = {
            'A': 28.33, 'B': 17.44, 'C': 34.66, 'D': 17.55, 'E': 21.0,
            'F': 22.89, 'G': 19.55, 'H': 26.66, 'I': 22.33,
        }
        for letter, w in fixed_widths.items():
            ws.column_dimensions[letter].width = w
        for c in range(10, 60):
            ws.column_dimensions[get_column_letter(c)].width = 20

        # =====================================================
        # GENERATE FILE
        # =====================================================

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Daily_MIS_Report_{report_date}.xlsx"

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)})

    finally:
        cur.close()
        conn.close()