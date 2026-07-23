from flask import render_template, session, redirect, url_for, jsonify, request
from datetime import date, datetime, timedelta
from functools import wraps
from database import get_db, get_cursor
from .. import bp
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment 
from openpyxl.utils import get_column_letter
import json as _json
import traceback
from collections import defaultdict
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def _classify_tide_types(rows):
    if not rows:
        return []
    heights = [float(r.get('tide_meters') or 0) for r in rows]
    current = 'HW' if len(heights) >= 2 and heights[0] > heights[1] else 'LW'
    types = []
    for _ in heights:
        types.append(current)
        current = 'LW' if current == 'HW' else 'HW'
    return types

def _parse_dt(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.fromisoformat(str(val).strip())
    except Exception:
        try:
            return datetime.strptime(str(val).strip(), '%Y-%m-%d %H:%M:%S')
        except Exception:
            return None

def _fmt_dt(val, strfmt='%d-%m-%Y %H:%M'):
    dt = _parse_dt(val)
    return dt.strftime(strfmt) if dt else ''

def _fetch_mother_vessels(from_datetime, to_datetime):

    conn = get_db()
    cur  = get_cursor(conn)

    cur.execute("""
        SELECT
            h.id,
            h.vcn_id,
            h.vessel_name,
            h.operation_type,
            h.nor_tendered,

            (
                SELECT MIN(a.discharge_started)
                FROM ldud_anchorage a
                WHERE a.ldud_id = h.id
            ) AS discharge_commenced,

            (
                SELECT MAX(a.discharge_commenced)
                FROM ldud_anchorage a
                WHERE a.ldud_id = h.id
            ) AS discharge_completed

        FROM ldud_header h
        WHERE h.nor_tendered IS NOT NULL
        ORDER BY h.nor_tendered
    """)

    all_vessels = [dict(r) for r in cur.fetchall()]

    vessels = []
    for v in all_vessels:
        commenced = _parse_dt(v.get("discharge_commenced"))
        completed = _parse_dt(v.get("discharge_completed"))

        if not commenced:
            continue

        if commenced > to_datetime:
            continue

        if completed and completed < from_datetime:
            continue

        vessels.append(v)

    ldud_ids = [v['id'] for v in vessels]
    vcn_ids  = [v['vcn_id'] for v in vessels if v.get('vcn_id')]

    bl_import        = {}
    bl_export        = {}
    vcn_meta         = {}
    ops_24h          = {}
    ops_till         = {}
    under_loading    = {}
    at_gull_loaded   = {}
    eta_to_dharamtar = {}
    mbc_eta_list     = []

    report_date = to_datetime.date()
    prev_date   = report_date - timedelta(days=1)

    if ldud_ids:
        cur.execute("""
            SELECT ldud_id, COALESCE(SUM(quantity),0) qty
            FROM ldud_vessel_operations
            WHERE ldud_id = ANY(%s)
              AND TO_DATE(start_time,'YYYY-MM-DD') = %s
            GROUP BY ldud_id
        """, (ldud_ids, prev_date))
        for r in cur.fetchall():
            ops_24h[r['ldud_id']] = float(r['qty'])

    if ldud_ids:
        cur.execute("""
            SELECT ldud_id, COALESCE(SUM(quantity),0) qty
            FROM ldud_vessel_operations
            WHERE ldud_id = ANY(%s)
              AND TO_DATE(start_time,'YYYY-MM-DD') <= %s
            GROUP BY ldud_id
        """, (ldud_ids, prev_date))
        for r in cur.fetchall():
            ops_till[r['ldud_id']] = float(r['qty'])

    if vcn_ids:
        cur.execute("""
            SELECT vcn_id, COALESCE(SUM(bl_quantity),0) total
            FROM vcn_cargo_declaration
            WHERE vcn_id = ANY(%s) GROUP BY vcn_id
        """, (vcn_ids,))
        bl_import = {r['vcn_id']: float(r['total']) for r in cur.fetchall()}

        cur.execute("""
            SELECT vcn_id, COALESCE(SUM(bl_quantity),0) total
            FROM vcn_export_cargo_declaration
            WHERE vcn_id = ANY(%s) GROUP BY vcn_id
        """, (vcn_ids,))
        bl_export = {r['vcn_id']: float(r['total']) for r in cur.fetchall()}

        cur.execute("""
            SELECT id, importer_exporter_name
            FROM vcn_header WHERE id = ANY(%s)
        """, (vcn_ids,))
        vcn_meta = {r['id']: r['importer_exporter_name'] or '' for r in cur.fetchall()}

    if ldud_ids:
        cur.execute("""
            SELECT ldud_id,
                   STRING_AGG(TRIM(barge_name), ', ' ORDER BY barge_name) AS barges
            FROM ldud_barge_lines
            WHERE commenced_loading IS NOT NULL
              AND completed_loading IS NULL
              AND ldud_id = ANY(%s)
            GROUP BY ldud_id
        """, (ldud_ids,))
        under_loading = {r['ldud_id']: r['barges'] for r in cur.fetchall()}

        cur.execute("""
            SELECT ldud_id,
                   STRING_AGG(TRIM(barge_name), ', ' ORDER BY barge_name) AS barges
            FROM ldud_barge_lines
            WHERE cast_off_mv IS NOT NULL
              AND (along_side_berth IS NULL OR TRIM(COALESCE(along_side_berth,'')) = '')
              AND ldud_id = ANY(%s)
            GROUP BY ldud_id
        """, (ldud_ids,))
        at_gull_loaded = {r['ldud_id']: r['barges'] for r in cur.fetchall()}

        cur.execute("""
            SELECT ldud_id,
                STRING_AGG(TRIM(barge_name), ', ' ORDER BY barge_name) AS barges
            FROM ldud_barge_lines
            WHERE anchored_gull_island IS NOT NULL
            AND cast_off_port IS NULL
            AND ldud_id = ANY(%s)
            GROUP BY ldud_id
        """, (ldud_ids,))
        eta_to_dharamtar = {r['ldud_id']: r['barges'] for r in cur.fetchall()}

    cur.execute("""
        SELECT
            h.mbc_name,
            p.departure_gull_island,
            h.cargo_name,
            COALESCE(h.bl_quantity, 0) AS bl_qty
        FROM mbc_header h
        JOIN mbc_discharge_port_lines p ON p.mbc_id = h.id
        WHERE p.departure_gull_island IS NOT NULL
          AND TRIM(COALESCE(p.departure_gull_island, '')) <> ''
          AND (
                p.vessel_arrival_port IS NULL
                OR TRIM(COALESCE(p.vessel_arrival_port, '')) = ''
              )
          AND NULLIF(TRIM(p.departure_gull_island), '')::timestamp <= %s
        ORDER BY p.departure_gull_island
    """, (to_datetime,))
 
    mbc_eta_rows = cur.fetchall()
 
    # Format: "MBC_NAME (cargo) ETA: departure_time"
    mbc_eta_list = []
    for r in mbc_eta_rows:
        dep_dt = _parse_dt(r['departure_gull_island'])
        dep_str = dep_dt.strftime('%d-%m %H:%M') if dep_dt else ''
        entry = f"{r['mbc_name']} ({r['cargo_name'] or ''}) - Dep Gull: {dep_str}"
        mbc_eta_list.append(entry)

    cur.close()
    conn.close()

    for i, v in enumerate(vessels):
        vid    = v.get('vcn_id')
        op     = v.get('operation_type', '')
        bl_qty = (bl_export.get(vid, 0) if op == 'Export' else bl_import.get(vid, 0))

        v['stevedore_group']  = vcn_meta.get(vid, '')
        v['bl_qty']           = bl_qty
        v['ops_24h']          = ops_24h.get(v['id'], 0)
        v['ops_till']         = ops_till.get(v['id'], 0)
        v['balance']          = round(bl_qty - ops_till.get(v['id'], 0), 2)
        v['under_loading']    = under_loading.get(v['id'], '')
        v['eta_to_dharamtar'] = eta_to_dharamtar.get(v['id'], '')
        v['wt_r19']           = ''
        v['at_gull_loaded']   = at_gull_loaded.get(v['id'], '')
        if i < len(mbc_eta_list):
          v['mbc_eta'] = mbc_eta_list[i]
        else:
            v['mbc_eta'] = ''

    return vessels



def _fetch_tide_data(from_datetime, to_datetime):

    conn = get_db()
    cur = get_cursor(conn)

    cur.execute("""
        SELECT
            tide_datetime,
            tide_meters
        FROM tide_master
        WHERE
            tide_datetime IS NOT NULL
            AND TRIM(tide_datetime) <> ''
            AND NULLIF(TRIM(tide_datetime), '')::timestamp >= %s
        ORDER BY NULLIF(TRIM(tide_datetime), '')::timestamp
        LIMIT 6
    """, (from_datetime,))

    rows = [dict(r) for r in cur.fetchall()]

    cur.close()
    conn.close()

    types = _classify_tide_types(rows)

    tide_data = []

    for row, tide_type in zip(rows, types):
        dt = _parse_dt(row["tide_datetime"])

        tide_data.append({
            "type": tide_type,
            "time": dt.strftime("%d/%m %H:%M") if dt else "",
            "height": row["tide_meters"],
        })

    return tide_data




def get_shift_code(dt):
    if not dt:
        return None
    hour = dt.hour
    if 6 <= hour < 14:
        return "A"
    elif 14 <= hour < 22:
        return "B"
    else:
        return "C"


def _fetch_all_barges(selected_date=None, selected_shift="ALL"):

    conn = get_db()
    cur  = get_cursor(conn)

    barges = []
    occupied_berth_set = set()

    selected_dt = None
    if selected_date:
        try:
            selected_dt = datetime.strptime(selected_date, "%Y-%m-%d").date()
        except Exception:
            pass

    cur.execute("""
        WITH discharge_sums AS (
            SELECT
                TRIM(UPPER(ll.barge_name)) AS barge_name,
                ll.source_id,
                SUM(COALESCE(ll.quantity,0)) AS discharged_qty
            FROM lueu_lines ll
            WHERE ll.is_deleted IS NOT TRUE
            AND ll.source_type = 'VCN'
            GROUP BY TRIM(UPPER(ll.barge_name)), ll.source_id
        )
        SELECT
            l.id,
            l.barge_name,
            l.trip_number,
            l.cargo_name,
            l.cast_off_port,
            l.along_side_berth,
            l.commence_discharge_berth,
            l.completed_discharge_berth,
            COALESCE(l.discharge_quantity,0) AS discharge_qty,
            (
                COALESCE(l.discharge_quantity,0)
                - COALESCE(ds.discharged_qty,0)
            ) AS balance_qty
        FROM ldud_barge_lines l
        LEFT JOIN ldud_header h
            ON h.id = l.ldud_id
        LEFT JOIN discharge_sums ds
            ON ds.barge_name = TRIM(
                UPPER(
                    CONCAT(
                        l.barge_name,
                        ' / ',
                        COALESCE(l.trip_number::text,'1')
                    )
                )
            )
            AND ds.source_id = h.vcn_id
        WHERE COALESCE(TRIM(l.barge_name),'') <> ''
    """)
    for row in cur.fetchall():
        row       = dict(row)
        balance_qty = max(float(row.get("balance_qty") or 0), 0)
        
        cutoff_dt = datetime(2026, 5, 1)

        arrival_dt = _parse_dt(row.get("along_side_berth"))

        # Skip barges that first arrived before 1 May 2026
        if arrival_dt and arrival_dt < cutoff_dt:
            continue

        # Skip completed
        if row.get("cast_off_port") and str(row.get("cast_off_port")).strip():
            continue
        if balance_qty <= 0:
           continue

        if row.get("completed_discharge_berth") and str(row.get("completed_discharge_berth")).strip():
            continue

        # Waiting — alongside berth set, discharge not started
        if (
            row.get("along_side_berth")
            and str(row.get("along_side_berth")).strip()
            and (
                row.get("commence_discharge_berth") is None
                or str(row.get("commence_discharge_berth")).strip() == ""
            )
        ):
            status = "Waiting"

        # Under Discharge — discharge started, not completed
        elif (
            row.get("commence_discharge_berth")
            and str(row.get("commence_discharge_berth")).strip()
            and (
                row.get("completed_discharge_berth") is None
                or str(row.get("completed_discharge_berth")).strip() == ""
            )
        ):
            status = "Under Discharge"

        # ── FIX: neither condition matched (e.g. barge hasn't reached
        #    berth yet, or both fields blank/whitespace) — skip it
        #    instead of crashing with UnboundLocalError.
        else:
            continue

        completed_date = _fmt_dt(row.get("cast_off_port")) if row.get("cast_off_port") else None
        berth = (row.get("commence_discharge_berth") or row.get("along_side_berth") or "")

        barges.append({
            "id":             row["id"],
            "type":           "BARGE",
            "completed_date": completed_date,
            "barge_name":     row["barge_name"],
            "name":           row["barge_name"],
            "cargo": row.get("cargo_name") or row.get("cargo_type") or "",
            "qty":            row["discharge_qty"],
            "discharge_qty":  float(row["discharge_qty"]),
            "total_qty":      float(row["discharge_qty"]),
            "balance_qty": balance_qty,
            "berth":          berth,
            "status":         status,
            "commence_discharge_berth": str(row.get("commence_discharge_berth") or "").strip(),
            "unloading_commenced":      "",   # barges use commence_discharge_berth
        })

    cur.execute("""
        WITH latest_mbc AS (
            SELECT
                h.id,
                h.mbc_name,
                h.cargo_name,
                COALESCE(h.bl_quantity,0) AS bl_qty,
                p.vessel_unloading_berth AS berth,
                p.vessel_arrival_port AS arrival_port,
                p.unloading_commenced,
                p.unloading_completed,
                p.vessel_cast_off AS mbc_cast_off,
                ROW_NUMBER() OVER (
                    PARTITION BY h.id
                    ORDER BY p.id DESC
                ) rn
            FROM mbc_header h
            JOIN mbc_discharge_port_lines p
                ON p.mbc_id = h.id
        )
        SELECT
            m.*,
            COALESCE(l.actual_qty,0) AS actual_qty
        FROM latest_mbc m
        LEFT JOIN (
            SELECT
                source_id,
                SUM(COALESCE(quantity,0)) AS actual_qty
            FROM lueu_lines
            WHERE source_type='MBC'
            AND is_deleted IS NOT TRUE
            GROUP BY source_id
        ) l
        ON l.source_id = m.id
        WHERE m.rn = 1
        AND m.arrival_port IS NOT NULL
        AND TRIM(COALESCE(m.arrival_port,'')) <> ''
        AND (
                m.unloading_completed IS NULL
                OR TRIM(COALESCE(m.unloading_completed,'')) = ''
        )
        AND (
                m.mbc_cast_off IS NULL
                OR TRIM(COALESCE(m.mbc_cast_off,'')) = ''
        )
        ORDER BY m.mbc_name
    """)

    for row in cur.fetchall():
        cutoff_dt = datetime(2026, 5, 1)
        

        arrival_dt = _parse_dt(row.get("along_side_berth"))

        # Skip barges arrived before 1 May 2026
        if arrival_dt and arrival_dt < cutoff_dt:
            continue

        cutoff_dt = datetime(2026, 5, 1)
        

        

        # Skip MBCs that arrived before 1 May 2026
        if arrival_dt and arrival_dt < cutoff_dt:
            continue

        bl_qty = float(row["bl_qty"] or 0)
        actual_qty = float(row["actual_qty"] or 0)
        balance_qty = max(bl_qty - actual_qty, 0)
        # Automatically hide completed MBC
        if balance_qty <= 0:
            continue

        # ---------------------------------------------------------
        # IMPORTANT:
        # Always send active MBCs to WAITING.
        # Saved berth_layout will decide whether they appear on a berth.
        # ---------------------------------------------------------

        status = "Waiting"
        berth = ""

        print(
            row["mbc_name"],
            "arrival =", row.get("arrival_port"),
            "db berth =", row.get("berth"),
            "started =", row.get("unloading_commenced"),
            "status =", status
        )
        

        # Automatically hide completed barges
        if balance_qty <= 0:
            continue
                
        barges.append({
            "id": row["id"],
            "type": "MBC",
            "completed_date": _fmt_dt(row.get("unloading_completed")),
            "barge_name": row["mbc_name"],
            "name": row["mbc_name"],
            "cargo": row.get("cargo_name") or "",
            "qty": bl_qty,
            "discharge_qty": bl_qty,
            "total_qty": bl_qty,
            "balance_qty": balance_qty,
            "berth": berth,
            "status": status,
            "unloading_commenced": str(row.get("unloading_commenced") or "").strip(),
            "commence_discharge_berth": "",
        })

        if status == "Discharging" and berth:
            occupied_berth_set.add(berth)

    # Sort alphabetically by name regardless of type
    barges.sort(key=lambda x: (x["name"] or "").upper())

    cur.close()
    conn.close()
    return barges, occupied_berth_set


# ── ROUTES — each defined exactly ONCE ───────────────────────────────────────

@bp.route('/module/RP01/bargeposition/')
@login_required
def bargeposition():

    barges, occupied_berth_set = _fetch_all_barges()

    waiting = [
    b for b in barges
    if b["status"] in ["Waiting", "Under Discharge"]
    ]

    discharging = [
        b for b in barges
        if b["status"] == "Discharging"
    ]
    occupied_berths = len(occupied_berth_set)

    today         = date.today()
    today_str     = today.strftime('%Y-%m-%d')

    from_date_str = request.args.get('from_date', today_str)
    from_time_str = request.args.get('from_time', '00:00')
    to_date_str   = request.args.get('to_date',   today_str)
    to_time_str   = request.args.get('to_time',   '23:59')

    to_datetime   = datetime.strptime(f"{to_date_str} {to_time_str}", '%Y-%m-%d %H:%M')

    # Widen window: start from previous day 00:00 so vessels that
    # completed early on the selected date are still included
    from_datetime = datetime.strptime(from_date_str, '%Y-%m-%d') - timedelta(days=1)
    from_datetime = from_datetime.replace(hour=0, minute=0, second=0)

    mother_vessels_raw = _fetch_mother_vessels(from_datetime, to_datetime)

    mother_vessels = [{
        'vessel_name':         v.get('vessel_name') or '',
        'stevedore_group':     v.get('stevedore_group') or '',
        'bl_qty':              v.get('bl_qty') or 0,
        'ops_24h':             v.get('ops_24h') or 0,
        'ops_till':            v.get('ops_till') or 0,
        'balance':             v.get('balance') or 0,
        'under_loading':       v.get('under_loading') or '',
        'eta_to_dharamtar':    v.get('eta_to_dharamtar') or '',
        'wt_r19':              v.get('wt_r19') or '',
        'at_gull_loaded':      v.get('at_gull_loaded') or '',
        'mbc_eta':             v.get('mbc_eta') or '',
        'nor_tendered':        _fmt_dt(v.get('nor_tendered')),
        'discharge_commenced': _fmt_dt(v.get('discharge_commenced')),
        'discharge_completed': _fmt_dt(v.get('discharge_completed')),
        'unloaded_till_date':  '',
        'disch_commenced':     '',
    } for v in mother_vessels_raw]
    
    conn = get_db()
    cur = get_cursor(conn)



    cur.execute("""
    SELECT berth_name
    FROM port_berth_master
    """)

    berths = [r["berth_name"].upper() for r in cur.fetchall()]

    old_berths = [
        "BERTH 1",
        "BERTH 2",
        "BERTH 3",
        "BERTH 4",
        "BERTH 5",
        "BERTH 5A",
    ]

    new_berths = [
        "BERTH 6",
        "BERTH 7",
        "BERTH 8",
        "BERTH 8A",
        "BERTH 9",
        "BERTH 10",
        "BERTH 11",
        "BERTH 12",
    ]

    cur.close()
    conn.close()

    tide_data = _fetch_tide_data(from_datetime, to_datetime)

    return render_template(
        "Barge_Position_Report/barge_dashboard.html",
        waiting=waiting,
        discharging=discharging,
        all_barges=barges,
        mother_vessels=mother_vessels,
        tide_data=tide_data,
         berths=berths,
        old_berths=old_berths,
        new_berths=new_berths,
        from_date=from_date_str,
        from_time=from_time_str,
        to_date=to_date_str,
        to_time=to_time_str,
        total_barges=len(barges),
        waiting_count=len(waiting),
        discharging_count=len(discharging),
        occupied_berths=occupied_berths,
        available_berths=max(0, 14 - occupied_berths),
        
    )
    
# ── API ROUTES ────────────────────────────────────────────────────────────────

@bp.route('/api/module/RP01/berth-occupancy')
@login_required
def api_berth_occupancy():
    include_completed = request.args.get('completed', '0') == '1'

    if include_completed:
        report_date = request.args.get('date', '')
        shift       = request.args.get('shift', 'ALL')

        if not report_date:
            return jsonify([])

        try:
            base = datetime.strptime(report_date, '%Y-%m-%d')
        except Exception:
            return jsonify([])

        # For date-only filtering — match the full selected date regardless of time
        date_start = base.replace(hour=0,  minute=0,  second=0)
        date_end   = base.replace(hour=23, minute=59, second=59)

        # If a specific shift is selected, narrow the window
        if shift.upper() != 'ALL':
            SHIFT_WINDOWS = {
                'A': (base.replace(hour=6,  minute=0),  base.replace(hour=14, minute=0)),
                'B': (base.replace(hour=14, minute=0),  base.replace(hour=22, minute=0)),
                'C': (base.replace(hour=22, minute=0),  (base + timedelta(days=1)).replace(hour=6, minute=0)),
            }
            from_dt, to_dt = SHIFT_WINDOWS.get(shift.upper(), (date_start, date_end))
        else:
            from_dt, to_dt = date_start, date_end

        conn = get_db()
        cur  = get_cursor(conn)
        results = []

        # Completed BARGES
        # Use flexible regex: matches YYYY-MM-DD with space OR T separator
        cur.execute(r"""
            SELECT *
            FROM (
                SELECT
                    bl.barge_name,
                    bl.cargo_name,
                    bl.commence_discharge_berth,
                    bl.along_side_berth,
                    bl.completed_discharge_berth,
                    bl.cast_off_port,
                    COALESCE(bl.discharge_quantity, 0) AS bl_qty,
                    CASE
                        WHEN bl.cast_off_port IS NOT NULL
                             AND TRIM(bl.cast_off_port) ~ '^\d{4}-\d{2}-\d{2}[T ]'
                        THEN SUBSTRING(TRIM(bl.cast_off_port), 1, 10)::date
                        WHEN bl.completed_discharge_berth IS NOT NULL
                             AND TRIM(bl.completed_discharge_berth) ~ '^\d{4}-\d{2}-\d{2}[T ]'
                        THEN SUBSTRING(TRIM(bl.completed_discharge_berth), 1, 10)::date
                        ELSE NULL
                    END AS completed_date
                FROM ldud_barge_lines bl
                JOIN ldud_header h ON h.id = bl.ldud_id
                WHERE COALESCE(TRIM(bl.barge_name),'') <> ''
                  AND (
                        (bl.cast_off_port IS NOT NULL
                         AND TRIM(bl.cast_off_port) ~ '^\d{4}-\d{2}-\d{2}[T ]')
                        OR
                        (bl.completed_discharge_berth IS NOT NULL
                         AND TRIM(bl.completed_discharge_berth) ~ '^\d{4}-\d{2}-\d{2}[T ]')
                      )
            ) sub
            WHERE sub.completed_date = %s::date
            ORDER BY sub.barge_name
        """, (report_date,))

        for row in cur.fetchall():
            row = dict(row)
            commenced = _fmt_dt(row.get('commence_discharge_berth') or row.get('along_side_berth'))
            completed = _fmt_dt(row.get('cast_off_port') or row.get('completed_discharge_berth'))
            results.append({
                'type':      'BARGE',
                'status':    'Completed',
                'name':      row['barge_name'],
                'cargo':     row.get('cargo_name') or '',
                'bl_qty':    float(row['bl_qty'] or 0),
                'commenced': commenced,
                'completed': completed,
            })

        # Completed MBCs
        cur.execute(r"""
            SELECT *
            FROM (
                SELECT
                    h.mbc_name,
                    h.cargo_name,
                    COALESCE(h.bl_quantity, 0) AS bl_qty,
                    p.unloading_commenced,
                    p.unloading_completed,
                    p.vessel_cast_off,
                    CASE
                        WHEN p.unloading_completed IS NOT NULL
                             AND TRIM(p.unloading_completed) ~ '^\d{4}-\d{2}-\d{2}[T ]'
                        THEN SUBSTRING(TRIM(p.unloading_completed), 1, 10)::date
                        WHEN p.vessel_cast_off IS NOT NULL
                             AND TRIM(p.vessel_cast_off) ~ '^\d{4}-\d{2}-\d{2}[T ]'
                        THEN SUBSTRING(TRIM(p.vessel_cast_off), 1, 10)::date
                        ELSE NULL
                    END AS completed_date
                FROM mbc_header h
                JOIN mbc_discharge_port_lines p ON p.mbc_id = h.id
                WHERE (
                        (p.unloading_completed IS NOT NULL
                         AND TRIM(p.unloading_completed) ~ '^\d{4}-\d{2}-\d{2}[T ]')
                        OR
                        (p.vessel_cast_off IS NOT NULL
                         AND TRIM(p.vessel_cast_off) ~ '^\d{4}-\d{2}-\d{2}[T ]')
                      )
            ) sub
            WHERE sub.completed_date = %s::date
            ORDER BY sub.mbc_name
        """, (report_date,))

        for row in cur.fetchall():
            row = dict(row)
            results.append({
                'type':      'MBC',
                'status':    'Completed',
                'name':      row['mbc_name'],
                'cargo':     row.get('cargo_name') or '',
                'bl_qty':    float(row['bl_qty'] or 0),
                'commenced': _fmt_dt(row.get('unloading_commenced')),
                'completed': _fmt_dt(row.get('unloading_completed') or row.get('vessel_cast_off')),
            })

        cur.close()
        conn.close()
        return jsonify(results)

    # Original logic — completely untouched
    barges, _ = _fetch_all_barges()
    return jsonify(barges)


@bp.route('/api/module/RP01/mother-vessel-data')
@login_required
def api_mother_vessel_data():
    from_datetime = _parse_dt(request.args.get('from_datetime'))
    to_datetime   = _parse_dt(request.args.get('to_datetime'))
    if not from_datetime or not to_datetime:
        return jsonify([])
    vessels_raw = _fetch_mother_vessels(from_datetime, to_datetime)
    vessels = [{
        'vessel_name':         v.get('vessel_name') or '',
        'discharge_commenced': _fmt_dt(v.get('discharge_commenced')),
        'discharge_completed': _fmt_dt(v.get('discharge_completed')),
        'under_loading':       v.get('under_loading') or '',
        'eta_to_dharamtar':    v.get('eta_to_dharamtar') or '',
        'wt_r19':              v.get('wt_r19') or '',
        'at_gull_loaded':      v.get('at_gull_loaded') or '',
        'mbc_eta':             v.get('mbc_eta') or '',
    } for v in vessels_raw]
    return jsonify(vessels)


@bp.route('/api/module/RP01/tide-data')
@login_required
def api_tide_data():
    from_datetime = _parse_dt(request.args.get('from_datetime'))
    to_datetime   = _parse_dt(request.args.get('to_datetime'))
    if not from_datetime or not to_datetime:
        return jsonify([])
    return jsonify(_fetch_tide_data(from_datetime, to_datetime))


@bp.route('/api/module/RP01/shift-details')
@login_required
def get_shift_details():
    conn = get_db()
    cur  = get_cursor(conn)
    cur.execute("SELECT name FROM port_shift_incharge ORDER BY name")
    shift_incharge_list = [r["name"] for r in cur.fetchall()]
    cur.execute("SELECT name FROM port_shift_operators ORDER BY name")
    crane_operator_list = [r["name"] for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({
        "shift_incharge_list": shift_incharge_list,
        "crane_operator_list": crane_operator_list,
    })


@bp.route('/api/module/RP01/shift-wise-discharge')
@login_required
def shift_wise_discharge():
    import traceback
    try:
        return _shift_wise_discharge_inner()
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

def _shift_wise_discharge_inner():
    selected_date = request.args.get('date', '')
    shift = request.args.get('shift', 'ALL')
    if not selected_date:
        return jsonify({'error': 'date is required'}), 400

    conn = get_db()
    cur = get_cursor(conn)
    is_all = shift.upper() == 'ALL'

    # ── 1. JETTY DISCHARGE ───────────────────────────────────────────────────
    if is_all:
        cur.execute("""
            SELECT cargo_name, COALESCE(SUM(quantity), 0) AS qty
            FROM lueu_lines
            WHERE entry_date = %s
              AND quantity > 0
              AND cargo_name IS NOT NULL AND cargo_name != ''
              AND is_deleted IS NOT TRUE
            GROUP BY cargo_name ORDER BY cargo_name
        """, (selected_date,))
    else:
        cur.execute("""
            SELECT cargo_name, COALESCE(SUM(quantity), 0) AS qty
            FROM lueu_lines
            WHERE entry_date = %s AND shift = %s
              AND quantity > 0
              AND cargo_name IS NOT NULL AND cargo_name != ''
              AND is_deleted IS NOT TRUE
            GROUP BY cargo_name ORDER BY cargo_name
        """, (selected_date, shift))
    jetty_rows = [dict(r) for r in cur.fetchall()]

    # ── 2. CLEANING DELAYS ───────────────────────────────────────────────────
    if is_all:
        cur.execute("""
            SELECT source_id, source_type, delay_name, barge_name
            FROM lueu_lines
            WHERE entry_date = %s
              AND is_deleted IS NOT TRUE
              AND delay_name IS NOT NULL AND delay_name != ''
                AND (
                    LOWER(delay_name) LIKE '%%payloader%%'
                    OR LOWER(delay_name) LIKE '%%labor cleaning%%'
                )
        """, (selected_date,))
    else:
        cur.execute("""
            SELECT source_id, source_type, delay_name,barge_name
            FROM lueu_lines
            WHERE entry_date = %s AND shift = %s
              AND is_deleted IS NOT TRUE
              AND delay_name IS NOT NULL AND delay_name != ''
                AND (
                    LOWER(delay_name) LIKE '%%payloader%%'
                    OR LOWER(delay_name) LIKE '%%labor cleaning%%'
                )
        """, (selected_date, shift))

    delay_map = {}
    for row in cur.fetchall():
        key = (row["source_id"], row["source_type"])

        if key not in delay_map:
            delay_map[key] = {
                "payloader": False,
                "labour": False
            }

        delay = (row["delay_name"] or "").strip().lower()

        print("DELAY ROW:", row["source_id"], row["source_type"], delay)

        if "payloader" in delay:
            delay_map[key]["payloader"] = True

        if "labor cleaning" in delay or "labour cleaning" in delay:
            delay_map[key]["labour"] = True

    print("DELAY MAP:", delay_map)


    # ── 3. BARGE DISCHARGE ──
    if is_all:
            cur.execute("""
                WITH actual AS (
                    SELECT
                        TRIM(UPPER(barge_name)) AS barge_key,
                        source_id,
                        SUM(COALESCE(quantity,0)) AS actual_qty
                    FROM lueu_lines
                    WHERE is_deleted IS NOT TRUE AND source_type = 'VCN'
                    AND entry_date = %s
                    GROUP BY 1, 2
                    HAVING SUM(COALESCE(quantity,0)) > 0
                )
                SELECT bl.id, bl.barge_name, bl.trip_number, bl.cargo_name,
                    COALESCE(bl.discharge_quantity, 0) AS bl_qty,
                    COALESCE(a.actual_qty, 0) AS actual_discharge,
                    bl.along_side_berth, bl.commence_discharge_berth,
                    bl.completed_discharge_berth, bl.cast_off_port, h.vcn_id
                FROM ldud_barge_lines bl
                JOIN ldud_header h ON h.id = bl.ldud_id
                LEFT JOIN actual a
                    ON a.barge_key = TRIM(
                        UPPER(
                            CONCAT(
                                bl.barge_name,
                                ' / ',
                                COALESCE(bl.trip_number::text,'1')
                            )
                        )
                    )
                AND a.source_id = h.vcn_id
                WHERE COALESCE(TRIM(bl.barge_name),'') <> ''
                AND COALESCE(a.actual_qty,0) > 0
                ORDER BY bl.barge_name
            """, (selected_date,))
    else:
        cur.execute("""
                WITH actual AS (
                    SELECT
                        TRIM(UPPER(barge_name)) AS barge_key,
                        source_id,
                        SUM(COALESCE(quantity,0)) AS actual_qty
                    FROM lueu_lines
                    WHERE is_deleted IS NOT TRUE AND source_type = 'VCN'
                    AND entry_date = %s AND shift = %s
                    GROUP BY 1, 2
                )
                SELECT bl.id, bl.barge_name,bl.trip_number, bl.cargo_name,
                    COALESCE(bl.discharge_quantity, 0) AS bl_qty,
                    COALESCE(a.actual_qty, 0) AS actual_discharge,
                    bl.along_side_berth, bl.commence_discharge_berth,
                    bl.completed_discharge_berth, bl.cast_off_port, h.vcn_id
                FROM ldud_barge_lines bl
                JOIN ldud_header h ON h.id = bl.ldud_id
                INNER JOIN actual a
                    ON a.barge_key = TRIM(
                        UPPER(
                            CONCAT(
                                bl.barge_name,
                                ' / ',
                                COALESCE(bl.trip_number::text,'1')
                            )
                        )
                    )
                AND a.source_id = h.vcn_id
                WHERE COALESCE(TRIM(bl.barge_name),'') <> ''
                AND COALESCE(a.actual_qty,0) > 0
                ORDER BY bl.barge_name
            """, (selected_date, shift))

    from datetime import date as date_cls
    cutoff = date_cls(2026, 5, 1)
    sel_date_obj = datetime.strptime(selected_date, '%Y-%m-%d').date()

    barge_discharge = []
    for row in cur.fetchall():
        row = dict(row)
        if sel_date_obj < cutoff:
            continue
        if row.get('cast_off_port') and str(row['cast_off_port']).strip():
            status = 'Completed'
        elif row.get('completed_discharge_berth') and str(row['completed_discharge_berth']).strip():
            status = 'Completed'
        elif row.get('commence_discharge_berth') and str(row['commence_discharge_berth']).strip():
            status = 'Under Discharge'
        else:
            status = 'Waiting'

        delays = delay_map.get(
            (row['vcn_id'], 'VCN'),
            {'payloader': False, 'labour': False}
        )
        print(
            "BARGE:",
            row["barge_name"],
            row["vcn_id"],
            delays
        )
        barge_discharge.append({
            'type': 'BARGE',
            'name': row['barge_name'],
            'cargo': row.get('cargo_name') or '',
            'bl_qty': float(row['bl_qty'] or 0),
            'actual_discharge': float(row['actual_discharge'] or 0),
            'status': status,

            'payloader_cl': row['barge_name'] if delays['payloader'] else '',
            'labour_cleaned': row['barge_name'] if delays['labour'] else '',
        })

        # ── 4. MBC DISCHARGE ─────────────────────────────────────────────────────

    if is_all:
            cur.execute("""
                SELECT
                    l.source_id AS id,
                    COALESCE(NULLIF(TRIM(l.barge_name), ''), h.mbc_name) AS mbc_name,
                    COALESCE(l.cargo_name, h.cargo_name) AS cargo_name,
                    COALESCE(h.bl_quantity, 0) AS bl_qty,
                    SUM(COALESCE(l.quantity,0)) AS actual_discharge
                FROM lueu_lines l
                JOIN mbc_header h
                    ON h.id = l.source_id
                WHERE l.source_type = 'MBC'
                AND l.is_deleted IS NOT TRUE
                AND l.entry_date = %s
                GROUP BY
                    l.source_id,
                    COALESCE(NULLIF(TRIM(l.barge_name), ''), h.mbc_name),
                    COALESCE(l.cargo_name, h.cargo_name),
                    h.bl_quantity
                HAVING SUM(COALESCE(l.quantity,0)) > 0
                ORDER BY
                    COALESCE(NULLIF(TRIM(l.barge_name), ''), h.mbc_name)
            """, (selected_date,))
    else:
            cur.execute("""
                SELECT
                    l.source_id AS id,
                    COALESCE(NULLIF(TRIM(l.barge_name), ''), h.mbc_name) AS mbc_name,
                    COALESCE(l.cargo_name, h.cargo_name) AS cargo_name,
                    COALESCE(h.bl_quantity, 0) AS bl_qty,
                    SUM(COALESCE(l.quantity,0)) AS actual_discharge
                FROM lueu_lines l
                JOIN mbc_header h
                    ON h.id = l.source_id
                WHERE l.source_type = 'MBC'
                AND l.is_deleted IS NOT TRUE
                AND l.entry_date = %s
                AND l.shift = %s
                GROUP BY
                    l.source_id,
                    COALESCE(NULLIF(TRIM(l.barge_name), ''), h.mbc_name),
                    COALESCE(l.cargo_name, h.cargo_name),
                    h.bl_quantity
                HAVING SUM(COALESCE(l.quantity,0)) > 0
                ORDER BY
                    COALESCE(NULLIF(TRIM(l.barge_name), ''), h.mbc_name)
            """, (selected_date, shift))

    mbc_discharge = []

    for row in cur.fetchall():
        cutoff_dt = datetime(2026, 5, 1)

        arrival_dt = _parse_dt(row.get("arrival_port"))

        # Skip MBCs arrived before 1 May 2026
        if arrival_dt and arrival_dt < cutoff_dt:
            continue

        delays = delay_map.get(
            (row['id'], 'MBC'),
            {'payloader': False, 'labour': False}
        )
        print(
            "MBC:",
            row["mbc_name"],
            row["id"],
            delays
        )

        mbc_discharge.append({
            'type': 'MBC',
            'name': row['mbc_name'],
            'cargo': row.get('cargo_name') or '',
            'bl_qty': float(row['bl_qty'] or 0),
            'actual_discharge': float(row['actual_discharge'] or 0),
            'status': 'Discharging',
            'payloader_cl': row['mbc_name'] if delays['payloader'] else '',
            'labour_cleaned': row['mbc_name'] if delays['labour'] else '',
        })

    cur.close()
    conn.close()

    return jsonify({
        'jetty_discharge': jetty_rows,
        'barge_discharge': barge_discharge,
        'mbc_discharge': mbc_discharge,
    })

@bp.route('/api/module/RP01/shift-report/save', methods=['POST'])
@login_required
def api_shift_report_save():
    """
    Save shift report data to barge_position_report table.
    
    Expected JSON:
    {
        "report_date": "2026-07-16",
        "shift": "A",
        "shift_incharge": "Name",
        "bpo": "Name",
        "crane_operator": "Name",
        "berth_layout": [...],  # Combined with waiting_area
        "notes": [...],
        "wt_r19": {...},
        "mbc_eta": {...},
        "eta_to_dharamtar": {...},
        "on_the_way_gull": {...},
        "shift_plan": {...},
        "movement_logs": [...]
    }
    
    Database Operation:
        INSERT INTO barge_position_report (...)
        ON CONFLICT (report_date, shift)
        DO UPDATE SET ...
    """
    
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    conn = get_db()
    cur = get_cursor(conn)

    try:
        # Prepare data
        report_date = data.get('report_date')
        shift = data.get('shift')
        shift_incharge = data.get('shift_incharge', '')
        bpo = data.get('bpo', '')
        crane_operator = data.get('crane_operator', '')
        
        # Separate berth_layout and waiting_area
        all_layout = data.get('berth_layout', [])
        berth_layout = [
            item for item in all_layout 
            if item.get('berth') != 'WAITING'
        ]
        waiting_area = [
            item for item in all_layout 
            if item.get('berth') == 'WAITING'
        ]
        
        # Convert to JSON for storage
        berth_layout_json = _json.dumps(berth_layout)
        waiting_area_json = _json.dumps(waiting_area)
        wt_r19_json = _json.dumps(data.get('wt_r19', {}))
        mbc_eta_json = _json.dumps(data.get('mbc_eta', {}))
        eta_to_dharamtar_json = _json.dumps(data.get('eta_to_dharamtar', {}))
        on_the_way_gull_json = _json.dumps(data.get('on_the_way_gull', {}))
        shift_plan_json = _json.dumps(data.get('shift_plan', {}))
        notes_json = _json.dumps(data.get('notes', []))
        movement_logs_json = _json.dumps(data.get('movement_logs', []))
        
        # ====================================================================
        # UPSERT: Insert or update based on (report_date, shift)
        # ====================================================================
        
        cur.execute("""
            INSERT INTO barge_position_report
                (report_date, shift, shift_incharge, bpo, crane_operator,
                 berth_layout, waiting_area, wt_r19, mbc_eta,
                 eta_to_dharamtar, on_the_way_gull, shift_plan,
                 notes, movement_logs, created_at, updated_at)
            VALUES 
                (%s, %s, %s, %s, %s,
                 %s::jsonb, %s::jsonb, %s::jsonb, %s::jsonb,
                 %s::jsonb, %s::jsonb, %s::jsonb,
                 %s::jsonb, %s::jsonb, NOW(), NOW())
            ON CONFLICT (report_date, shift)
            DO UPDATE SET
                shift_incharge   = EXCLUDED.shift_incharge,
                bpo              = EXCLUDED.bpo,
                crane_operator   = EXCLUDED.crane_operator,
                berth_layout     = EXCLUDED.berth_layout,
                waiting_area     = EXCLUDED.waiting_area,
                wt_r19           = EXCLUDED.wt_r19,
                mbc_eta          = EXCLUDED.mbc_eta,
                eta_to_dharamtar = EXCLUDED.eta_to_dharamtar,
                on_the_way_gull  = EXCLUDED.on_the_way_gull,
                shift_plan       = EXCLUDED.shift_plan,
                notes            = EXCLUDED.notes,
                movement_logs    = EXCLUDED.movement_logs,
                updated_at       = NOW()
        """, (
            report_date, shift, shift_incharge, bpo, crane_operator,
            berth_layout_json, waiting_area_json,
            wt_r19_json, mbc_eta_json, eta_to_dharamtar_json,
            on_the_way_gull_json, shift_plan_json,
            notes_json, movement_logs_json,
        ))
        
        conn.commit()
        return jsonify({'ok': True})

    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    
    finally:
        cur.close()
        conn.close()

@bp.route('/api/module/RP01/shift-report/load')
@login_required
def api_shift_report_load():
    """
    TRACK 1 — BERTH SECTION (shift-specific, save/carry-forward):
        Only actual berth assignments (berth_layout). Waiting area is
        intentionally EXCLUDED from this track now — it must never be
        read from saved or carried-forward DB rows.

    WAITING AREA:
        Always computed LIVE from _fetch_all_barges() at request time.
        Never taken from `own` saved row, never carried forward from a
        previous shift/date. What's live in the backend right now is
        what gets shown — that's it.

    TRACK 2 — DATE-SPECIFIC FIELDS (unchanged):
        notes, wt_r19, mbc_eta, eta_to_dharamtar, on_the_way_gull —
        merged oldest -> newest across eligible rows.
    """

    report_date = request.args.get('date')
    shift = request.args.get('shift', 'ALL')

    if not report_date:
        return jsonify({'found': False})

    conn = get_db()
    cur = get_cursor(conn)

    def shift_index(s):
        return {'A': 0, 'B': 1, 'C': 2}.get((s or '').upper(), 3)

    def has_berth_content(row):
        """True only if THIS row has real BERTH assignments.
        Waiting area no longer counts here — a row saved with only
        waiting-area entries and no actual berths must NOT be treated
        as having berth content."""
        return bool(row.get('berth_layout') or [])

    try:
        requested_date_obj = datetime.strptime(report_date, '%Y-%m-%d').date()
    except Exception:
        return jsonify({'found': False, 'error': 'invalid date'}), 400

    requested_idx = shift_index(shift)
    requested_key = (requested_date_obj, requested_idx)

    def row_key(r):
        rd = r["report_date"]
        if isinstance(rd, datetime):
            rd = rd.date()
        elif isinstance(rd, str):
            rd = datetime.strptime(rd[:10], "%Y-%m-%d").date()
        return (rd, shift_index(r.get("shift")))

    try:
        # ── Exact-match row for THIS shift+date ──────────────────────────
        cur.execute("""
            SELECT * FROM barge_position_report
            WHERE report_date = %s AND shift = %s
            LIMIT 1
        """, (report_date, shift))
        exact_row = cur.fetchone()
        exact_row = dict(exact_row) if exact_row else None

        # ── All rows up to and including the requested date ──────────────
        cur.execute("""
            SELECT
                report_date, shift, berth_layout, waiting_area,
                shift_incharge, bpo, crane_operator, notes,
                wt_r19, mbc_eta, eta_to_dharamtar, on_the_way_gull,
                movement_logs, shift_plan, updated_at
            FROM barge_position_report
            WHERE report_date <= %s
            ORDER BY
            report_date DESC,
            CASE shift
                WHEN 'C' THEN 3
                WHEN 'B' THEN 2
                WHEN 'A' THEN 1
                ELSE 0
            END DESC
        """, (report_date,))
        all_rows = [dict(r) for r in cur.fetchall()]

        # ══════════════════════════════════════════════════════════════
        # TRACK 1 — BERTH SECTION ONLY (waiting area excluded)
        # ══════════════════════════════════════════════════════════════
        berth_source = "none"
        berth_layout_out = []
        shift_incharge_out = ""
        bpo_out = ""
        crane_operator_out = ""
        movement_logs_out = []
        berth_updated_at = None

        if (
            exact_row
            and exact_row.get("shift") != "ALL"
            and has_berth_content(exact_row)
        ):
            latest = exact_row
            berth_source = "own"
        else:
            latest = None
            valid_rows = []
            for row in all_rows:
                if (row.get("shift") or "").upper() == "ALL":
                    continue
                if not has_berth_content(row):
                    continue
                if row_key(row) >= requested_key:
                    continue
                valid_rows.append(row)
            valid_rows.sort(key=row_key, reverse=True)
            if valid_rows:
                latest = valid_rows[0]
                berth_source = "carried"

        if latest:
            combined = latest.get("berth_layout") or []  # berths only, no waiting

            shift_incharge_out = latest.get("shift_incharge", "")
            bpo_out = latest.get("bpo", "")
            crane_operator_out = latest.get("crane_operator", "")
            berth_updated_at = latest.get("updated_at")

            if berth_source == "own":
                movement_logs_out = latest.get("movement_logs") or []
            else:
                movement_logs_out = []

            active_keys = set()

            cur.execute("""
                SELECT TRIM(UPPER(barge_name)) AS key
                FROM ldud_barge_lines
                WHERE COALESCE(TRIM(barge_name),'') <> ''
                AND (cast_off_port IS NULL OR TRIM(cast_off_port)='')
                AND (completed_discharge_berth IS NULL OR TRIM(completed_discharge_berth)='')
            """)
            for r in cur.fetchall():
                r = dict(r)
                if r.get("key"):
                    active_keys.add(r["key"])

            cur.execute("""
                SELECT TRIM(UPPER(h.mbc_name)) AS key
                FROM mbc_header h
                JOIN mbc_discharge_port_lines p
                ON p.mbc_id = h.id
                WHERE p.vessel_arrival_port IS NOT NULL
                AND TRIM(COALESCE(p.vessel_arrival_port,'')) <> ''
                AND (p.unloading_completed IS NULL OR TRIM(p.unloading_completed)='')
                AND (p.vessel_cast_off IS NULL OR TRIM(p.vessel_cast_off)='')
            """)
            for r in cur.fetchall():
                r = dict(r)
                if r.get("key"):
                    active_keys.add(r["key"])

            berth_layout_out = [
                item
                for item in combined
                if (item.get("name") or "").strip().upper() in active_keys
            ]
        else:
            berth_source = "none"

        # ══════════════════════════════════════════════════════════════
        # WAITING AREA — ALWAYS LIVE, NEVER SAVED / NEVER CARRIED
        # ══════════════════════════════════════════════════════════════
        live_barges, _ = _fetch_all_barges()
        live_map = {
            (b["name"] or "").strip().upper(): b
            for b in live_barges
        }
        waiting_area_out = [
            {
                'berth': 'WAITING',
                'type': b.get('type', 'BARGE'),
                'name': b.get('name', ''),
                'cargo': b.get('cargo', ''),
                'total': b.get('total_qty', b.get('qty', 0)),
                'balance': b.get('balance_qty', 0),
            }
            for b in live_barges
            if b.get('status') in ('Waiting', 'Under Discharge')
        ]

        saved_names = {
            (item.get("name") or "").strip().upper()
            for item in berth_layout_out
        }

        waiting_area_out = [
            item
            for item in waiting_area_out
            if (item.get("name") or "").strip().upper() not in saved_names
        ]
        for item in berth_layout_out:
            key = (item.get("name") or "").strip().upper()

            if key in live_map:
                live = live_map[key]

                print(
                    "MERGE:",
                    key,
                    "saved =", item.get("balance"),
                    "live =", live.get("balance_qty")
                )

                item["balance"] = live.get("balance_qty", 0)
                item["balance_qty"] = live.get("balance_qty", 0)

                item["total"] = live.get("total_qty", live.get("qty", 0))
                item["total_qty"] = live.get("total_qty", live.get("qty", 0))

                item["cargo"] = live.get("cargo", item.get("cargo", ""))
                item["status"] = live.get("status", item.get("status"))

        berth_layout_out.extend(waiting_area_out)

        # ══════════════════════════════════════════════════════════════
        # TRACK 2 — DATE-SPECIFIC META FIELDS (unchanged)
        # ══════════════════════════════════════════════════════════════
        meta_candidates = [r for r in all_rows if row_key(r) <= requested_key]
        meta_candidates.sort(key=row_key)  # oldest -> newest

        merged_wt_r19 = {}
        merged_mbc_eta = {}
        merged_eta_dharamtar = {}
        merged_gull = {}
        merged_notes = []
        latest_meta_updated_at = None
        latest_notes_row = None

        for r in meta_candidates:
            for k, v in (r.get('wt_r19') or {}).items():
                if v:
                    merged_wt_r19[k] = v

            for k, v in (r.get('mbc_eta') or {}).items():
                if v:
                    merged_mbc_eta[k] = v

            for k, v in (r.get('eta_to_dharamtar') or {}).items():
                if v:
                    merged_eta_dharamtar[k] = v

            for k, v in (r.get('on_the_way_gull') or {}).items():
                if v:
                    merged_gull[k] = v

            # Keep only the latest notes
            if r.get("notes"):
                latest_notes_row = r

            if r.get("updated_at"):
                latest_meta_updated_at = r.get("updated_at")

        # Carry forward only the latest saved notes
        if latest_notes_row:
            merged_notes = latest_notes_row.get("notes") or []

        found = (berth_source == 'own')
        carried = (berth_source == 'carried')
        updated_at = berth_updated_at or latest_meta_updated_at

        result = {
            'found': found,
            'carried': carried,
            'berth_source': berth_source,
            'shift_incharge': shift_incharge_out,
            'bpo': bpo_out,
            'crane_operator': crane_operator_out,
            'berth_layout': berth_layout_out,
            'notes': merged_notes,
            'wt_r19': merged_wt_r19,
            'mbc_eta': merged_mbc_eta,
            'eta_to_dharamtar': merged_eta_dharamtar,
            'on_the_way_gull': merged_gull,
            'shift_plan': (exact_row.get('shift_plan') or {}) if exact_row else {},
            'movement_logs': movement_logs_out,
            'updated_at': str(updated_at or ''),
        }
        print("LIVE BARGES:", len(live_barges))
        for b in live_barges:
            if b["name"] == "SURLA":
                  print("LIVE:", b)
            print(
                b["name"],
                b["type"],
                b["status"]
            )

        print("WAITING AREA:", len(waiting_area_out))
        for b in waiting_area_out:
            print(
                b["name"],
                b["type"],
                
            )
        return jsonify(result)

    except Exception as e:
        return jsonify({'found': False, 'error': str(e)}), 500

    finally:
        cur.close()
        conn.close()
    
@bp.route('/api/module/RP01/download-barge-position-excel')
@login_required
def download_barge_position_excel():
    report_date = request.args.get('date', '')
    shift          = request.args.get('shift', 'ALL')
    shift_incharge = request.args.get('shift_incharge', '')
    bpo            = request.args.get('bpo', '')
    operator       = request.args.get('operator', '')

    if not report_date:
        return jsonify({'error': 'date is required'}), 400

    # ── SHIFT WINDOW (same logic as frontend SHIFT_WINDOWS) ────────────────
    SHIFT_WINDOWS = {
        'ALL': {'fh': 6,  'fm': 0, 'th': 6,  'tm': 0, 'next_day': True},
        'A':   {'fh': 6,  'fm': 0, 'th': 14, 'tm': 0, 'next_day': False},
        'B':   {'fh': 14, 'fm': 0, 'th': 22, 'tm': 0, 'next_day': False},
        'C':   {'fh': 22, 'fm': 0, 'th': 6,  'tm': 0, 'next_day': True},
    }
    win  = SHIFT_WINDOWS.get(shift.upper(), SHIFT_WINDOWS['ALL'])
    base = datetime.strptime(report_date, '%Y-%m-%d')
    from_dt = base.replace(hour=win['fh'], minute=win['fm'], second=0)
    to_base = base + timedelta(days=1) if win['next_day'] else base
    to_dt   = to_base.replace(hour=win['th'], minute=win['tm'], second=0)

    # ── DATA PULLS ───────────────────────────────────────────────────────
    mother_vessels = _fetch_mother_vessels(from_dt, to_dt)
    tide_data      = _fetch_tide_data(from_dt, to_dt)
    barges, occupied_berth_set = _fetch_all_barges()
    waiting     = [b for b in barges if b['status'] in ('Waiting', 'Under Discharge')]
    discharging = [b for b in barges if b['status'] == 'Discharging']

    conn = get_db()
    cur  = get_cursor(conn)

    # ── SAVED REPORT OVERLAY (editable fields, notes, logs, plan, layout) ──
    saved = {}
    try:
        cur.execute("""
            SELECT * FROM barge_position_report
            WHERE report_date = %s AND shift = %s
        """, (report_date, shift))
        row = cur.fetchone()
        if row:
            saved = dict(row)
    except Exception:
        saved = {}

    wt_r19_map          = saved.get('wt_r19') or {}
    mbc_eta_map          = saved.get('mbc_eta') or {}
    eta_dharamtar_map    = saved.get('eta_to_dharamtar') or {}
    on_the_way_gull_map  = saved.get('on_the_way_gull') or {}
    notes_saved          = saved.get('notes') or []
    movement_logs        = saved.get('movement_logs') or []
    shift_plan           = saved.get('shift_plan') or {}
    berth_layout_saved   = saved.get('berth_layout') or []
    waiting_area_saved   = saved.get('waiting_area') or []

    for v in mother_vessels:
        name = v.get('vessel_name') or ''

        backend_mbc_eta   = v.get('mbc_eta', '')
        backend_eta_dhrmt = v.get('eta_to_dharamtar', '')

        # WT @ R19 and ON THE WAY TO GULL have no live backend source —
        # always use the saved/carried user-entered value.
        v['wt_r19']          = wt_r19_map.get(name, '')
        v['on_the_way_gull']  = on_the_way_gull_map.get(name, '')

        # MBC ETA and ETA TO DHARAMTAR DO have a live backend value.
        # Prefer that if it exists; only fall back to the saved/user value
        # when the backend has nothing for this vessel right now.
        v['mbc_eta'] = backend_mbc_eta if backend_mbc_eta else mbc_eta_map.get(name, '')

        v['eta_to_dharamtar'] = (
            backend_eta_dhrmt if backend_eta_dhrmt
            else eta_dharamtar_map.get(name, v.get('at_gull_loaded', ''))
        )

    # If a report was saved for this date/shift, its berth layout wins
    use_saved_layout = bool(berth_layout_saved or waiting_area_saved)

    # ── SHIFT WISE DISCHARGE (jetty / barge / mbc) ──────────────────────────
    is_all = shift.upper() == 'ALL'
    if is_all:
        cur.execute("""
            SELECT cargo_name, COALESCE(SUM(quantity), 0) AS qty
            FROM lueu_lines
            WHERE entry_date = %s AND quantity > 0
              AND cargo_name IS NOT NULL AND cargo_name != '' AND is_deleted IS NOT TRUE
            GROUP BY cargo_name ORDER BY cargo_name
        """, (report_date,))
    else:
        cur.execute("""
            SELECT cargo_name, COALESCE(SUM(quantity), 0) AS qty
            FROM lueu_lines
            WHERE entry_date = %s AND shift = %s AND quantity > 0
              AND cargo_name IS NOT NULL AND cargo_name != '' AND is_deleted IS NOT TRUE
            GROUP BY cargo_name ORDER BY cargo_name
        """, (report_date, shift))
    jetty_rows = [dict(r) for r in cur.fetchall()]

    if is_all:
        cur.execute("""
            SELECT source_id, source_type, delay_name, barge_name
            FROM lueu_lines
            WHERE entry_date = %s AND is_deleted IS NOT TRUE
              AND delay_name IS NOT NULL AND delay_name != ''
              AND (LOWER(delay_name) LIKE '%%payloader%%' OR LOWER(delay_name) LIKE '%%labor cleaning%%')
        """, (report_date,))
    else:
        cur.execute("""
            SELECT source_id, source_type, delay_name, barge_name
            FROM lueu_lines
            WHERE entry_date = %s AND shift = %s AND is_deleted IS NOT TRUE
              AND delay_name IS NOT NULL AND delay_name != ''
              AND (LOWER(delay_name) LIKE '%%payloader%%' OR LOWER(delay_name) LIKE '%%labor cleaning%%')
        """, (report_date, shift))
    delay_map = {}
    for r in cur.fetchall():
        key = (r['source_id'], r['source_type'])
        d = (r['delay_name'] or '').strip().lower()
        e = delay_map.setdefault(key, {'payloader': False, 'labour': False})
        if 'payloader' in d: e['payloader'] = True
        if 'labor cleaning' in d or 'labour cleaning' in d: e['labour'] = True

    if is_all:
        cur.execute("""
            WITH actual AS (
                SELECT TRIM(UPPER(barge_name)) AS barge_key, source_id,
                       SUM(COALESCE(quantity,0)) AS actual_qty
                FROM lueu_lines
                WHERE is_deleted IS NOT TRUE AND source_type = 'VCN' AND entry_date = %s
                GROUP BY 1, 2 HAVING SUM(COALESCE(quantity,0)) > 0
            )
            SELECT bl.barge_name, bl.cargo_name,
                   COALESCE(bl.discharge_quantity, 0) AS bl_qty,
                   COALESCE(a.actual_qty, 0) AS actual_discharge, h.vcn_id
            FROM ldud_barge_lines bl
            JOIN ldud_header h ON h.id = bl.ldud_id
            LEFT JOIN actual a
                ON a.barge_key = TRIM(UPPER(CONCAT(bl.barge_name, ' / ', COALESCE(bl.trip_number::text,'1'))))
               AND a.source_id = h.vcn_id
            WHERE COALESCE(TRIM(bl.barge_name),'') <> '' AND COALESCE(a.actual_qty,0) > 0
            ORDER BY bl.barge_name
        """, (report_date,))
    else:
        cur.execute("""
            WITH actual AS (
                SELECT TRIM(UPPER(barge_name)) AS barge_key, source_id,
                       SUM(COALESCE(quantity,0)) AS actual_qty
                FROM lueu_lines
                WHERE is_deleted IS NOT TRUE AND source_type = 'VCN'
                  AND entry_date = %s AND shift = %s
                GROUP BY 1, 2
            )
            SELECT bl.barge_name, bl.cargo_name,
                   COALESCE(bl.discharge_quantity, 0) AS bl_qty,
                   COALESCE(a.actual_qty, 0) AS actual_discharge, h.vcn_id
            FROM ldud_barge_lines bl
            JOIN ldud_header h ON h.id = bl.ldud_id
            INNER JOIN actual a
                ON a.barge_key = TRIM(UPPER(CONCAT(bl.barge_name, ' / ', COALESCE(bl.trip_number::text,'1'))))
               AND a.source_id = h.vcn_id
            WHERE COALESCE(TRIM(bl.barge_name),'') <> '' AND COALESCE(a.actual_qty,0) > 0
            ORDER BY bl.barge_name
        """, (report_date, shift))

    barge_discharge = []
    for r in cur.fetchall():
        r = dict(r)
        d = delay_map.get((r['vcn_id'], 'VCN'), {'payloader': False, 'labour': False})
        barge_discharge.append({
            'type': 'BARGE', 'name': r['barge_name'], 'cargo': r.get('cargo_name') or '',
            'bl_qty': float(r['bl_qty'] or 0), 'actual_discharge': float(r['actual_discharge'] or 0),
            'payloader_cl': r['barge_name'] if d['payloader'] else '',
            'labour_cleaned': r['barge_name'] if d['labour'] else '',
        })

    if is_all:
        cur.execute("""
            SELECT l.source_id AS id,
                   COALESCE(NULLIF(TRIM(l.barge_name), ''), h.mbc_name) AS mbc_name,
                   COALESCE(l.cargo_name, h.cargo_name) AS cargo_name,
                   COALESCE(h.bl_quantity, 0) AS bl_qty,
                   SUM(COALESCE(l.quantity,0)) AS actual_discharge
            FROM lueu_lines l JOIN mbc_header h ON h.id = l.source_id
            WHERE l.source_type = 'MBC' AND l.is_deleted IS NOT TRUE AND l.entry_date = %s
            GROUP BY l.source_id, COALESCE(NULLIF(TRIM(l.barge_name), ''), h.mbc_name),
                     COALESCE(l.cargo_name, h.cargo_name), h.bl_quantity
            HAVING SUM(COALESCE(l.quantity,0)) > 0
            ORDER BY COALESCE(NULLIF(TRIM(l.barge_name), ''), h.mbc_name)
        """, (report_date,))
    else:
        cur.execute("""
            SELECT l.source_id AS id,
                   COALESCE(NULLIF(TRIM(l.barge_name), ''), h.mbc_name) AS mbc_name,
                   COALESCE(l.cargo_name, h.cargo_name) AS cargo_name,
                   COALESCE(h.bl_quantity, 0) AS bl_qty,
                   SUM(COALESCE(l.quantity,0)) AS actual_discharge
            FROM lueu_lines l JOIN mbc_header h ON h.id = l.source_id
            WHERE l.source_type = 'MBC' AND l.is_deleted IS NOT TRUE
              AND l.entry_date = %s AND l.shift = %s
            GROUP BY l.source_id, COALESCE(NULLIF(TRIM(l.barge_name), ''), h.mbc_name),
                     COALESCE(l.cargo_name, h.cargo_name), h.bl_quantity
            HAVING SUM(COALESCE(l.quantity,0)) > 0
            ORDER BY COALESCE(NULLIF(TRIM(l.barge_name), ''), h.mbc_name)
        """, (report_date, shift))

    mbc_discharge = []
    for r in cur.fetchall():
        r = dict(r)
        d = delay_map.get((r['id'], 'MBC'), {'payloader': False, 'labour': False})
        mbc_discharge.append({
            'type': 'MBC', 'name': r['mbc_name'], 'cargo': r.get('cargo_name') or '',
            'bl_qty': float(r['bl_qty'] or 0), 'actual_discharge': float(r['actual_discharge'] or 0),
            'payloader_cl': r['mbc_name'] if d['payloader'] else '',
            'labour_cleaned': r['mbc_name'] if d['labour'] else '',
        })

    barge_rows_combined = barge_discharge + mbc_discharge

    # ── SHIFT SUMMARY (A/B/C discharge totals) ─────────────────────────────
    shift_discharge_totals = {'A': 0, 'B': 0, 'C': 0}
    for s in ['A', 'B', 'C']:
        cur.execute("""
            WITH actual AS (
                SELECT TRIM(UPPER(barge_name)) AS barge_key, source_id,
                       SUM(COALESCE(quantity,0)) AS actual_qty
                FROM lueu_lines
                WHERE is_deleted IS NOT TRUE AND source_type = 'VCN'
                  AND entry_date = %s AND shift = %s
                GROUP BY 1, 2
            )
            SELECT COALESCE(SUM(a.actual_qty),0) AS total FROM actual a
        """, (report_date, s))
        barge_total = float(cur.fetchone()['total'] or 0)

        cur.execute("""
            SELECT COALESCE(SUM(quantity),0) AS total
            FROM lueu_lines
            WHERE source_type = 'MBC' AND is_deleted IS NOT TRUE
              AND entry_date = %s AND shift = %s
        """, (report_date, s))
        mbc_total = float(cur.fetchone()['total'] or 0)

        shift_discharge_totals[s] = barge_total + mbc_total

    total_discharge = sum(shift_discharge_totals.values())
    a_plan = float(shift_plan.get('a_plan', 0) or 0)
    b_plan = float(shift_plan.get('b_plan', 0) or 0)
    c_plan = float(shift_plan.get('c_plan', 0) or 0)
    total_plan = a_plan + b_plan + c_plan

    slag_qty = sum(r['actual_discharge'] for r in barge_rows_combined if 'slag' in (r['cargo'] or '').lower())
    clinker_qty = sum(r['actual_discharge'] for r in barge_rows_combined if (r['cargo'] or '').strip().lower() == 'clinker')
    slag_clinker_total = slag_qty + clinker_qty
    steel_plant = total_discharge - slag_clinker_total

    cur.close()
    conn.close()

    # ── BERTH LAYOUT (matrix + waiting) ─────────────────────────────────
    old_berths = ["BERTH 1", "BERTH 2", "BERTH 3", "BERTH 4", "BERTH 5", "BERTH 5A"]
    new_berths = ["BERTH 6", "BERTH 7", "BERTH 8", "BERTH 8A", "BERTH 9", "BERTH 10", "BERTH 11", "BERTH 12"]
    positions  = ['A/S', 'D/B', 'T/B', 'F/B', 'S/B']

    matrix = {}   # (berth, position) -> item dict
    waiting_list = []

    if use_saved_layout:
        for item in berth_layout_saved:
            matrix[(item.get('berth'), item.get('position'))] = item
        waiting_list = waiting_area_saved
    else:
        for b in discharging:
            berth = (b.get('berth') or '').upper()
            if berth:
                matrix[(berth, 'A/S')] = b
        waiting_list = waiting

    # ══════════════════════════════════════════════════════════════════
    #  BUILD WORKBOOK  (side-by-side layout, bordered/centered titles)
    # ══════════════════════════════════════════════════════════════════
    wb = Workbook()
    ws = wb.active
    ws.title = "Barge Position Report"
    ws.sheet_view.showGridLines = False

    # ── Colors ───────────────────────────────────────────────────────
    C_TEXT_PRIMARY   = "0F172A"
    C_TEXT_MUTED     = "64748B"
    C_BORDER         = "000000"   # solid black grid lines to match on-screen report
    C_APP_BG         = "F8FAFC"

    C_WAITING_BG     = "FEF08A"
    C_WAITING_TEXT   = "713F12"

    C_DISCHARGE_BG   = "32DF6E"
    C_DISCHARGE_TEXT = "14532D"

    C_WAITING_DISCHARGE_BG = "43E6F1"   # cyan — alongside but discharge not started yet

    C_VESSEL_HEADER  = "4D8CCD"
    C_SECTION_TITLE  = "1D4ED8"

    C_METRIC_WAITING   = "B45309"
    C_METRIC_DISCHARGE = "166534"
    C_METRIC_OCCUPIED  = "1D4ED8"

    C_TIDE_HW_BG   = "DCFCE7"; C_TIDE_HW_TX = "166534"
    C_TIDE_LW_BG   = "DBEAFE"; C_TIDE_LW_TX = "1E40AF"

    C_BADGE_BARGE_BG = "FEF9C3"; C_BADGE_BARGE_TX = "92400E"
    C_BADGE_MBC_BG   = "DBEAFE"; C_BADGE_MBC_TX   = "1E40AF"

    C_COMPLETED_BG   = "DCFCE7"; C_COMPLETED_TX = "166534"

    thin = Side(style='thin', color=C_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # title_border now identical to border since both are black — kept as a
    # separate name so section-header calls (which pass brdr=title_border)
    # still resolve to the same solid black grid.
    title_border = border

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a = Alignment(horizontal='left', vertical='center', wrap_text=True)

    title_font       = Font(bold=True, size=18, color=C_TEXT_PRIMARY)
    meta_font        = Font(size=11, color=C_TEXT_PRIMARY)
    section_font     = Font(bold=True, size=13, color=C_SECTION_TITLE)
    header_font      = Font(bold=True, size=10, color="FFFFFF")
    grey_header_font = Font(bold=True, size=9, color=C_TEXT_MUTED)
    grey_header_fill = PatternFill("solid", fgColor=C_APP_BG)
    vessel_header_fill = PatternFill("solid", fgColor=C_VESSEL_HEADER)
    waiting_fill  = PatternFill("solid", fgColor=C_WAITING_BG)
    discharge_fill= PatternFill("solid", fgColor=C_DISCHARGE_BG)
    waiting_discharge_fill = PatternFill("solid", fgColor=C_WAITING_DISCHARGE_BG)
    empty_fill    = PatternFill("solid", fgColor="F1F5F9")
    total_fill    = PatternFill("solid", fgColor=C_APP_BG)
    card_fill     = PatternFill("solid", fgColor="FFFFFF")
    tide_hw_fill  = PatternFill("solid", fgColor=C_TIDE_HW_BG)
    tide_lw_fill  = PatternFill("solid", fgColor=C_TIDE_LW_BG)
    completed_fill= PatternFill("solid", fgColor=C_COMPLETED_BG)
    note_num_fill = PatternFill("solid", fgColor="E2E8F0")

    def put(r, c, value, font=None, fill=None, align=center, brdr=border):
        cell = ws.cell(row=r, column=c, value=value)
        cell.font = font or Font(size=10, color=C_TEXT_PRIMARY)
        if fill: cell.fill = fill
        cell.alignment = align or center
        if brdr: cell.border = brdr
        return cell

    def merge(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        # apply border to every cell in the merged range so the black grid
        # doesn't visually "cut off" partway through a merged block
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(row=rr, column=cc).border = title_border

    # ── Row 1: Title (centered, no border) ────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    put(1, 1, "Daily Barge Position Report — RP01", font=title_font, brdr=None, align=Alignment(horizontal='center'))

    # ── Row 2: Date / Shift / Doc info (bordered mini-table) ─────────
    put(2, 1, "Date:", font=Font(bold=True, size=11), brdr=title_border, align=Alignment(horizontal='left'))
    put(2, 2, f" {report_date}  ", font=meta_font, brdr=title_border, align=Alignment(horizontal='left'))
    put(2, 3, " Shift: ", font=Font(bold=True, size=11), brdr=title_border, align=Alignment(horizontal='left'))
    put(2, 4, f"{shift} -shift", font=meta_font, brdr=title_border, align=Alignment(horizontal='left'))
    merge(2, 5, 2, 6)
    put(2, 5, "DOC NO.OPE/0100/F/01", font=Font(size=9, color=C_TEXT_MUTED), brdr=title_border, align=Alignment(horizontal='left'))
    put(2, 7, "ISUUENO.02", font=Font(size=9, color=C_TEXT_MUTED), brdr=title_border, align=Alignment(horizontal='left'))
    merge(2, 8, 2, 9)
    put(2, 8, "ISUUE DATE: 01.04.2022", font=Font(size=9, color=C_TEXT_MUTED), brdr=title_border, align=Alignment(horizontal='left'))

    # ── Row 3: spacer ──────────────────────────────────────────────────
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)
    put(3, 1, "", brdr=None)

    # ── Row 4: Shift Incharge (full-width merge, wrapped, dynamic height) ──
    put(4, 1, " Shift Incharge:", font=Font(bold=True, size=11), brdr=title_border, align=Alignment(horizontal='left', vertical='center'))
    merge(4, 2, 4, 14)
    put(4, 2, f" {shift_incharge}", font=meta_font, brdr=title_border, align=left_a)

    # ── Row 5: Crane Operator / BPO (full-width merge, wrapped) ───────
    put(5, 1, " Crane Operator:", font=Font(bold=True, size=11), brdr=title_border, align=Alignment(horizontal='left', vertical='center'))
    merge(5, 2, 5, 14)
    put(5, 2, f" {operator}", font=meta_font, brdr=title_border, align=left_a)

    put(6, 1, " BPO:", font=Font(bold=True, size=11), brdr=title_border, align=Alignment(horizontal='left', vertical='center'))
    merge(6, 2, 6, 14)
    put(6, 2, f" {bpo}", font=meta_font, brdr=title_border, align=left_a)

    # Dynamically size rows 4-6 so long comma-separated names wrap cleanly
    # inside their bordered cell instead of overflowing past it.
    def est_row_height(text, chars_per_line=170):
        lines = max(1, -(-len(str(text or '')) // chars_per_line))
        return max(18, lines * 14)

    ws.row_dimensions[4].height = est_row_height(shift_incharge)
    ws.row_dimensions[5].height = est_row_height(operator)
    ws.row_dimensions[6].height = est_row_height(bpo)

    # ── Row 8: spacer ──────────────────────────────────────────────────
    row = 8

    # ── Row 8: berth section titles (bordered, centered, side by side) ──
    berth_title_row = row
    merge(berth_title_row, 1, berth_title_row, 6)
    put(berth_title_row, 1, f"OLD BERTH  ({len(old_berths)} BERTH)", font=section_font, brdr=title_border, align=center)
    merge(berth_title_row, 8, berth_title_row, 13)
    put(berth_title_row, 8, f"NEW BERTHS  ({len(new_berths)} BERTH)", font=section_font, brdr=title_border, align=center)
    row += 1

    # ── berth headers ───────────────────────────────────────
    header_row_berth = row
    put(header_row_berth, 1, "BERTH", font=grey_header_font, fill=grey_header_fill)
    for i, p in enumerate(positions):
        put(header_row_berth, i + 2, p, font=grey_header_font, fill=grey_header_fill)
    put(header_row_berth, 8, "BERTH", font=grey_header_font, fill=grey_header_fill)
    for i, p in enumerate(positions):
        put(header_row_berth, i + 9, p, font=grey_header_font, fill=grey_header_fill)
    row += 1

    # ── berth matrices side by side ─────────────────────────
    def berth_cell_text(item):
        return (f"⚓ {item.get('type','BARGE')} — {item.get('name','')}\n"
                f"{item.get('cargo','')}\n"
                f"BL: {item.get('total', item.get('qty',0))} MT   "
                f"Bal: {item.get('balance',0)} MT")

    def berth_cell_fill(item):
        unloading_commenced = str(item.get('unloading_commenced') or '').strip()
        commence_discharge_berth = str(item.get('commence_discharge_berth') or '').strip()
        unload_started = bool(unloading_commenced) or bool(commence_discharge_berth)
        return discharge_fill if unload_started else waiting_discharge_fill

    berth_matrix_start = row
    max_berth_rows = max(len(old_berths), len(new_berths))
    for r_i in range(max_berth_rows):
        r = berth_matrix_start + r_i
        ws.row_dimensions[r].height = 46
        if r_i < len(old_berths):
            b = old_berths[r_i]
            put(r, 1, b, font=Font(bold=True, size=9), align=left_a)
            for i, p in enumerate(positions):
                item = matrix.get((b, p))
                if item:
                    put(r, i + 2, berth_cell_text(item), font=Font(size=8, bold=True, color=C_DISCHARGE_TEXT), fill=berth_cell_fill(item), align=left_a)
                else:
                    put(r, i + 2, "—", font=Font(size=9, color="94A3B8"), fill=empty_fill)
        if r_i < len(new_berths):
            b = new_berths[r_i]
            put(r, 8, b, font=Font(bold=True, size=9), align=left_a)
            for i, p in enumerate(positions):
                item = matrix.get((b, p))
                if item:
                    put(r, i + 9, berth_cell_text(item), font=Font(size=8, bold=True, color=C_DISCHARGE_TEXT), fill=berth_cell_fill(item), align=left_a)
                else:
                    put(r, i + 9, "—", font=Font(size=9, color="94A3B8"), fill=empty_fill)

    row = berth_matrix_start + max_berth_rows + 2  # spacer then section title row

    # ── Mother Vessel + Waiting Area (side by side) ──────────────────
    mv_title_row = row
    merge(mv_title_row, 1, mv_title_row, 8)
    put(mv_title_row, 1, "MOTHER VESSEL", font=section_font, brdr=title_border, align=center)
    merge(mv_title_row, 10, mv_title_row, 14)
    put(mv_title_row, 10, f"WAITING AREA  ({len(waiting_list)})", font=section_font, brdr=title_border, align=center)
    row += 1

    header_row = row
    put(header_row, 1, "Parameter", font=header_font, fill=vessel_header_fill)
    for i, v in enumerate(mother_vessels):
        put(header_row, i + 2, f"Vessel {i+1}\n{v.get('vessel_name','')}", font=header_font, fill=vessel_header_fill)
    for i, h in enumerate(["Type", "Name", "Cargo", "BL Qty (MT)", "Balance (MT)"]):
        put(header_row, i + 10, h, font=grey_header_font, fill=grey_header_fill)
    row += 1

    mv_rows = [
        ("VSL DISCH COMMNACED", 'discharge_commenced'),
        ("VSL DISCHARGE COMPLITED", 'discharge_completed'),
        ("UNDER LOADING", 'under_loading'),
        ("ETA TO DHARAMTAR", 'eta_to_dharamtar'),
        ("WT @ R19", 'wt_r19'),
        ("ON THE WAY TO GULL", 'on_the_way_gull'),
        ("MBC ETA", 'mbc_eta'),
    ]
    mv_end_row = row + len(mv_rows) - 1
    waiting_end_row = row + len(waiting_list) - 1
    for idx, (label, key) in enumerate(mv_rows):
        r = row + idx
        row_fill = PatternFill("solid", fgColor="FAFAFA") if idx % 2 == 0 else None
        put(r, 1, label, font=Font(bold=True, size=9), fill=row_fill, align=left_a)
        for i, v in enumerate(mother_vessels):
            put(r, i + 2, v.get(key, ''), fill=row_fill, align=left_a)

    for idx, item in enumerate(waiting_list):
        r = row + idx
        put(r, 10, item.get('type', 'BARGE'), font=Font(bold=True, size=9, color=C_WAITING_TEXT), fill=waiting_fill)
        put(r, 11, item.get('name', ''), font=Font(bold=True, size=10, color=C_WAITING_TEXT), fill=waiting_fill, align=left_a)
        put(r, 12, item.get('cargo', ''), font=Font(size=9, color=C_WAITING_TEXT), fill=waiting_fill, align=left_a)
        put(r, 13, float(item.get('total', item.get('total_qty', item.get('qty', 0))) or 0),
            font=Font(bold=True, size=9, color=C_WAITING_TEXT), fill=waiting_fill)
        put(r, 14, float(item.get('balance', item.get('balance_qty', 0)) or 0),
            font=Font(bold=True, size=9, color=C_METRIC_WAITING), fill=waiting_fill)

    # ── FIX: Notes/Tide (columns 1-8) only need to wait for the Mother
    #    Vessel table (also columns 1-8) to finish — NOT for the Waiting
    #    Area (columns 10-14), which is visually separate. Previously this
    #    used max(mv_end_row, waiting_end_row), which left a large empty
    #    gap under the Mother Vessel columns whenever Waiting Area was
    #    longer than Mother Vessel. Waiting Area is allowed to keep
    #    running past this point in its own column range.
    row = mv_end_row + 2

    # ── Notes + Tide Table (side by side) ─────────────────────────────
    nt_title_row = row
    merge(nt_title_row, 1, nt_title_row, 4)
    put(nt_title_row, 1, "NOTES", font=section_font, brdr=title_border, align=center)
    merge(nt_title_row, 6, nt_title_row, 8)
    put(nt_title_row, 6, "TIDE TABLE", font=section_font, brdr=title_border, align=center)
    row += 1

    tide_header_row = row
    for i, h in enumerate(["Type", "Time", "Height (m)"]):
        put(tide_header_row, i + 6, h, font=grey_header_font, fill=grey_header_fill)
    row += 1

    # ── NOTES LOOP ────────────────────────────────────────────
    notes_list = notes_saved or ["3B to 5A plug problem — informed electrical Mr. Koli."]
    for i, n in enumerate(notes_list, start=1):
        r = row + i - 1
        est_lines = max(1, -(-len(n) // 55))   # ceil division
        ws.row_dimensions[r].height = max(22, est_lines * 14)

        put(r, 1, i, font=Font(bold=True, size=9, color="1E3A8A"), fill=note_num_fill)

        merge(r, 2, r, 4)
        put(r, 2, n, font=Font(size=9), fill=grey_header_fill, align=left_a)

    for i, t in enumerate(tide_data):
        r = row + i
        tag = (t.get('type') or '').upper()
        tfill = tide_hw_fill if tag == 'HW' else tide_lw_fill
        ttext = C_TIDE_HW_TX if tag == 'HW' else C_TIDE_LW_TX
        put(r, 6, tag, font=Font(bold=True, size=9, color=ttext), fill=tfill)
        put(r, 7, t.get('time', ''), font=Font(size=9))
        put(r, 8, t.get('height', ''), font=Font(size=9))

    notes_tide_end_row = row + max(len(notes_list), len(tide_data)) - 1

    # ── Now reconcile with Waiting Area, since the Shift Wise Discharge
    #    section below spans the FULL width again (columns 1-8) and must
    #    not start until both the Notes/Tide block AND the Waiting Area
    #    (which may still be running in columns 10-14) have finished.
    row = max(notes_tide_end_row, waiting_end_row) + 2

    # ── Shift Wise Discharge Report ───────────────────────────────────
    merge(row, 1, row, 8)
    put(row, 1, "SHIFT WISE DISCHARGE REPORT", font=section_font, brdr=title_border, align=center)
    row += 1

    header_row1 = row
    merge(header_row1, 1, header_row1, 2)
    put(header_row1, 1, "SHIFT JETTY DISCHARGE", font=grey_header_font, fill=grey_header_fill)
    merge(header_row1, 3, header_row1, 8)
    put(header_row1, 3, "BARGE DISCHARGE", font=grey_header_font, fill=grey_header_fill)
    row += 1

    header_row2 = row
    labels2 = ["CARGO", "QTY", "BARGES AND MBC", "CARGO", "BL QTY", "ACTUAL DISCHARGE", "PAYLOADER CL", "LABOUR CLEANED"]
    for i, h in enumerate(labels2):
        put(header_row2, i + 1, h, font=grey_header_font, fill=grey_header_fill)
    row += 1

    payloader_names = [r['payloader_cl'] for r in barge_rows_combined if r.get('payloader_cl')]
    labour_names    = [r['labour_cleaned'] for r in barge_rows_combined if r.get('labour_cleaned')]

    max_rows = max(len(jetty_rows), len(barge_rows_combined), 1)
    jetty_total = bl_total = actual_total = 0
    for i in range(max_rows):
        j = jetty_rows[i] if i < len(jetty_rows) else None
        b = barge_rows_combined[i] if i < len(barge_rows_combined) else None
        if j: jetty_total += float(j['qty'] or 0)
        if b: bl_total += b['bl_qty']; actual_total += b['actual_discharge']

        put(row, 1, j['cargo_name'] if j else '', font=Font(size=9), align=left_a)
        put(row, 2, float(j['qty']) if j else '', font=Font(size=9))

        if b:
            badge_fill = C_BADGE_MBC_BG if b['type'] == 'MBC' else C_BADGE_BARGE_BG
            badge_tx   = C_BADGE_MBC_TX if b['type'] == 'MBC' else C_BADGE_BARGE_TX
            cell = put(row, 3, f"{b['type']}  {b['name']}", font=Font(size=9, bold=True, color=badge_tx), align=left_a)
            cell.fill = PatternFill("solid", fgColor=badge_fill)
        else:
            put(row, 3, '', font=Font(size=9), align=left_a)

        put(row, 4, b['cargo'] if b else '', font=Font(size=9), align=left_a)
        put(row, 5, b['bl_qty'] if b else '', font=Font(size=9))
        put(row, 6, b['actual_discharge'] if b else '', font=Font(bold=True, size=9))
        put(row, 7, payloader_names[i] if i < len(payloader_names) else '', font=Font(size=8, color=C_TEXT_MUTED))
        put(row, 8, labour_names[i] if i < len(labour_names) else '', font=Font(size=8, color=C_TEXT_MUTED))
        row += 1

    put(row, 1, "TOTAL", font=Font(bold=True, size=11), fill=total_fill)
    put(row, 2, jetty_total, font=Font(bold=True, size=11), fill=total_fill)
    merge(row, 3, row, 4)
    put(row, 3, '', fill=total_fill)
    put(row, 5, bl_total, font=Font(bold=True, size=11), fill=total_fill)
    put(row, 6, actual_total, font=Font(bold=True, size=11), fill=total_fill)
    put(row, 7, '', fill=total_fill)
    put(row, 8, '', fill=total_fill)
    row += 2

    # ── Shift summary + Slag/Clinker (side by side) ─────────────────
    summary_top = row
    for i, h in enumerate(["SHIFT", "DISCHARGE", "PLAN", "DIFF"]):
        put(summary_top, i + 1, h, font=grey_header_font, fill=grey_header_fill)

    put(summary_top, 6, "SLAG", font=Font(bold=True, size=9), fill=total_fill, align=left_a)
    put(summary_top, 7, slag_qty, font=Font(size=9), fill=card_fill)
    row += 1

    for s, plan in zip(['A', 'B', 'C'], [a_plan, b_plan, c_plan]):
        put(row, 1, s, font=Font(size=9))
        put(row, 2, shift_discharge_totals[s], font=Font(size=9))
        put(row, 3, plan if plan else '', font=Font(size=9))
        put(row, 4, (plan - shift_discharge_totals[s]) if plan else '', font=Font(size=9))
        row += 1

    put(summary_top + 1, 6, "CLINKER", font=Font(bold=True, size=9), fill=total_fill, align=left_a)
    put(summary_top + 1, 7, clinker_qty, font=Font(size=9), fill=card_fill)
    put(summary_top + 2, 6, "TOTAL", font=Font(bold=True, size=9), fill=total_fill, align=left_a)
    put(summary_top + 2, 7, slag_clinker_total, font=Font(size=9), fill=card_fill)

    put(row, 1, "TOTAL", font=Font(bold=True), fill=total_fill)
    put(row, 2, total_discharge, font=Font(bold=True), fill=total_fill)
    put(row, 3, total_plan if total_plan else '', font=Font(bold=True), fill=total_fill)
    put(row, 4, (total_plan - total_discharge) if total_plan else '', font=Font(bold=True), fill=total_fill)
    put(row, 6, "STEEL PLANT", font=Font(bold=True, size=9), fill=total_fill, align=left_a)
    put(row, 7, steel_plant, font=Font(bold=True, size=9), fill=card_fill)
    row += 3

    # ── Column widths ──────────────────────────────────────────────
    widths = {
        'A': 20, 'B': 20, 'C': 19, 'D': 20, 'E': 18, 'F': 18,
        'G': 16, 'H': 14, 'I': 15, 'J': 16, 'K': 20, 'L': 16, 'M': 16, 'N': 14
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A8"

    # ── Output ─────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Daily_Barge_Position_Report_{report_date}_{shift}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
@bp.route('/api/module/RP01/download-movement-logs-excel', methods=['POST'])
@login_required
def download_movement_logs_excel():
    data = request.get_json() or {}
    logs = data.get('movement_logs', [])
    report_date = data.get('report_date', '')
    shift = data.get('shift', 'ALL')

    wb = Workbook()
    ws = wb.active
    ws.title = "Movement Logs"
    ws.sheet_view.showGridLines = False

    thin = Side(style='thin', color='D9E2EC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a = Alignment(horizontal='left', vertical='center', wrap_text=True)

    headers = ["Name", "From", "To", "Date", "Time", "Shift", "Shift Incharge"]

    # ── Title row ────────────────────────────────────────────────
    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = f"MOVEMENT LOGS   (Total: {len(logs)})"
    c.font = Font(bold=True, size=16, color="2563EB")
    c.alignment = center
    ws.row_dimensions[1].height = 26

    # ── Header row ───────────────────────────────────────────────
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, size=10, color="6B7280")
        cell.fill = PatternFill("solid", fgColor="F8FAFC")
        cell.alignment = center
        cell.border = border

    # ── Data rows ────────────────────────────────────────────────
    r = 3
    if not logs:
        ws.merge_cells(f'A{r}:G{r}')
        cell = ws.cell(row=r, column=1, value="No movement logs found.")
        cell.font = Font(size=10, color="94A3B8")
        cell.alignment = center
        r += 1
    else:
        for log in logs:
            incharge = (log.get('shiftIncharge') or '').strip() or '—'
            vals = [
                log.get('name', ''), log.get('from', ''), log.get('to', ''),
                log.get('reportDate', ''), log.get('time', ''),
                log.get('shift', ''), incharge
            ]
            for i, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=i, value=v)
                cell.border = border
                cell.alignment = left_a if i in (1, 2, 3, 7) else center
                cell.font = Font(bold=(i == 1), size=10, color="0F172A")
            r += 1

    widths = {'A': 18, 'B': 16, 'C': 16, 'D': 14, 'E': 14, 'F': 10, 'G': 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"Movement_Logs_{report_date or 'report'}_{shift}_{stamp}.xlsx"
    )
    
@bp.route('/api/module/RP01/download-completed-excel')
@login_required
def download_completed_excel():
    report_date = request.args.get('date', '')
    shift = request.args.get('shift', 'ALL')
    if not report_date:
        return jsonify({'error': 'date is required'}), 400

    # ── Fetch completed BARGES + MBC for this date ──────────────────
    conn = get_db()
    cur  = get_cursor(conn)
    items = []

    cur.execute(r"""
        SELECT * FROM (
            SELECT bl.barge_name, bl.cargo_name, bl.commence_discharge_berth,
                   bl.along_side_berth, bl.completed_discharge_berth, bl.cast_off_port,
                   COALESCE(bl.discharge_quantity, 0) AS bl_qty,
                   CASE
                       WHEN bl.cast_off_port IS NOT NULL AND TRIM(bl.cast_off_port) ~ '^\d{4}-\d{2}-\d{2}[T ]'
                       THEN SUBSTRING(TRIM(bl.cast_off_port), 1, 10)::date
                       WHEN bl.completed_discharge_berth IS NOT NULL AND TRIM(bl.completed_discharge_berth) ~ '^\d{4}-\d{2}-\d{2}[T ]'
                       THEN SUBSTRING(TRIM(bl.completed_discharge_berth), 1, 10)::date
                       ELSE NULL
                   END AS completed_date
            FROM ldud_barge_lines bl
            JOIN ldud_header h ON h.id = bl.ldud_id
            WHERE COALESCE(TRIM(bl.barge_name),'') <> ''
              AND ((bl.cast_off_port IS NOT NULL AND TRIM(bl.cast_off_port) ~ '^\d{4}-\d{2}-\d{2}[T ]')
                   OR (bl.completed_discharge_berth IS NOT NULL AND TRIM(bl.completed_discharge_berth) ~ '^\d{4}-\d{2}-\d{2}[T ]'))
        ) sub WHERE sub.completed_date = %s::date
        ORDER BY sub.barge_name
    """, (report_date,))
    for row in cur.fetchall():
        row = dict(row)
        items.append({
            'type': 'BARGE', 'name': row['barge_name'], 'cargo': row.get('cargo_name') or '',
            'bl_qty': float(row['bl_qty'] or 0),
            'commenced': _fmt_dt(row.get('commence_discharge_berth') or row.get('along_side_berth')),
            'completed': _fmt_dt(row.get('cast_off_port') or row.get('completed_discharge_berth')),
        })

    cur.execute(r"""
        SELECT * FROM (
            SELECT h.mbc_name, h.cargo_name, COALESCE(h.bl_quantity, 0) AS bl_qty,
                   p.unloading_commenced, p.unloading_completed, p.vessel_cast_off,
                   CASE
                       WHEN p.unloading_completed IS NOT NULL AND TRIM(p.unloading_completed) ~ '^\d{4}-\d{2}-\d{2}[T ]'
                       THEN SUBSTRING(TRIM(p.unloading_completed), 1, 10)::date
                       WHEN p.vessel_cast_off IS NOT NULL AND TRIM(p.vessel_cast_off) ~ '^\d{4}-\d{2}-\d{2}[T ]'
                       THEN SUBSTRING(TRIM(p.vessel_cast_off), 1, 10)::date
                       ELSE NULL
                   END AS completed_date
            FROM mbc_header h JOIN mbc_discharge_port_lines p ON p.mbc_id = h.id
            WHERE (p.unloading_completed IS NOT NULL AND TRIM(p.unloading_completed) ~ '^\d{4}-\d{2}-\d{2}[T ]')
               OR (p.vessel_cast_off IS NOT NULL AND TRIM(p.vessel_cast_off) ~ '^\d{4}-\d{2}-\d{2}[T ]')
        ) sub WHERE sub.completed_date = %s::date
        ORDER BY sub.mbc_name
    """, (report_date,))
    for row in cur.fetchall():
        row = dict(row)
        items.append({
            'type': 'MBC', 'name': row['mbc_name'], 'cargo': row.get('cargo_name') or '',
            'bl_qty': float(row['bl_qty'] or 0),
            'commenced': _fmt_dt(row.get('unloading_commenced')),
            'completed': _fmt_dt(row.get('unloading_completed') or row.get('vessel_cast_off')),
        })

    cur.close()
    conn.close()

    # ── Build styled workbook ────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Completed Barges & MBC"
    ws.sheet_view.showGridLines = False

    thin = Side(style='thin', color='D9E2EC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a = Alignment(horizontal='left', vertical='center', wrap_text=True)

    headers = ["Type", "Name", "Cargo", "BL Qty (MT)", "Commenced", "Completed", "Status"]

    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = f"✓ COMPLETED BARGES & MBC   (Total: {len(items)})"
    c.font = Font(bold=True, size=16, color="166534")
    c.alignment = center
    ws.row_dimensions[1].height = 26

    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, size=10, color="6B7280")
        cell.fill = PatternFill("solid", fgColor="F8FAFC")
        cell.alignment = center
        cell.border = border

    badge_barge_fill = PatternFill("solid", fgColor="FEF9C3")
    badge_mbc_fill   = PatternFill("solid", fgColor="DBEAFE")
    status_fill      = PatternFill("solid", fgColor="DCFCE7")

    r = 3
    if not items:
        ws.merge_cells(f'A{r}:G{r}')
        cell = ws.cell(row=r, column=1, value="No completed barges / MBC found.")
        cell.font = Font(size=10, color="94A3B8")
        cell.alignment = center
        r += 1
    else:
        for idx, item in enumerate(items):
            row_fill = PatternFill("solid", fgColor="FFFFFF" if idx % 2 == 0 else "F8FAFC")
            is_mbc = item['type'] == 'MBC'

            cell = ws.cell(row=r, column=1, value=item['type'])
            cell.fill = badge_mbc_fill if is_mbc else badge_barge_fill
            cell.font = Font(bold=True, size=9, color="1E40AF" if is_mbc else "92400E")
            cell.alignment = center
            cell.border = border

            cell = ws.cell(row=r, column=2, value=item['name'])
            cell.font = Font(bold=True, size=10, color="0F172A")
            cell.fill = row_fill
            cell.alignment = left_a
            cell.border = border

            cell = ws.cell(row=r, column=3, value=item['cargo'])
            cell.font = Font(size=10, color="0F172A")
            cell.fill = row_fill
            cell.alignment = left_a
            cell.border = border

            cell = ws.cell(row=r, column=4, value=item['bl_qty'])
            cell.font = Font(size=10, color="0F172A")
            cell.fill = row_fill
            cell.alignment = center
            cell.border = border
            cell.number_format = '#,##0'

            cell = ws.cell(row=r, column=5, value=item['commenced'])
            cell.font = Font(size=9, color="1D4ED8")
            cell.fill = row_fill
            cell.alignment = center
            cell.border = border

            cell = ws.cell(row=r, column=6, value=item['completed'])
            cell.font = Font(bold=True, size=9, color="166534")
            cell.fill = row_fill
            cell.alignment = center
            cell.border = border

            cell = ws.cell(row=r, column=7, value="✓ Completed")
            cell.font = Font(bold=True, size=9, color="166534")
            cell.fill = status_fill
            cell.alignment = center
            cell.border = border

            r += 1

    widths = {'A': 12, 'B': 22, 'C': 20, 'D': 14, 'E': 18, 'F': 18, 'G': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"Completed_Barges_MBC_{report_date}_{shift}_{stamp}.xlsx"
    )        
    
@bp.route('/api/module/RP01/download-movement-logs-range-excel')
@login_required
def download_movement_logs_range_excel():
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')

    if not from_date or not to_date:
        return jsonify({'error': 'from_date and to_date are required'}), 400

    conn = get_db()
    cur = get_cursor(conn)

    all_logs = []
    try:
        cur.execute("""
            SELECT report_date, shift, movement_logs
            FROM barge_position_report
            WHERE report_date BETWEEN %s AND %s
              AND movement_logs IS NOT NULL
            ORDER BY report_date ASC, shift ASC
        """, (from_date, to_date))

        for row in cur.fetchall():
            row = dict(row)
            logs = row.get('movement_logs') or []
            for log in logs:
                all_logs.append(log)
    finally:
        cur.close()
        conn.close()

    # Sort chronologically by reportDate + time where available
    def sort_key(log):
        return (log.get('reportDate', ''), log.get('time', ''))
    all_logs.sort(key=sort_key)

    # ── Build workbook (same style as the single-shift logs export) ────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Movement Logs"
    ws.sheet_view.showGridLines = False

    thin = Side(style='thin', color='D9E2EC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a = Alignment(horizontal='left', vertical='center', wrap_text=True)

    headers = ["Name", "From", "To", "Date", "Time", "Shift", "Shift Incharge"]

    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = f"MOVEMENT LOGS   ({from_date} to {to_date})   Total: {len(all_logs)}"
    c.font = Font(bold=True, size=16, color="2563EB")
    c.alignment = center
    ws.row_dimensions[1].height = 26

    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, size=10, color="6B7280")
        cell.fill = PatternFill("solid", fgColor="F8FAFC")
        cell.alignment = center
        cell.border = border

    r = 3
    if not all_logs:
        ws.merge_cells(f'A{r}:G{r}')
        cell = ws.cell(row=r, column=1, value="No movement logs found for this date range.")
        cell.font = Font(size=10, color="94A3B8")
        cell.alignment = center
        r += 1
    else:
        for log in all_logs:
            incharge = (log.get('shiftIncharge') or '').strip() or '—'
            vals = [
                log.get('name', ''), log.get('from', ''), log.get('to', ''),
                log.get('reportDate', ''), log.get('time', ''),
                log.get('shift', ''), incharge
            ]
            for i, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=i, value=v)
                cell.border = border
                cell.alignment = left_a if i in (1, 2, 3, 7) else center
                cell.font = Font(bold=(i == 1), size=10, color="0F172A")
            r += 1

    widths = {'A': 18, 'B': 16, 'C': 16, 'D': 14, 'E': 14, 'F': 10, 'G': 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"Movement_Logs_{from_date}_to_{to_date}_{stamp}.xlsx"
    )   



# ══════════════════════════════════════════════════════════════════════════
#  NEW — Date/Shift-wise Jetty Cargo Report (separate route, does not
#  touch or replace the existing /api/module/RP01/jetty-cargo-report)
# ══════════════════════════════════════════════════════════════════════════

_JETTY_V2_SHIFT_WINDOWS = {
    'ALL': {'fh': 6,  'fm': 0, 'th': 6,  'tm': 0, 'next_day': True},
    'A':   {'fh': 6,  'fm': 0, 'th': 14, 'tm': 0, 'next_day': False},
    'B':   {'fh': 14, 'fm': 0, 'th': 22, 'tm': 0, 'next_day': False},
    'C':   {'fh': 22, 'fm': 0, 'th': 6,  'tm': 0, 'next_day': True},
}


def _jetty_v2_shift_window(report_date_str, shift):
    """Returns (from_dt, to_dt) for the given date+shift."""
    win = _JETTY_V2_SHIFT_WINDOWS.get((shift or 'ALL').upper(), _JETTY_V2_SHIFT_WINDOWS['ALL'])
    base = datetime.strptime(report_date_str, '%Y-%m-%d')
    from_dt = base.replace(hour=win['fh'], minute=win['fm'], second=0)
    to_base = base + timedelta(days=1) if win['next_day'] else base
    to_dt = to_base.replace(hour=win['th'], minute=win['tm'], second=0)
    return from_dt, to_dt


def _fetch_barges_asof_v2(to_dt, report_date_str, shift='ALL'):
    """Independent copy — computes barge status AS OF `to_dt` (end of the
    selected shift window) instead of 'right now'. Does not call or modify
    _fetch_all_barges()."""

    conn = get_db()
    cur = get_cursor(conn)

    barges = []
    cutoff_dt = datetime(2026, 5, 1)
    shift = (shift or 'ALL').upper()

    cur.execute("""
        WITH discharge_sums AS (
            SELECT
                TRIM(UPPER(ll.barge_name)) AS barge_name,
                ll.source_id,
                SUM(COALESCE(ll.quantity,0)) AS discharged_qty
            FROM lueu_lines ll
            WHERE ll.is_deleted IS NOT TRUE
              AND ll.source_type = 'VCN'
              AND ll.entry_date <= %s
            GROUP BY TRIM(UPPER(ll.barge_name)), ll.source_id
        ),
        latest_berth AS (
            SELECT DISTINCT ON (
                source_id,
                TRIM(UPPER(barge_name))
            )
                source_id,
                TRIM(UPPER(barge_name)) AS barge_key,
                berth_name
            FROM lueu_lines
            WHERE source_type = 'VCN'
              AND berth_name IS NOT NULL
              AND TRIM(berth_name) <> ''
              AND entry_date <= %s
              AND (%s = 'ALL' OR shift = %s)
            ORDER BY
                source_id,
                TRIM(UPPER(barge_name)),
                entry_date DESC,
                id DESC
        )
        SELECT
            l.id,
            l.barge_name,
            l.trip_number,
            l.cargo_name,
            lb.berth_name,
            l.anchored_gull_island,
            l.cast_off_mv,
            l.cast_off_port,
            l.along_side_berth,
            l.commence_discharge_berth,
            l.completed_discharge_berth,
            COALESCE(l.discharge_quantity,0) AS discharge_qty,
            (
                COALESCE(l.discharge_quantity,0)
                - COALESCE(ds.discharged_qty,0)
            ) AS balance_qty
        FROM ldud_barge_lines l
        LEFT JOIN ldud_header h
            ON h.id = l.ldud_id
        LEFT JOIN discharge_sums ds
            ON ds.barge_name = TRIM(
                UPPER(
                    CONCAT(
                        l.barge_name,
                        ' / ',
                        COALESCE(l.trip_number::text,'1')
                    )
                )
            )
            AND ds.source_id = h.vcn_id
        LEFT JOIN latest_berth lb
            ON lb.source_id = h.vcn_id
            AND lb.barge_key = TRIM(
                UPPER(
                    CONCAT(
                        l.barge_name,
                        ' / ',
                        COALESCE(l.trip_number::text,'1')
                    )
                )
            )
        WHERE COALESCE(TRIM(l.barge_name),'') <> ''
    """, (report_date_str, report_date_str, shift, shift))

    for row in cur.fetchall():
        row = dict(row)
        balance_qty = max(float(row.get("balance_qty") or 0), 0)

        alongside_dt   = _parse_dt(row.get("along_side_berth"))
        commence_dt    = _parse_dt(row.get("commence_discharge_berth"))
        completed_dt   = _parse_dt(row.get("completed_discharge_berth"))
        cast_off_dt    = _parse_dt(row.get("cast_off_port"))
        cast_off_mv_dt = _parse_dt(row.get("cast_off_mv"))

        if alongside_dt and alongside_dt < cutoff_dt:
            continue

        cast_off_by_then  = bool(cast_off_dt and cast_off_dt <= to_dt)
        completed_by_then = bool(completed_dt and completed_dt <= to_dt)

        status = None
        if alongside_dt and alongside_dt <= to_dt and not cast_off_by_then:
            if commence_dt and commence_dt <= to_dt and not completed_by_then:
                status = "Under Discharge"
            elif not commence_dt or commence_dt > to_dt:
                status = "Waiting"

        # ── ETA = "Loaded & Transit": cast off mother vessel by to_dt,
        # but NOT yet alongside the discharge berth as of to_dt.
        alongside_by_then = bool(alongside_dt and alongside_dt <= to_dt)
        eta_active = bool(
            cast_off_mv_dt
            and cast_off_mv_dt <= to_dt
            and not alongside_by_then
        )

        berth = (row.get("berth_name") or "").upper()

        barges.append({
            "id": row["id"],
            "type": "BARGE",
            "barge_name": row["barge_name"],
            "name": row["barge_name"],
            "cargo": row.get("cargo_name") or "",
            "balance_qty": balance_qty,
            "berth": berth,
            "status": status,
            "eta_active": eta_active,
        })

    cur.close()
    conn.close()
    return barges

def _fetch_mbc_asof_v2(to_dt, report_date_str, shift='ALL'):
    """MBC equivalent of _fetch_barges_asof_v2 — status/balance as of `to_dt`."""

    conn = get_db()
    cur = get_cursor(conn)
    cutoff_dt = datetime(2026, 5, 1)
    shift = (shift or 'ALL').upper()

    cur.execute("""
        WITH actual AS (
            SELECT source_id, SUM(COALESCE(quantity,0)) AS actual_qty
            FROM lueu_lines
            WHERE source_type = 'MBC'
              AND is_deleted IS NOT TRUE
              AND entry_date <= %s
            GROUP BY source_id
        ),
        latest_berth AS (
            SELECT DISTINCT ON (mbc_id)
                mbc_id,
                vessel_unloading_berth AS berth_name
            FROM mbc_discharge_port_lines
            WHERE vessel_unloading_berth IS NOT NULL
              AND TRIM(vessel_unloading_berth) <> ''
            ORDER BY mbc_id, id DESC
        )
        SELECT
            h.id,
            h.mbc_name,
            h.cargo_name,
            COALESCE(h.bl_quantity,0) AS bl_qty,
            p.vessel_arrival_port,
            p.unloading_commenced,
            p.unloading_completed,
            p.vessel_cast_off,
            lb.berth_name,
            COALESCE(a.actual_qty,0) AS actual_qty
        FROM mbc_header h
        JOIN mbc_discharge_port_lines p ON p.mbc_id = h.id
        LEFT JOIN actual a ON a.source_id = h.id
        LEFT JOIN latest_berth lb ON lb.mbc_id = h.id
        WHERE p.vessel_arrival_port IS NOT NULL
          AND TRIM(COALESCE(p.vessel_arrival_port,'')) <> ''
    """, (report_date_str,))

    mbcs = []
    for row in cur.fetchall():
        row = dict(row)
        arrival_dt   = _parse_dt(row.get("vessel_arrival_port"))
        commence_dt  = _parse_dt(row.get("unloading_commenced"))
        completed_dt = _parse_dt(row.get("unloading_completed"))
        cast_off_dt  = _parse_dt(row.get("vessel_cast_off"))

        if arrival_dt and arrival_dt < cutoff_dt:
            continue
        if not arrival_dt or arrival_dt > to_dt:
            continue

        cast_off_by_then  = bool(cast_off_dt and cast_off_dt <= to_dt)
        completed_by_then = bool(completed_dt and completed_dt <= to_dt)
        if cast_off_by_then or completed_by_then:
            continue  # finished before/at this point — not "active as of" to_dt

        balance_qty = max(float(row["bl_qty"] or 0) - float(row["actual_qty"] or 0), 0)
        if balance_qty <= 0:
            continue

        unloading_started = bool(commence_dt and commence_dt <= to_dt)

        mbcs.append({
            "id": row["id"],
            "type": "MBC",
            "name": row["mbc_name"],
            "cargo": row.get("cargo_name") or "",
            "balance_qty": balance_qty,
            "berth": (row.get("berth_name") or "").upper(),
            "status": "Under Discharge" if unloading_started else "Waiting",
            "unloading_commenced": str(row.get("unloading_commenced") or "").strip(),
        })

    cur.close()
    conn.close()
    return mbcs


@bp.route('/api/module/RP01/jetty-cargo-report-v2')
@login_required
def api_jetty_cargo_report_v2():
    """New, separate endpoint — date/shift-wise version.
    Does not affect /api/module/RP01/jetty-cargo-report."""
    report_date = request.args.get('date', '')
    shift = request.args.get('shift', 'ALL')
    if not report_date:
        return jsonify({'jetty_waiting': [], 'total_jetty': 0, 'total_eta': 0, 'berth_discharge': []})

    from_dt, to_dt = _jetty_v2_shift_window(report_date, shift)
    barges = _fetch_barges_asof_v2(to_dt, report_date, shift)
    mbcs   = _fetch_mbc_asof_v2(to_dt, report_date, shift)

    jetty_counts = defaultdict(int)
    eta_counts = defaultdict(int)
    berth_rows = []

    # ── Barges still feed JETTY (Waiting) and ETA (Loaded/Transit) counts ──
    for b in barges:
        cargo = (b.get('cargo') or 'UNKNOWN').strip().upper() or 'UNKNOWN'
        if b['status'] == 'Waiting':
            jetty_counts[cargo] += 1
        if b.get('eta_active'):
            eta_counts[cargo] += 1
        # NOTE: barges no longer added to berth_rows — that table is MBC-only now.

    # ── CARGO / BERTH / CARGO BAL table — MBC only ──────────────────────────
    for m in mbcs:
        if (
            m['status'] == 'Under Discharge'
            and (m.get('berth') or '').strip()
            and float(m.get('balance_qty') or 0) > 0
        ):
            berth_rows.append({
                'cargo': m.get('cargo') or '',
                'berth': (m.get('berth') or '').upper(),
                'balance': m.get('balance_qty', 0),
            })

    all_cargos = sorted(set(jetty_counts) | set(eta_counts))
    jetty_waiting = [
        {'cargo': c, 'jetty': jetty_counts.get(c, 0), 'eta': eta_counts.get(c, 0)}
        for c in all_cargos
    ]

    return jsonify({
        'jetty_waiting': jetty_waiting,
        'total_jetty': sum(jetty_counts.values()),
        'total_eta': sum(eta_counts.values()),
        'berth_discharge': berth_rows,
    })

    
@bp.route('/api/module/RP01/jetty-cargo-report')
@login_required
def api_jetty_cargo_report():
    report_date = request.args.get('date', '')
    shift = request.args.get('shift', 'ALL')
    if not report_date:
        return jsonify({'jetty_waiting': [], 'total_jetty': 0, 'total_eta': 0, 'berth_discharge': []})

    barges, _ = _fetch_all_barges()

    jetty_counts = defaultdict(int)
    for b in barges:
        if b['type'] == 'BARGE' and b['status'] == 'Waiting':
            cargo = (b.get('cargo') or 'UNKNOWN').strip().upper() or 'UNKNOWN'
            jetty_counts[cargo] += 1

    conn = get_db()
    cur = get_cursor(conn)

    cur.execute("""
        SELECT COALESCE(NULLIF(TRIM(cargo_name),''),'UNKNOWN') AS cargo, COUNT(*) AS cnt
        FROM ldud_barge_lines
        WHERE anchored_gull_island IS NOT NULL
          AND (cast_off_port IS NULL OR TRIM(cast_off_port) = '')
        GROUP BY 1
    """)
    eta_counts = {(r['cargo'] or 'UNKNOWN').upper(): r['cnt'] for r in cur.fetchall()}

    all_cargos = sorted(set(jetty_counts) | set(eta_counts))
    jetty_waiting = [
        {'cargo': c, 'jetty': jetty_counts.get(c, 0), 'eta': eta_counts.get(c, 0)}
        for c in all_cargos
    ]
    total_jetty = sum(jetty_counts.values())
    total_eta = sum(eta_counts.values())

    berth_rows = []
    for b in barges:
        if b['type'] == 'BARGE' and b['status'] == 'Under Discharge':
            berth = (b.get('berth') or b.get('commence_discharge_berth') or '').upper()
            berth_rows.append({
                'cargo': b.get('cargo') or '',
                'berth': berth,
                'balance': b.get('balance_qty', 0),
            })

    cur.execute("""
        SELECT berth_layout FROM barge_position_report
        WHERE report_date = %s AND shift = %s
    """, (report_date, shift))
    row = cur.fetchone()
    saved_layout = (dict(row).get('berth_layout') if row else []) or []
    mbc_live = {b['name'].upper(): b for b in barges if b['type'] == 'MBC'}

    for item in saved_layout:
        if (item.get('type') or '').upper() != 'MBC':
            continue
        name = (item.get('name') or '').upper()
        live = mbc_live.get(name)
        unloading_started = bool(str(item.get('unloading_commenced') or '').strip())
        if live and unloading_started:
            berth_rows.append({
                'cargo': live.get('cargo') or item.get('cargo') or '',
                'berth': (item.get('berth') or '').upper(),
                'balance': live.get('balance_qty', item.get('balance', 0)),
            })

    cur.close()
    conn.close()

    return jsonify({
        'jetty_waiting': jetty_waiting,
        'total_jetty': total_jetty,
        'total_eta': total_eta,
        'berth_discharge': berth_rows,
    })