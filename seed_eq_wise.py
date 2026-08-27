"""Seed rp01_daily_throughput from the 'Eq Wise.xlsx' daily equipment sheet.

    python -m alembic upgrade head           # creates the table (rp01dailythr01)
    python seed_eq_wise.py "C:/Users/Shubham/Downloads/Eq Wise.xlsx"

The sheet is chronological but has a known defect: one FY's block was entered
a year late (FY2022-23 rows stamped 2023-04-01 .. 2024-02-28, colliding with
the real FY2023-24). clean_rows() repairs that by year-shifting any run whose
dates only become contiguous once a year is subtracted, then asserts the whole
series is contiguous and duplicate-free — so a *new* dating defect fails the
seed loudly instead of silently poisoning the forecast.
"""
import sys
from datetime import date, datetime, timedelta

from database import get_db, get_cursor

SHEET = 'Yearly'
DATE_COL, TOTAL_COL = 0, 12


def _shift_year(d, years):
    try:
        return d.replace(year=d.year - years)
    except ValueError:          # Feb 29 -> Feb 28
        return d.replace(year=d.year - years, day=28)


def clean_rows(raw):
    """[(date, total, {equipment: qty})] in sheet order -> same, dates repaired.

    Walks in sheet order carrying a year-offset. When the next date breaks
    day-by-day contiguity but does line up once shifted back a year, the offset
    flips on (and back off again when the unshifted date is the contiguous one).
    """
    out, offset, prev = [], 0, None
    for d, total, eq in raw:
        cand = _shift_year(d, offset)
        if prev is not None and cand != prev + timedelta(days=1):
            alt = _shift_year(d, 1 - offset)
            if alt == prev + timedelta(days=1):
                offset = 1 - offset
                cand = alt
        out.append((cand, total, eq))
        prev = cand

    dates = [r[0] for r in out]
    dupes = len(dates) - len(set(dates))
    assert not dupes, f"{dupes} duplicate dates survived cleaning"
    span = (dates[-1] - dates[0]).days + 1
    assert span == len(dates), f"series has {span - len(dates)} missing day(s)"
    return out


def read_sheet(path):
    """-> (equipment column names, [(date, total, {equipment: qty})]) in sheet order."""
    import openpyxl
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True)[SHEET]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header = next(r for r in rows if str(r[DATE_COL]).strip() == 'Date')
    names = [str(h).strip() for h in header[1:TOTAL_COL]]

    raw = []
    for r in rows:
        d = r[DATE_COL]
        if not isinstance(d, (datetime, date)):
            continue
        d = d.date() if isinstance(d, datetime) else d
        eq = {n: float(v) for n, v in zip(names, r[1:TOTAL_COL])
              if isinstance(v, (int, float))}
        raw.append((d, float(r[TOTAL_COL] or 0), eq))
    return names, raw


def fix_excel(src, dst):
    """Write a date-repaired, chronologically sorted copy of the sheet.

    Same year-shift repair the seeder applies, plus TOTAL recomputed from the
    equipment columns (the source has a couple of arithmetic typos). Reports
    every change so nothing is corrected silently.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    names, raw = read_sheet(src)
    before = {i: d for i, (d, _, _) in enumerate(raw)}
    rows = clean_rows(raw)

    moved = [(before[i], d) for i, (d, _, _) in enumerate(rows) if before[i] != d]
    retotalled = []
    out = []
    for d, total, eq in rows:
        s = round(sum(eq.values()), 3)
        if abs(s - total) > 1:
            retotalled.append((d, total, s))
        out.append((d, s, eq))
    out.sort(key=lambda r: r[0])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    hdr = ['Date'] + names + ['TOTAL']
    for ci, h in enumerate(hdr, start=1):
        c = ws.cell(1, ci, h)
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='DDEBF7')
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[c.column_letter].width = max(12, len(h) + 3)
    ws.freeze_panes = 'A2'

    for ri, (d, total, eq) in enumerate(out, start=2):
        ws.cell(ri, 1, d).number_format = 'DD-MM-YYYY'
        for ci, n in enumerate(names, start=2):
            v = eq.get(n)
            if v is not None:
                ws.cell(ri, ci, v).number_format = '#,##0'
        ws.cell(ri, len(hdr), total).number_format = '#,##0'
    wb.save(dst)

    print(f"wrote {dst}")
    print(f"  {len(out)} rows, {out[0][0]} -> {out[-1][0]} (sorted, contiguous, no duplicates)")
    print(f"  re-dated {len(moved)} rows (the year-late block)")
    if moved:
        print(f"    {moved[0][0]} -> {moved[0][1]}  ...  {moved[-1][0]} -> {moved[-1][1]}")
    print(f"  recomputed {len(retotalled)} TOTAL cells")
    for d, was, now in retotalled:
        print(f"    {d}  {was:,.0f} -> {now:,.0f}")
    return out


def require_table(cur):
    """The table is owned by alembic revision rp01dailythr01, not by this script.

    A data-loading tool has no business creating schema — that is how a table
    ends up on a database nobody migrated. Fail with the fix instead.
    """
    cur.execute("SELECT to_regclass('rp01_daily_throughput') AS t")
    if not cur.fetchone()['t']:
        raise SystemExit(
            "rp01_daily_throughput does not exist.\n"
            "Run the migration first:  python -m alembic upgrade head")


def seed(path):
    import json
    _, raw = read_sheet(path)
    rows = clean_rows(raw)
    conn = get_db()
    cur = get_cursor(conn)
    try:
        require_table(cur)
        for d, total, eq in rows:
            cur.execute("""
                INSERT INTO rp01_daily_throughput (entry_date, equipment, total)
                VALUES (%s, %s, %s)
                ON CONFLICT (entry_date) DO UPDATE
                   SET equipment = EXCLUDED.equipment,
                       total     = EXCLUDED.total,
                       updated_at = NOW()
            """, (d, json.dumps(eq), total))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return rows


def check_overlap():
    """Compare seeded Excel totals against lueu_lines on the dates they share.

    The forecast splices Excel history onto live lueu_lines days. If the two
    disagree about the same day, the level estimate jumps at the seam, so this
    is worth eyeballing on prod before trusting the numbers.
    """
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("""
            SELECT t.entry_date, t.total AS excel, l.q AS lueu
            FROM rp01_daily_throughput t
            JOIN (
                SELECT entry_date::date AS d, SUM(COALESCE(quantity, 0)) AS q
                FROM lueu_lines WHERE is_deleted IS NOT TRUE GROUP BY 1
            ) l ON l.d = t.entry_date
            ORDER BY t.entry_date
        """)
        rows = cur.fetchall()
    finally:
        conn.close()

    if not rows:
        print('No overlapping dates — nothing to compare.')
        return

    diffs = [(r['entry_date'], float(r['excel']), float(r['lueu'])) for r in rows]
    worst = sorted(diffs, key=lambda x: -abs(x[1] - x[2]))[:10]
    tot_e = sum(d[1] for d in diffs)
    tot_l = sum(d[2] for d in diffs)
    print(f"{len(diffs)} overlapping days: {diffs[0][0]} -> {diffs[-1][0]}")
    print(f"  Excel total {tot_e:,.0f} MT   lueu_lines total {tot_l:,.0f} MT"
          f"   delta {(tot_l - tot_e) / tot_e * 100 if tot_e else 0:+.2f}%")
    print("  largest daily gaps:")
    for d, e, l in worst:
        print(f"    {d}  excel {e:>10,.0f}  lueu {l:>10,.0f}  diff {l - e:>+10,.0f}")


def demo():
    """Self-check: the repair fixes a year-late block and rejects a real gap."""
    mk = lambda y, m, dd: (date(y, m, dd), 100.0, {})
    good = [mk(2022, 3, 30), mk(2022, 3, 31),
            mk(2023, 4, 1), mk(2023, 4, 2),          # year-late block
            mk(2022, 4, 3)]                          # back to correct stamping
    assert [r[0] for r in clean_rows(good)] == [
        date(2022, 3, 30), date(2022, 3, 31),
        date(2022, 4, 1), date(2022, 4, 2), date(2022, 4, 3)]

    try:
        clean_rows([mk(2022, 3, 30), mk(2022, 4, 5)])
    except AssertionError:
        pass
    else:
        raise AssertionError("a genuine gap should have failed the assert")
    print("clean_rows self-check ok")


if __name__ == '__main__':
    arg = sys.argv[1] if len(sys.argv) > 1 else None
    if arg is None:
        demo()
    elif arg == '--check':
        check_overlap()
    elif arg == '--fix-excel':
        if len(sys.argv) < 4:
            sys.exit('usage: seed_eq_wise.py --fix-excel <in.xlsx> <out.xlsx>')
        fix_excel(sys.argv[2], sys.argv[3])
    else:
        rows = seed(arg)
        print(f"seeded {len(rows)} days: {rows[0][0]} -> {rows[-1][0]}")
        check_overlap()
